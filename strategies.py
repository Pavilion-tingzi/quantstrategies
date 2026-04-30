import re
import numpy as np
import pandas as pd
import os
from pathlib import Path
from typing import Optional, List, Dict, Tuple, Union
import matplotlib.dates as mdates
import matplotlib.pyplot as plt
import matplotlib
import warnings
import chardet
from tools import TechnicalIndicators, RiskManage, BaseTool, MyCommInfo, MyRMSizer, MyAllInSizer, MyBroker, SignalEffectiveness, MyObserver
import backtrader as bt
import backtrader.analyzers as btanalyzers
import math

# 设置matplotlib中文字体
matplotlib.rcParams['font.sans-serif'] = ['SimHei', 'Arial Unicode MS', 'DejaVu Sans']
matplotlib.rcParams['axes.unicode_minus'] = False

class BuyHoldStrategy:
    """买入持有策略类，用于计算买入持有策略的各项绩效指标"""

    def __init__(self, initial_capital: float, commission_rate: float,
                 stamp_duty_rate: float, min_commission: float):
        """
        初始化买入持有策略

        Parameters:
        -----------
        initial_capital : float
            初始资金
        commission_rate : float
            佣金费率
        stamp_duty_rate : float
            印花税费率
        min_commission : float
            最低佣金
        """
        self.initial_capital = initial_capital
        self.commission_rate = commission_rate
        self.stamp_duty_rate = stamp_duty_rate
        self.min_commission = min_commission
        self.data = None
        self.net_values = None
        self.metrics = {}

    def run_backtest(self, data: pd.DataFrame) -> pd.Series:
        """
        运行买入持有策略回测

        Parameters:
        -----------
        data : pd.DataFrame
            包含日期和收盘价的数据

        Returns:
        --------
        pd.Series : 每日净值序列
        """
        self.data = data.copy()
        first_close = float(self.data.iloc[0]['收盘'])

        # 计算可买入的最大股数
        max_shares = int(self.initial_capital / (first_close * 100)) * 100
        shares = max_shares

        if shares > 0:
            buy_value = first_close * shares
            commission = max(buy_value * self.commission_rate, self.min_commission)
            cash_after_buy = self.initial_capital - buy_value - commission
        else:
            cash_after_buy = self.initial_capital
            shares = 0

        # 计算每日净值
        net_values = []
        for i, row in self.data.iterrows():
            if i == 0:
                net_values.append(self.initial_capital)
            else:
                total_value = cash_after_buy + shares * float(row['收盘'])
                net_values.append(total_value)

        self.net_values = pd.Series(net_values, index=pd.to_datetime(self.data['日期']))
        return self.net_values

    def calculate_metrics(self, risk_free_rate: float = 0.02) -> Dict:
        """
        计算买入持有策略的绩效指标

        Parameters:
        -----------
        risk_free_rate : float
            无风险利率

        Returns:
        --------
        Dict : 绩效指标字典
        """
        if self.net_values is None:
            raise ValueError("请先运行回测")

        # 总收益率
        total_return = (self.net_values.iloc[-1] / self.net_values.iloc[0]) - 1

        # 年化收益率
        days = (self.net_values.index[-1] - self.net_values.index[0]).days
        if days > 0:
            annual_return = (1 + total_return) ** (252 / days) - 1
        else:
            annual_return = 0

        # 最大回撤
        rolling_max = self.net_values.expanding().max()
        drawdown = (self.net_values - rolling_max) / rolling_max
        max_drawdown = drawdown.min()

        # 夏普比率
        daily_returns = self.net_values.pct_change().dropna()
        if len(daily_returns) > 0 and daily_returns.std() > 0:
            excess = daily_returns - risk_free_rate / 252
            sharpe = excess.mean() / daily_returns.std() * np.sqrt(252)
        else:
            sharpe = 0

        self.metrics = {
            '总收益率': total_return,
            '年化收益率': annual_return,
            '最大回撤': max_drawdown,
            '夏普比率': sharpe,
            '交易次数': 0,
            '胜率': 0,
            '盈亏比': 0
        }

        return self.metrics

class DoubleMovingAverageStrategy:
    """
    双均线策略回测类

    参数可自定义：
    - short_ma: 短期均线周期
    - long_ma: 长期均线周期
    - initial_capital: 初始资金
    - commission_rate: 手续费率
    - min_commission: 最低手续费
    - stamp_tax_rate: 印花税率 (仅卖出)
    - stop_loss: 止损线
    - risk_free_rate: 无风险利率
    """

    def __init__(self,
                 short_ma: int = 10,
                 long_ma: int = 120,
                 initial_capital: float = 100000,
                 commission_rate: float = 0.0001,
                 min_commission: float = 5,
                 stamp_tax_rate: float = 0.001,
                 stop_loss: float = None,
                 risk_free_rate: float = 0.02):
        """
        初始化策略参数

        Parameters:
        -----------
        short_ma : int
            短期均线周期 (默认: 10)
        long_ma : int
            长期均线周期 (默认: 120)
        initial_capital : float
            初始资金 (默认: 100000)
        commission_rate : float
            手续费率 (默认: 0.0001, 万1)
        min_commission : float
            最低手续费 (默认: 5)
        stamp_tax_rate : float
            印花税率 (默认: 0.001, 千1, 仅卖出)
        stop_loss : float, optional
            止损线，如0.1表示-10%止损 (默认: None)
        risk_free_rate : float
            无风险利率 (默认: 0.02, 用于夏普比率计算)
        """
        # 策略参数
        self.short_ma = short_ma
        self.long_ma = long_ma
        self.initial_capital = initial_capital
        self.commission_rate = commission_rate
        self.min_commission = min_commission
        self.stamp_tax_rate = stamp_tax_rate
        self.stop_loss = stop_loss
        self.risk_free_rate = risk_free_rate

        # 验证参数有效性
        if short_ma >= long_ma:
            warnings.warn("警告：短期均线周期应小于长期均线周期，否则可能导致信号异常")

        # 数据相关
        self.stock_code = None
        self.stock_name = None
        self.data = None
        self.results = None

        # 绩效指标
        self.metrics = {}
        self.buyhold_metrics = {}

        # 均线列名
        self.ma_short_col = f'ma_{self.short_ma}'
        self.ma_long_col = f'ma_{self.long_ma}'

    def _detect_encoding(self, filepath: str) -> str:
        """检测文件编码"""
        encodings = ['gbk', 'utf-8', 'gb2312', 'gb18030']
        for encoding in encodings:
            try:
                with open(filepath, 'r', encoding=encoding) as f:
                    f.readline()
                return encoding
            except:
                continue
        return 'gbk'  # 默认返回gbk

    def load_data(self, filepath: str) -> 'DoubleMovingAverageStrategy':
        """
        加载数据

        Parameters:
        -----------
        filepath : str
            数据文件路径

        Returns:
        --------
        self : DoubleMovingAverageStrategy
            策略实例
        """
        # 检测文件编码
        encoding = self._detect_encoding(filepath)

        # 加载数据
        self.data = pd.read_csv(filepath, encoding=encoding)

        # 提取股票代码和名称
        filename = os.path.basename(filepath)

        # 尝试从文件名提取股票代码（6位数字）
        match = re.search(r'(\d{6})', filename)
        if match:
            self.stock_code = match.group(1)
        else:
            self.stock_code = filename.split('.')[0]

        # 尝试提取股票名称
        name_match = re.search(r'[^a-zA-Z0-9](\w{2,4})[^a-zA-Z0-9]', filename)
        self.stock_name = name_match.group(1) if name_match else self.stock_code

        # 确保日期列存在并处理
        if '日期' not in self.data.columns:
            # 尝试其他可能的日期列名
            for col in self.data.columns:
                if 'date' in col.lower() or '时间' in col or '交易日期' in col:
                    self.data.rename(columns={col: '日期'}, inplace=True)
                    break
            else:
                raise ValueError("数据中缺少'日期'列")

        # 转换日期格式
        try:
            self.data['日期'] = pd.to_datetime(self.data['日期'])
        except:
            try:
                self.data['日期'] = pd.to_datetime(self.data['日期'], format='%Y%m%d')
            except:
                pass

        # 按日期排序
        self.data = self.data.sort_values('日期').reset_index(drop=True)

        print(f"股票 {self.stock_code} {self.stock_name} 数据加载完成，共{len(self.data)}条记录")
        return self

    def preprocess_data(self) -> 'DoubleMovingAverageStrategy':
        """
        数据处理：计算策略指标和交易信号字段（正确处理停牌和涨跌停）
        """
        if self.data is None:
            raise ValueError("请先调用load_data加载数据")

        # 确保必要的列存在
        required_cols = ['开盘', '收盘', '最高', '最低']
        for col in required_cols:
            if col not in self.data.columns:
                for possible in [col, col.lower(), col.capitalize()]:
                    if possible in self.data.columns:
                        self.data.rename(columns={possible: col}, inplace=True)
                        break
                else:
                    warnings.warn(f"数据中缺少'{col}'列")

        # 创建停牌标识
        if '异常情况' in self.data.columns:
            self.data['is_suspend'] = (self.data['异常情况'] == '停牌').astype(int)
        else:
            self.data['is_suspend'] = 0
            warnings.warn("数据中缺少'异常情况'字段，将假设所有交易日正常")

        # 创建异常涨跌幅标识
        if '异常涨跌幅' in self.data.columns:
            self.data['abnormal_buy_forbidden'] = (self.data['异常涨跌幅'] > 0).astype(int)
            self.data['abnormal_sell_forbidden'] = (self.data['异常涨跌幅'] < 0).astype(int)
        else:
            self.data['abnormal_buy_forbidden'] = 0
            self.data['abnormal_sell_forbidden'] = 0

        # 初始化交易相关字段
        self.data['position'] = 0
        self.data['shares'] = 0
        self.data['cash'] = 0.0
        self.data['hold_value'] = 0.0
        self.data['commission'] = 0.0
        self.data['stamp_tax'] = 0.0
        self.data['total_asset'] = 0.0
        self.data['daily_return'] = 0.0
        self.data['signal'] = ''
        self.data['raw_signal'] = ''
        self.data['signal_executed'] = ''

        # 使用TechnicalIndicators计算均线（传入suspend_col排除停牌）

        self.data[self.ma_short_col] = TechnicalIndicators.calculate_moving_average(
            self.data, self.short_ma, close_col='收盘', shift=1, suspend_col='is_suspend'
        )
        self.data[self.ma_long_col] = TechnicalIndicators.calculate_moving_average(
            self.data, self.long_ma, close_col='收盘', shift=1, suspend_col='is_suspend'
        )

        # 识别金叉和死叉
        golden_cross, death_cross = TechnicalIndicators.detect_golden_death_cross(
            self.data[self.ma_short_col],
            self.data[self.ma_long_col]
        )
        self.data['golden_cross'] = golden_cross
        self.data['death_cross'] = death_cross

        # 生成原始信号
        for i in range(len(self.data)):
            if self.data.loc[i, 'golden_cross']:
                self.data.loc[i, 'raw_signal'] = 'buy'
            elif self.data.loc[i, 'death_cross']:
                self.data.loc[i, 'raw_signal'] = 'sell'

        # 处理异常情况导致的交易延迟
        self._process_abnormal_signals()

        # 计算ATR指标（排除停牌日）
        if '最高' in self.data.columns and '最低' in self.data.columns:
            self.data['atr'] = TechnicalIndicators.calculate_atr(
                self.data, period=14,
                high_col='最高', low_col='最低', close_col='收盘',
                suspend_col='is_suspend'
            )
            print(f"ATR计算完成，有效ATR值数量: {(self.data['atr'] > 0).sum()}")

        # 添加shift后的收盘价
        self.data['收盘_shift1'] = self.data['收盘'].shift(1)

        print("数据处理完成，指标计算完毕")

        return self

    def _process_abnormal_signals(self) -> None:
        """
        处理异常情况导致的交易延迟
        包括：停牌、涨停（不能买入）、跌停（不能卖出）
        """
        # 先过滤掉停牌日的信号
        suspend_mask = self.data['is_suspend'] == 1
        self.data.loc[suspend_mask, 'raw_signal'] = ''

        # 处理涨跌停限制
        # 涨停日不能买入
        buy_forbidden_mask = self.data['abnormal_buy_forbidden'] == 1
        self.data.loc[buy_forbidden_mask & (self.data['raw_signal'] == 'buy'), 'raw_signal'] = ''

        # 跌停日不能卖出
        sell_forbidden_mask = self.data['abnormal_sell_forbidden'] == 1
        self.data.loc[sell_forbidden_mask & (self.data['raw_signal'] == 'sell'), 'raw_signal'] = ''

        # 使用TechnicalIndicators生成最终信号（买卖交替、最小间隔）
        self.data['signal'] = TechnicalIndicators.filter_signals_with_interval(
            self.data, signal_col='raw_signal', min_interval=5, first_must_buy=True
        )

        # 处理延迟执行信号（当信号遇到异常情况时延迟到下一个交易日）
        self.data['pending_signal'] = ''
        pending = None

        for i in range(len(self.data)):
            is_abnormal = (self.data.loc[i, 'is_suspend'] == 1 or
                           self.data.loc[i, 'abnormal_buy_forbidden'] == 1 or
                           self.data.loc[i, 'abnormal_sell_forbidden'] == 1)

            current_signal = self.data.loc[i, 'signal']

            if pending is not None and not is_abnormal:
                # 如果有待执行的信号且当前不是异常日，执行延迟信号
                self.data.loc[i, 'signal'] = pending
                pending = None
            elif current_signal != '' and is_abnormal:
                # 如果当前有信号但遇到异常日，延迟执行
                pending = current_signal
                self.data.loc[i, 'signal'] = ''

        # 统计信号
        buy_signals = (self.data['signal'] == 'buy').sum()
        sell_signals = (self.data['signal'] == 'sell').sum()

        print(f"信号生成完成，买入信号: {buy_signals}次，卖出信号: {sell_signals}次")

    def _generate_final_signals(self) -> None:
        """
        生成最终交易信号，满足以下条件：
        1. 首笔交易必须是买入
        2. 买卖信号必须交替出现
        3. 相同方向信号之间至少间隔5天
        """
        # 设置最小信号间隔
        min_interval = 5

        # 获取所有原始信号
        signal_indices = []
        signal_types = []

        for i in range(len(self.data)):
            if self.data.loc[i, 'raw_signal'] in ['buy', 'sell']:
                signal_indices.append(i)
                signal_types.append(self.data.loc[i, 'raw_signal'])

        if not signal_indices:
            return

        # 找到第一个买入信号的位置
        first_buy_idx = None
        for i, signal_type in enumerate(signal_types):
            if signal_type == 'buy':
                first_buy_idx = i
                break

        if first_buy_idx is None:
            return

        # 从第一个买入信号开始筛选
        selected_indices = [signal_indices[first_buy_idx]]
        selected_types = ['buy']
        last_idx = signal_indices[first_buy_idx]
        last_type = 'buy'

        # 遍历后续信号
        for i in range(first_buy_idx + 1, len(signal_indices)):
            current_idx = signal_indices[i]
            current_type = signal_types[i]

            # 检查信号类型是否交替（不能连续相同）
            if current_type == last_type:
                continue

            # 检查时间间隔
            if current_idx - last_idx < min_interval:
                continue

            # 通过所有检查，接受该信号
            selected_indices.append(current_idx)
            selected_types.append(current_type)
            last_idx = current_idx
            last_type = current_type

        # 先清空原有的signal列
        self.data['signal'] = ''

        # 将筛选后的信号写入signal列
        for idx, signal_type in zip(selected_indices, selected_types):
            self.data.loc[idx, 'signal'] = signal_type

    def _check_stop_loss(self, current_price: float, buy_price: float) -> bool:
        """检查是否触发止损"""
        if self.stop_loss is None or buy_price == 0:
            return False
        return (current_price - buy_price) / buy_price <= -self.stop_loss

    def run_backtest(self, risk_manager=None) -> 'DoubleMovingAverageStrategy':
        """
        运行回测
        """
        if self.data is None:
            raise ValueError("请先加载和预处理数据")

        cash = self.initial_capital
        shares = 0
        avg_cost = 0.0
        pending = None
        last_trade_date = None

        # 初始化数据列
        self.data['shares'] = 0
        self.data['avg_cost'] = 0.0
        self.data['cash'] = 0.0
        self.data['hold_value'] = 0.0
        self.data['total_asset'] = 0.0
        self.data['daily_return'] = 0.0
        self.data['signal_executed'] = ''
        self.data['exit_reason'] = ''

        # 风险管理相关字段
        if risk_manager is not None:
            self.data['risk_status'] = 0
            self.data['add_count'] = 0
            self.data['stop_loss_price'] = 0.0
            self.data['stop_profit_price'] = 0.0
            self.data['last_buy_price'] = 0.0
            self.data['initial_atr'] = 0.0

        # 获取异常情况标记
        is_suspend = self.data['is_suspend'] if 'is_suspend' in self.data.columns else pd.Series(0,
                                                                                                 index=self.data.index)
        abnormal_buy_forbidden = self.data[
            'abnormal_buy_forbidden'] if 'abnormal_buy_forbidden' in self.data.columns else pd.Series(0,
                                                                                                      index=self.data.index)
        abnormal_sell_forbidden = self.data[
            'abnormal_sell_forbidden'] if 'abnormal_sell_forbidden' in self.data.columns else pd.Series(0,
                                                                                                        index=self.data.index)

        # 检查ATR是否存在
        if risk_manager is not None and 'atr' not in self.data.columns:
            print("警告: ATR列不存在，将使用原有逻辑（无风险管理）")
            risk_manager = None

        for i in range(len(self.data)):
            row = self.data.iloc[i]
            signal = ''

            # 判断是否为异常日（不能交易）
            is_abnormal = (is_suspend.iloc[i] == 1 or
                           abnormal_buy_forbidden.iloc[i] == 1 or
                           abnormal_sell_forbidden.iloc[i] == 1)

            # 风险管理：检查止损止盈
            if risk_manager is not None and shares > 0:
                prev_close = self.data.iloc[i - 1]['收盘'] if i > 0 else None
                prev_atr = self.data.iloc[i-1]['atr'] if i > 0 and 'atr' in self.data.columns else None

                should_exit, exit_reason = risk_manager.check_stop(
                    prev_close=prev_close,
                    current_atr=prev_atr  # 注意：传入前一日ATR
                )

                if should_exit and not is_abnormal:
                    signal = 'sell'
                    self.data.at[i, 'exit_reason'] = exit_reason

                    # 执行卖出
                    price = row['收盘']
                    trade_amount = shares * price
                    commission = max(trade_amount * self.commission_rate, self.min_commission)
                    stamp_tax = trade_amount * self.stamp_tax_rate
                    cash += trade_amount - commission - stamp_tax
                    self.data.at[i, 'commission'] = commission
                    self.data.at[i, 'stamp_tax'] = stamp_tax

                    # 记录卖出前的状态
                    state_before = risk_manager.get_risk_state()
                    self.data.at[i, 'stop_loss_price'] = state_before['stop_loss']
                    self.data.at[i, 'stop_profit_price'] = state_before['stop_profit']

                    self.data.at[i, 'signal_executed'] = 'sell'

                    shares = 0
                    avg_cost = 0.0
                    pending = None
                    risk_manager.update_risk_state('close')
                    last_trade_date = row['日期']


            # 记录当前风险状态（每行都记录）
            if risk_manager is not None:
                state = risk_manager.get_risk_state()
                self.data.at[i, 'risk_status'] = state['position_status']
                self.data.at[i, 'add_count'] = state['add_count']
                self.data.at[i, 'avg_cost'] = state['avg_cost']
                self.data.at[i, 'stop_loss_price'] = state['stop_loss']
                self.data.at[i, 'stop_profit_price'] = state['stop_profit']
                self.data.at[i, 'last_buy_price'] = state['last_buy_price']
                self.data.at[i, 'initial_atr'] = state['initial_atr']
                # 打印详细信息
                print(f"日期={row['日期']},"
                      f"执行信号={signal}, 止盈价={state['stop_profit']:.6f}, "
                      f"最高收盘价={state['highest_close']:.6f}")

            # 执行待处理的延迟信号
            if pending is not None and not is_abnormal:
                if pending == 'buy' and signal == '':
                    signal = 'buy'
                    pending = None
                elif pending == 'sell' and signal == '':
                    signal = 'sell'
                    pending = None
                    if risk_manager:
                        risk_manager.update_risk_state('close')

            # 触发新信号 - 关键修改
            final_signal = row.get('signal', '')

            if signal == '' and not is_abnormal:
                # 加仓判断：只要持仓中且价格达到加仓条件就加仓
                # 不依赖买入信号，因为加仓是基于价格条件而非金叉信号
                if risk_manager is not None and shares > 0:
                    prev_close = self.data.iloc[i - 1]['收盘'] if i > 0 else None
                    can_add = risk_manager.is_can_add(prev_close)

                    if can_add:
                        signal = 'add'
                        # 可选：打印加仓触发日志
                        # print(f"📊 触发加仓条件: 日期={row['日期']}, 价格={row['收盘']:.4f}")

                # 建仓判断（无持仓且信号为买入）
                elif shares == 0 and final_signal == 'buy':
                    signal = 'buy'

                # 平仓判断（有持仓且信号为卖出）
                elif shares > 0 and final_signal == 'sell':
                    signal = 'sell'
                    if risk_manager:
                        risk_manager.update_risk_state('close')

            # 延迟执行（遇到异常日）
            elif signal == '' and is_abnormal and final_signal != '':
                if shares == 0 and final_signal == 'buy':
                    pending = 'buy'
                elif shares > 0 and final_signal == 'sell':
                    pending = 'sell'

            # 执行建仓
            if signal == 'buy':
                price = row['收盘']
                atr_value = row['atr'] if 'atr' in self.data.columns else None

                if risk_manager is not None and atr_value is not None and atr_value > 0:
                    pos_info = risk_manager.calculate_position_size(price, atr_value, cash)
                    shares_to_buy = pos_info['shares']

                    if shares_to_buy > 0:
                        trade_amount = shares_to_buy * price
                        commission = max(trade_amount * self.commission_rate, self.min_commission)
                        total_cost = trade_amount + commission

                        if cash >= total_cost:
                            shares = shares_to_buy
                            avg_cost = price
                            cash -= total_cost
                            self.data.at[i, 'commission'] = commission
                            self.data.at[i, 'stamp_tax'] = 0
                            self.data.at[i, 'signal_executed'] = 'buy'

                            risk_manager.update_risk_state(
                                'open',
                                price=price,
                                stop_loss=pos_info['stop_loss'],
                                stop_profit=pos_info['stop_profit'],
                                atr_value=atr_value,
                                add_shares_list=pos_info['add_shares_list']
                            )

                            updated_state = risk_manager.get_risk_state()
                            self.data.at[i, 'risk_status'] = updated_state['position_status']
                            self.data.at[i, 'add_count'] = updated_state['add_count']
                            self.data.at[i, 'avg_cost'] = updated_state['avg_cost']
                            self.data.at[i, 'stop_loss_price'] = updated_state['stop_loss']
                            self.data.at[i, 'stop_profit_price'] = updated_state['stop_profit']
                            self.data.at[i, 'last_buy_price'] = updated_state['last_buy_price']

                            print(f"🏠 建仓成功: 日期={row['日期']}, 价格={price:.4f}, 股数={shares_to_buy}, "
                                  f"止损价={updated_state['stop_loss']:.4f}")
                else:
                    # 原有买入逻辑
                    max_shares = int(cash / price / 100) * 100
                    while max_shares > 0:
                        trade_amount = max_shares * price
                        commission = max(trade_amount * self.commission_rate, self.min_commission)
                        total_cost = trade_amount + commission

                        if cash >= total_cost:
                            shares = max_shares
                            avg_cost = price
                            cash -= total_cost
                            self.data.at[i, 'commission'] = commission
                            self.data.at[i, 'stamp_tax'] = 0
                            self.data.at[i, 'signal_executed'] = 'buy'
                            break
                        else:
                            max_shares -= 100

            # 执行加仓
            elif signal == 'add' and risk_manager is not None and shares > 0:
                price = row['收盘']
                old_shares = shares

                add_info = risk_manager.calculate_add_position_size(price, old_shares, cash)
                shares_to_add = add_info['shares']

                if shares_to_add > 0:
                    trade_amount = shares_to_add * price
                    commission = max(trade_amount * self.commission_rate, self.min_commission)
                    total_cost = trade_amount + commission

                    if cash >= total_cost:
                        old_avg_cost = avg_cost
                        new_avg_cost = (old_avg_cost * old_shares + price * shares_to_add) / (
                                    old_shares + shares_to_add)
                        shares += shares_to_add
                        avg_cost = new_avg_cost
                        cash -= total_cost
                        self.data.at[i, 'commission'] = commission
                        self.data.at[i, 'stamp_tax'] = 0
                        self.data.at[i, 'signal_executed'] = 'add'

                        risk_manager.update_risk_state(
                            'add',
                            old_shares=old_shares,
                            new_shares=shares_to_add,
                            price=price,
                            new_stop_loss=add_info.get('new_stop_loss')
                        )

                        updated_state = risk_manager.get_risk_state()
                        self.data.at[i, 'risk_status'] = updated_state['position_status']
                        self.data.at[i, 'add_count'] = updated_state['add_count']
                        self.data.at[i, 'avg_cost'] = updated_state['avg_cost']
                        self.data.at[i, 'stop_loss_price'] = updated_state['stop_loss']
                        self.data.at[i, 'stop_profit_price'] = updated_state['stop_profit']
                        self.data.at[i, 'last_buy_price'] = updated_state['last_buy_price']

                        print(f"➕ 加仓成功: 日期={row['日期']}, 价格={price:.4f}, "
                              f"加仓股数={shares_to_add}, 总持仓={shares}, "
                              f"新成本={updated_state['avg_cost']:.4f}, "
                              f"新止损价={updated_state['stop_loss']:.4f},"
                              f"止盈价={updated_state['stop_profit']:.6f}")

            # 执行卖出
            elif signal == 'sell' and shares > 0:
                price = row['收盘']
                trade_amount = shares * price
                commission = max(trade_amount * self.commission_rate, self.min_commission)
                stamp_tax = trade_amount * self.stamp_tax_rate
                cash += trade_amount - commission - stamp_tax

                self.data.at[i, 'commission'] = commission
                self.data.at[i, 'stamp_tax'] = stamp_tax
                self.data.at[i, 'signal_executed'] = 'sell'

                if risk_manager:
                    state_before = risk_manager.get_risk_state()
                    print(f"💰 卖出: 日期={row['日期']}, 价格={price:.4f}, 股数={shares}, "
                          f"成本={state_before['avg_cost']:.4f}, 止损价={state_before['stop_loss']:.4f}")

                shares = 0
                avg_cost = 0.0

                if risk_manager:
                    risk_manager.update_risk_state('close')

            # 记录每日状态
            self.data.at[i, 'shares'] = shares
            self.data.at[i, 'position'] = 1 if shares > 0 else 0
            self.data.at[i, 'avg_cost'] = avg_cost
            self.data.at[i, 'cash'] = cash
            self.data.at[i, 'hold_value'] = shares * row['收盘']
            self.data.at[i, 'total_asset'] = cash + shares * row['收盘']

        # 计算日收益率
        self.data['daily_return'] = self.data['total_asset'].pct_change()
        self.data['cumulative_returns'] = self.data['total_asset'] / self.initial_capital - 1

        # 计算回撤
        self.data['cum_max'] = self.data['total_asset'].cummax()
        mask = self.data['cum_max'] > 0
        self.data['drawdown'] = 0.0
        self.data.loc[mask, 'drawdown'] = 1 - self.data.loc[mask, 'total_asset'] / self.data.loc[mask, 'cum_max']

        # 统计
        executed_buy = (self.data['signal_executed'] == 'buy').sum()
        executed_sell = (self.data['signal_executed'] == 'sell').sum()
        executed_add = (self.data['signal_executed'] == 'add').sum()

        print(f"\n回测执行完成，买入: {executed_buy}次，加仓: {executed_add}次，卖出: {executed_sell}次")

        # 检查风险状态
        if risk_manager is not None:
            stop_loss_records = self.data[self.data['stop_loss_price'] > 0]
            print(f"\n=== 风险管理状态检查 ===")
            print(f"止损价>0的记录数: {len(stop_loss_records)}")

        return self

    def calculate_metrics(self) -> Dict:
        """
        计算绩效指标

        Returns:
        --------
        Dict : 策略绩效指标字典
        """
        if self.data is None:
            raise ValueError("请先运行回测")

        try:
            # 最终资产
            final_asset = self.data['total_asset'].iloc[-1]

            # 基础指标
            trading_days = len(self.data)
            self.metrics['总收益率'] = final_asset / self.initial_capital - 1

            # 年化收益率
            if self.metrics['总收益率'] > -1:
                try:
                    self.metrics['年化收益率'] = (1 + self.metrics['总收益率']) ** (252 / trading_days) - 1
                except (ValueError, ZeroDivisionError):
                    self.metrics['年化收益率'] = 0
            else:
                self.metrics['年化收益率'] = -1

            # 最大回撤
            self.data['cum_max'] = self.data['total_asset'].cummax()
            mask = self.data['cum_max'] > 0
            self.data['drawdown'] = 0.0
            self.data.loc[mask, 'drawdown'] = 1 - self.data.loc[mask, 'total_asset'] / self.data.loc[mask, 'cum_max']
            self.metrics['最大回撤率'] = self.data['drawdown'].max()

            # 夏普比率
            daily_returns = self.data['daily_return'].dropna()
            if len(daily_returns) > 0:
                daily_std = daily_returns.std()
                if daily_std > 0 and not np.isnan(daily_std) and not np.isinf(daily_std):
                    excess_return = daily_returns.mean() * 252 - self.risk_free_rate
                    volatility = daily_std * np.sqrt(252)
                    if volatility > 0 and not np.isnan(volatility) and not np.isinf(volatility):
                        self.metrics['夏普比率'] = excess_return / volatility
                    else:
                        self.metrics['夏普比率'] = 0
                else:
                    self.metrics['夏普比率'] = 0
            else:
                self.metrics['夏普比率'] = 0

            # 交易统计
            self.metrics['交易次数'], self.metrics['胜率'], self.metrics['盈亏比'] = self._calculate_trade_stats()

            # 处理无效值
            for key in self.metrics:
                value = self.metrics[key]
                if pd.isna(value) or np.isinf(value):
                    self.metrics[key] = 0

        except Exception as e:
            print(f"计算策略指标时出错: {e}")
            self.metrics = {
                '总收益率': 0,
                '年化收益率': 0,
                '最大回撤率': 0,
                '夏普比率': 0,
                '交易次数': 0,
                '胜率': 0,
                '盈亏比': 0
            }

        # 计算买入持有策略指标
        self._calculate_buyhold_metrics()

        return self.metrics

    def _calculate_trade_stats(self) -> Tuple[int, float, float]:
        """
        计算交易统计

        Returns:
        --------
        Tuple[int, float, float] : 交易次数, 胜率, 盈亏比
        """
        try:
            # 使用 signal_executed 字段而不是 signal
            if 'signal_executed' in self.data.columns:
                trades = self.data[self.data['signal_executed'] != ''][
                    ['日期', 'signal_executed', '收盘', 'shares', 'commission', 'stamp_tax']].copy()
            else:
                trades = self.data[self.data['signal'] != ''][
                    ['日期', 'signal', '收盘', 'position', 'commission', 'stamp_tax']].copy()

            if trades.empty:
                print("没有交易记录")
                return 0, 0, 1

            print(f"找到 {len(trades)} 条交易信号")

            profits = []
            buy_info = None

            for idx, row in trades.iterrows():
                signal_type = row['signal_executed'] if 'signal_executed' in row else row['signal']

                if signal_type == 'buy':
                    # 记录买入信息
                    buy_info = {
                        'date': row['日期'],
                        'price': row['收盘'],
                        'shares': row['shares'] if 'shares' in row else row['position'],
                        'commission': row.get('commission', 0)
                    }
                    print(f"买入: 日期={buy_info['date']}, 价格={buy_info['price']:.4f}, 股数={buy_info['shares']}")

                elif signal_type == 'add' and buy_info is not None:
                    # 加仓：更新买入信息（加权平均）
                    add_price = row['收盘']
                    add_shares = row['shares'] - buy_info['shares']
                    add_commission = row.get('commission', 0)

                    total_cost = buy_info['price'] * buy_info['shares'] + buy_info['commission']
                    total_cost += add_price * add_shares + add_commission
                    buy_info['shares'] += add_shares
                    buy_info['price'] = total_cost / buy_info['shares']
                    buy_info['commission'] = 0  # 手续费已包含在成本中
                    print(
                        f"加仓: 日期={row['日期']}, 价格={add_price:.4f}, 加仓股数={add_shares}, 新成本={buy_info['price']:.4f}")

                elif signal_type == 'sell' and buy_info is not None:
                    # 计算盈亏
                    sell_price = row['收盘']
                    sell_commission = row.get('commission', 0)
                    sell_stamp_tax = row.get('stamp_tax', 0)

                    sell_amount = buy_info['shares'] * sell_price
                    sell_cost = sell_commission + sell_stamp_tax
                    buy_cost = buy_info['price'] * buy_info['shares']

                    profit = sell_amount - sell_cost - buy_cost
                    profit_rate = profit / buy_cost if buy_cost > 0 else 0

                    profits.append(profit)
                    print(
                        f"卖出: 日期={row['日期']}, 价格={sell_price:.4f}, 股数={buy_info['shares']}, 盈亏={profit:.2f}, 收益率={profit_rate:.2%}")

                    buy_info = None  # 重置

            if not profits:
                print("没有完整的买卖对")
                return 0, 0, 1

            # 计算统计指标
            win_count = sum(1 for p in profits if p > 0)
            loss_count = sum(1 for p in profits if p < 0)
            total_trades = win_count + loss_count

            win_rate = win_count / total_trades if total_trades > 0 else 0

            total_profit = sum(p for p in profits if p > 0)
            total_loss = abs(sum(p for p in profits if p < 0))
            profit_loss_ratio = total_profit / total_loss if total_loss > 0 else float('inf')

            # 处理无限大
            if np.isinf(profit_loss_ratio) or profit_loss_ratio > 100:
                profit_loss_ratio = 0

            print(
                f"交易统计: 总交易={total_trades}, 盈利={win_count}, 亏损={loss_count}, 胜率={win_rate:.2%}, 盈亏比={profit_loss_ratio:.2f}")

            return total_trades, win_rate, profit_loss_ratio

        except Exception as e:
            print(f"计算交易统计时出错: {e}")
            import traceback
            traceback.print_exc()
            return 0, 0, 1

    def _calculate_buyhold_metrics(self) -> Dict:
        """计算买入持有策略指标"""
        try:
            # 调用买入持有策略类
            bh_strategy = BuyHoldStrategy(
                initial_capital=self.initial_capital,
                commission_rate=self.commission_rate,
                stamp_duty_rate=self.stamp_tax_rate,
                min_commission=self.min_commission
            )

            # 运行回测
            bh_strategy.run_backtest(self.data)

            # 计算指标
            bh_metrics = bh_strategy.calculate_metrics(risk_free_rate=self.risk_free_rate)

            # 获取最大回撤并取绝对值
            max_drawdown = bh_metrics.get('最大回撤', 0)
            if max_drawdown is not None:
                max_drawdown = abs(max_drawdown)

            # 转换指标名称
            self.buyhold_metrics = {
                '总收益率': bh_metrics.get('总收益率', 0),
                '年化收益率': bh_metrics.get('年化收益率', 0),
                '最大回撤率': max_drawdown,
                '夏普比率': bh_metrics.get('夏普比率', 0)
            }

        except Exception as e:
            print(f"计算买入持有指标时出错: {e}")
            self.buyhold_metrics = {
                '总收益率': 0,
                '年化收益率': 0,
                '最大回撤率': 0,
                '夏普比率': 0
            }

        return self.buyhold_metrics

    def plot_result(self, output_path: str = None) -> plt.Figure:
        """
        绘制回测结果图表

        Parameters:
        -----------
        output_path : str, optional
            图表保存路径，如果不提供则只显示不保存

        Returns:
        --------
        plt.Figure : matplotlib图表对象
        """
        if self.data is None:
            raise ValueError("请先运行回测")

        # 创建4个子图
        fig, axes = plt.subplots(4, 1, figsize=(14, 16), sharex=True)

        # 子图1：价格 + 均线 + 买卖点
        ax1 = axes[0]
        ax1.plot(self.data['日期'], self.data['收盘'], label='收盘价', color='black', linewidth=1)
        ax1.plot(self.data['日期'], self.data[self.ma_short_col],
                 label=f'MA{self.short_ma}', color='orange', alpha=0.7)
        ax1.plot(self.data['日期'], self.data[self.ma_long_col],
                 label=f'MA{self.long_ma}', color='blue', alpha=0.7)

        # 标记买卖点
        buy_points = self.data[self.data['signal'] == 'buy']
        sell_points = self.data[self.data['signal'] == 'sell']
        ax1.scatter(buy_points['日期'], buy_points['收盘'],
                    marker='^', color='red', s=100, label='买入', zorder=5)
        ax1.scatter(sell_points['日期'], sell_points['收盘'],
                    marker='v', color='green', s=100, label='卖出', zorder=5)

        ax1.set_title(f'价格走势与交易信号 (MA{self.short_ma} vs MA{self.long_ma})', fontsize=12)
        ax1.legend(loc='upper left')
        ax1.set_ylabel('价格')
        ax1.grid(True, alpha=0.3)

        # 子图2：持仓状态
        ax2 = axes[1]
        ax2.fill_between(self.data['日期'], 0, self.data['position'] > 0,
                         alpha=0.3, color='blue', label='持仓')
        ax2.set_title('持仓状态', fontsize=12)
        ax2.set_ylabel('持仓(1=有,0=无)')
        ax2.set_ylim(-0.1, 1.1)
        ax2.grid(True, alpha=0.3)

        # 子图3：策略净值 vs 买入持有净值（对数坐标）
        ax3 = axes[2]

        # 归一化到1起点
        strategy_nav = self.data['total_asset'] / self.initial_capital

        # 计算买入持有净值
        bh_strategy = BuyHoldStrategy(
            initial_capital=self.initial_capital,
            commission_rate=self.commission_rate,
            stamp_duty_rate=self.stamp_tax_rate,
            min_commission=self.min_commission
        )
        bh_net_values = bh_strategy.run_backtest(self.data)
        buyhold_nav = bh_net_values / self.initial_capital

        ax3.semilogy(self.data['日期'], strategy_nav, label='策略净值', color='red', linewidth=1.5)
        ax3.semilogy(self.data['日期'], buyhold_nav, label='买入持有净值', color='gray', linewidth=1.5, alpha=0.7)
        ax3.set_title('策略净值 vs 买入持有净值（对数坐标）', fontsize=12)
        ax3.legend(loc='upper left')
        ax3.set_ylabel('净值')
        ax3.grid(True, alpha=0.3)

        # 子图4：回撤曲线
        ax4 = axes[3]
        ax4.fill_between(self.data['日期'], 0, self.data['drawdown'],
                         alpha=0.5, color='red', label='回撤')
        ax4.axhline(y=self.metrics['最大回撤率'], color='darkred', linestyle='--',
                    label=f"最大回撤 {self.metrics['最大回撤率']:.2%}")
        ax4.set_title('回撤曲线', fontsize=12)
        ax4.set_ylabel('回撤幅度')
        ax4.set_xlabel('日期')
        ax4.legend(loc='lower left')
        ax4.grid(True, alpha=0.3)

        plt.tight_layout()

        # 保存图表
        if output_path:
            plt.savefig(output_path, dpi=300, bbox_inches='tight')
            print(f"图表已保存至: {output_path}")

        plt.show()
        return fig

    def write_report(self, output_path: str = None, risk_manager=None) -> str:
        """
        编写回测报告

        Parameters:
        -----------
        output_path : str, optional
            报告保存路径，如果不提供则自动生成
        risk_manager : RiskManage, optional
            风险管理器实例，用于生成详细的交易记录

        Returns:
        --------
        str : 生成的文件路径
        """
        if self.data is None or not self.metrics:
            raise ValueError("请先运行回测并计算指标")

        # 如果没有提供输出路径，自动生成
        if output_path is None:
            # 从股票代码生成报告文件名
            report_filename = f"{self.stock_code}_双均线_回测结果.xlsx"
            output_path = os.path.join(os.getcwd(), report_filename)
        else:
            # 如果提供了目录路径，在目录下生成文件
            if os.path.isdir(output_path):
                report_filename = f"{self.stock_code}_双均线_回测结果.xlsx"
                output_path = os.path.join(output_path, report_filename)
            # 如果提供了完整文件路径，检查文件名格式
            else:
                # 确保文件名包含股票代码
                dir_name = os.path.dirname(output_path)
                file_name = os.path.basename(output_path)
                if self.stock_code not in file_name:
                    # 如果文件名中不包含股票代码，重新生成
                    new_file_name = f"{self.stock_code}_双均线_回测结果.xlsx"
                    output_path = os.path.join(dir_name if dir_name else '.', new_file_name)

        print(f"\n生成回测报告: {output_path}")

        # 创建Excel写入器
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:

            # ========== Sheet1: 参数说明 ==========
            params_data = {
                '参数名称': [
                    '股票代码',
                    '短期均线天数',
                    '长期均线天数',
                    '初始资金',
                    '手续费率',
                    '最低手续费',
                    '印花税率',
                    '止损线',
                    '无风险利率',
                    '最小信号间隔',
                    '数据起始日期',
                    '数据结束日期',
                    '交易天数'
                ],
                '参数值': [
                    self.stock_code,
                    self.short_ma,
                    self.long_ma,
                    f"{self.initial_capital:,.2f}",
                    f"{self.commission_rate:.4%}",
                    f"{self.min_commission:.2f}",
                    f"{self.stamp_tax_rate:.4%}",
                    f"{self.stop_loss:.2%}" if self.stop_loss else '无',
                    f"{self.risk_free_rate:.2%}",
                    '5天',
                    self.data['日期'].iloc[0].strftime('%Y-%m-%d'),
                    self.data['日期'].iloc[-1].strftime('%Y-%m-%d'),
                    len(self.data)
                ]
            }

            # 添加风险管理参数
            if risk_manager is not None:
                state = risk_manager.get_risk_state()
                params_data['参数名称'].extend([
                    '风险管理', '风险预算比率', '加仓比例', '加仓间隔(ATR倍数)',
                    '止损/止盈(ATR倍数)'
                ])
                params_data['参数值'].extend([
                    '启用',
                    f"{risk_manager.risk_percent:.2%}",
                    str(risk_manager.add_ratios),
                    str(risk_manager.add_atr_multiple),
                    str(risk_manager.stop_atr_multiple)
                ])
            else:
                params_data['参数名称'].append('风险管理')
                params_data['参数值'].append('未启用')

            params_df = pd.DataFrame(params_data)
            params_df.to_excel(writer, sheet_name='参数说明', index=False)

            # ========== Sheet2: 绩效指标 ==========
            metrics_data = {
                '指标名称': ['总收益率', '年化收益率', '最大回撤', '夏普比率', '交易次数', '胜率', '盈亏比'],
                '本策略指标值': [
                    f"{self.metrics.get('总收益率', 0):.2%}",
                    f"{self.metrics.get('年化收益率', 0):.2%}",
                    f"{self.metrics.get('最大回撤率', 0):.2%}",
                    f"{self.metrics.get('夏普比率', 0):.2f}",
                    str(self.metrics.get('交易次数', 0)),
                    f"{self.metrics.get('胜率', 0):.2%}",
                    f"{self.metrics.get('盈亏比', 0):.2f}"
                ],
                '买入持有指标值': [
                    f"{self.buyhold_metrics.get('总收益率', 0):.2%}",
                    f"{self.buyhold_metrics.get('年化收益率', 0):.2%}",
                    f"{self.buyhold_metrics.get('最大回撤率', 0):.2%}",
                    f"{self.buyhold_metrics.get('夏普比率', 0):.2f}",
                    '1',
                    '100%',
                    '-'
                ]
            }
            metrics_df = pd.DataFrame(metrics_data)
            metrics_df.to_excel(writer, sheet_name='绩效指标', index=False)

            # ========== Sheet3: 日度数据 ==========
            # 计算累计收益率
            self.data['累计收益率'] = (self.data['total_asset'] / self.initial_capital - 1)

            # 选择需要的列
            daily_cols = ['日期', '开盘', '收盘', '最高', '最低', '成交量', '异常情况',
                          self.ma_short_col, self.ma_long_col, 'signal',
                          'total_asset', 'hold_value', 'cash', 'position',
                          'drawdown', 'daily_return', '累计收益率']

            # 如果存在风险管理相关字段，也加入
            if 'atr' in self.data.columns:
                daily_cols.append('atr')
            if 'risk_status' in self.data.columns:
                daily_cols.extend(['risk_status', 'add_count', 'avg_cost', 'stop_loss_price', 'stop_profit_price'])

            # 只保留存在的列
            existing_cols = [col for col in daily_cols if col in self.data.columns]
            daily_df = self.data[existing_cols].copy()

            # 重命名列名使其更友好
            column_mapping = {
                self.ma_short_col: f'MA{self.short_ma}',
                self.ma_long_col: f'MA{self.long_ma}',
                'signal': '交易信号',
                'total_asset': '总资产',
                'hold_value': '持仓市值',
                'cash': '可用资金',
                'position': '持仓股数',
                'drawdown': '回撤率',
                'daily_return': '日收益率',
                '累计收益率': '累计收益率',
                'atr': 'ATR',
                'risk_status': '持仓状态',
                'add_count': '已加仓次数',
                'avg_cost': '平均成本',
                'stop_loss_price': '止损价',
                'stop_profit_price': '止盈价'
            }
            daily_df.rename(columns=column_mapping, inplace=True)

            # 格式化百分比列
            for col in ['回撤率', '日收益率', '累计收益率']:
                if col in daily_df.columns:
                    daily_df[col] = daily_df[col].apply(lambda x: f"{x:.2%}" if pd.notna(x) else '')

            daily_df.to_excel(writer, sheet_name='日度数据', index=False)

            # ========== Sheet4: 交易记录 ==========
            if risk_manager is not None:
                # 使用风险管理器的详细交易记录
                trade_records = self._generate_detailed_trade_records(risk_manager)
            else:
                # 原有简单交易记录
                trade_records = self._generate_simple_trade_records()

            trade_records.to_excel(writer, sheet_name='交易记录', index=False)

            # 调整列宽
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

        print(f"回测报告已保存至: {output_path}")
        return output_path

    def _generate_simple_trade_records(self) -> pd.DataFrame:
        """生成简单交易记录（无风险管理）"""
        trades = self.data[self.data['signal'] != ''].copy()

        if len(trades) < 2:
            return pd.DataFrame(columns=['买入日期', '卖出日期', '买入价', '卖出价',
                                         '股数', '收益率', '盈亏金额', '交易类型'])

        trade_records = []
        i = 0
        while i < len(trades) - 1:
            if trades.iloc[i]['signal'] == 'buy':
                buy_row = trades.iloc[i]

                for j in range(i + 1, len(trades)):
                    if trades.iloc[j]['signal'] == 'sell':
                        sell_row = trades.iloc[j]

                        buy_price = buy_row['收盘']
                        sell_price = sell_row['收盘']
                        shares = buy_row['position']

                        if shares > 0:
                            buy_amount = shares * buy_price
                            buy_commission = buy_row.get('commission', 0)
                            buy_cost = buy_amount + buy_commission

                            sell_amount = shares * sell_price
                            sell_commission = sell_row.get('commission', 0)
                            sell_stamp_tax = sell_row.get('stamp_tax', 0)
                            sell_revenue = sell_amount - sell_commission - sell_stamp_tax

                            profit = sell_revenue - buy_cost
                            profit_rate = profit / buy_cost if buy_cost > 0 else 0

                            if self.stop_loss and profit_rate <= -self.stop_loss:
                                trade_type = '止损'
                            else:
                                trade_type = '正常'

                            trade_records.append({
                                '买入日期': buy_row['日期'],
                                '卖出日期': sell_row['日期'],
                                '买入价': f"{buy_price:.4f}",
                                '卖出价': f"{sell_price:.4f}",
                                '股数': shares,
                                '收益率': f"{profit_rate:.2%}",
                                '盈亏金额': f"{profit:.2f}",
                                '交易类型': trade_type
                            })

                        i = j + 1
                        break
                else:
                    i += 1
            else:
                i += 1

        return pd.DataFrame(trade_records)

    def _generate_detailed_trade_records(self, risk_manager) -> pd.DataFrame:
        """生成详细交易记录（使用风险管理器）"""
        # 获取所有实际执行的信号，使用 signal_executed 字段
        trades = self.data[self.data['signal_executed'] != ''].copy()
        trades = trades.sort_values('日期')

        if len(trades) == 0:
            return pd.DataFrame()

        # 获取加仓次数（从风险管理器获取）
        add_count_total = risk_manager.add_count_total if hasattr(risk_manager, 'add_count_total') else 0

        # 构建列名
        columns = ['建仓日期', '建仓价', '建仓股数', '建仓止损价', '卖出日期', '卖出价',
                   '平均成本', '总股数', '收益率', '盈亏金额', '持仓天数', '离场类型']

        for i in range(1, add_count_total + 1):
            columns.append(f'第{i}次加仓日期')
            columns.append(f'第{i}次加仓股数')
            columns.append(f'第{i}次加仓价格')
            columns.append(f'第{i}次加仓止损价')

        trade_records = []
        idx = 0

        while idx < len(trades):
            row = trades.iloc[idx]

            # 找到建仓信号
            if row['signal_executed'] == 'buy':
                # 建仓信息
                buy_date = row['日期']
                buy_price = row['收盘']
                buy_shares = int(row['shares'])
                buy_commission = row.get('commission', 0)

                # 建仓总成本（包含手续费）
                buy_total_cost = buy_price * buy_shares + buy_commission

                # 获取建仓止损价
                buy_stop_loss = row.get('stop_loss_price', 0)
                buy_stop_loss_display = f"{buy_stop_loss:.4f}" if buy_stop_loss > 0 else '无'

                # 收集加仓记录
                add_records = []
                next_idx = idx + 1

                # 收集所有加仓信号
                while next_idx < len(trades) and trades.iloc[next_idx]['signal_executed'] == 'add':
                    add_row = trades.iloc[next_idx]

                    # 计算加仓股数
                    if next_idx > idx + 1:
                        prev_shares = trades.iloc[next_idx - 1]['shares']
                        current_shares = add_row['shares']
                        add_shares = int(current_shares - prev_shares)
                    else:
                        add_shares = int(add_row['shares'] - buy_shares)

                    if add_shares <= 0:
                        next_idx += 1
                        continue

                    # 加仓信息（包含手续费）
                    add_price = add_row['收盘']
                    add_commission = add_row.get('commission', 0)
                    add_total_cost = add_price * add_shares + add_commission

                    # 获取加仓止损价
                    add_stop_loss = add_row.get('stop_loss_price', 0)
                    add_stop_loss_display = f"{add_stop_loss:.4f}" if add_stop_loss > 0 else '无'

                    add_records.append({
                        '日期': add_row['日期'],
                        '价格': add_price,
                        '股数': add_shares,
                        '手续费': add_commission,
                        '总成本': add_total_cost,
                        '止损价': add_stop_loss_display
                    })
                    next_idx += 1

                # 找到卖出信号
                if next_idx < len(trades) and trades.iloc[next_idx]['signal_executed'] == 'sell':
                    sell_row = trades.iloc[next_idx]
                    sell_date = sell_row['日期']
                    sell_price = sell_row['收盘']
                    sell_commission = sell_row.get('commission', 0)
                    sell_stamp_tax = sell_row.get('stamp_tax', 0)

                    # 计算总股数
                    total_shares = buy_shares + sum(a['股数'] for a in add_records)

                    # 计算总成本（包含所有买入手续费）
                    total_cost = buy_total_cost + sum(a['总成本'] for a in add_records)
                    avg_cost = total_cost / total_shares if total_shares > 0 else 0

                    # 计算卖出收入（扣除手续费和印花税）
                    sell_amount = total_shares * sell_price
                    sell_revenue = sell_amount - sell_commission - sell_stamp_tax

                    # 计算盈亏
                    profit = sell_revenue - total_cost
                    profit_rate = profit / total_cost if total_cost > 0 else 0

                    # 持仓天数
                    holding_days = (sell_date - buy_date).days

                    # 离场类型
                    exit_reason = sell_row.get('exit_reason', '卖出信号')
                    if exit_reason not in ['止损', '止盈', '卖出信号']:
                        exit_reason = '卖出信号'

                    # 构建记录
                    record = {
                        '建仓日期': buy_date.strftime('%Y-%m-%d') if hasattr(buy_date, 'strftime') else str(buy_date),
                        '建仓价': f"{buy_price:.4f}",
                        '建仓股数': buy_shares,
                        '建仓止损价': buy_stop_loss_display,
                        '卖出日期': sell_date.strftime('%Y-%m-%d') if hasattr(sell_date, 'strftime') else str(
                            sell_date),
                        '卖出价': f"{sell_price:.4f}",
                        '平均成本': f"{avg_cost:.4f}",
                        '总股数': total_shares,
                        '收益率': f"{profit_rate:.2%}",
                        '盈亏金额': f"{profit:.2f}",
                        '持仓天数': holding_days,
                        '离场类型': exit_reason
                    }

                    # 添加加仓记录
                    for k, add in enumerate(add_records, 1):
                        record[f'第{k}次加仓日期'] = add['日期'].strftime('%Y-%m-%d') if hasattr(add['日期'],
                                                                                                 'strftime') else str(
                            add['日期'])
                        record[f'第{k}次加仓股数'] = add['股数']
                        record[f'第{k}次加仓价格'] = f"{add['价格']:.4f}"
                        record[f'第{k}次加仓止损价'] = add['止损价']

                    # 填充未使用的加仓字段
                    actual_add_count = len(add_records)
                    for k in range(actual_add_count + 1, add_count_total + 1):
                        record[f'第{k}次加仓日期'] = ''
                        record[f'第{k}次加仓股数'] = 0
                        record[f'第{k}次加仓价格'] = ''
                        record[f'第{k}次加仓止损价'] = ''

                    trade_records.append(record)
                    idx = next_idx + 1
                else:
                    idx += 1
            else:
                idx += 1

        if not trade_records:
            print("警告: 没有生成任何交易记录")
            return pd.DataFrame(columns=columns)

        print(f"生成了 {len(trade_records)} 条交易记录")
        return pd.DataFrame(trade_records)

    def run_complete_analysis(self,
                              filepath: str,
                              output_folder: str = './results',
                              risk_manager=None) -> 'DoubleMovingAverageStrategy':
        """
        运行完整的分析流程

        Parameters:
        -----------
        filepath : str
            数据文件路径
        output_folder : str
            输出文件夹路径
        risk_manager : RiskManage, optional
            风险管理器实例，如果为None则不启用风险管理

        Returns:
        --------
        self : DoubleMovingAverageStrategy
            策略实例
        """
        # 创建输出文件夹
        os.makedirs(output_folder, exist_ok=True)

        # 运行分析流程
        self.load_data(filepath)
        self.preprocess_data()
        self.run_backtest(risk_manager=risk_manager)
        self.calculate_metrics()

        # 生成报告和图表
        base_name = f"{self.stock_code}_{self.stock_name}_MA{self.short_ma}_{self.long_ma}"

        # 保存图表
        chart_path = os.path.join(output_folder, f"{base_name}_回测图表.png")
        self.plot_result(chart_path)

        # 保存报告（传入风险管理器）
        report_path = os.path.join(output_folder, f"{base_name}_回测报告.xlsx")
        self.write_report(report_path, risk_manager=risk_manager)

        return self

    @staticmethod
    def compare_stocks(file_list: Union[str, List[str]],
                       output_folder: str = './stock_comparison',
                       short_ma: int = 10,
                       long_ma: int = 120,
                       initial_capital: float = 100000,
                       commission_rate: float = 0.0001,
                       min_commission: float = 5,
                       stamp_tax_rate: float = 0.001,
                       stop_loss: float = None,
                       risk_free_rate: float = 0.02) -> pd.DataFrame:
        """
        多股票对比

        Parameters:
        -----------
        file_list : Union[str, List[str]]
            文件夹路径或文件路径列表
        output_folder : str
            输出文件夹路径
        short_ma : int
            短期均线周期
        long_ma : int
            长期均线周期
        initial_capital : float
            初始资金
        commission_rate : float
            手续费率
        min_commission : float
            最低手续费
        stamp_tax_rate : float
            印花税率
        stop_loss : float, optional
            止损线
        risk_free_rate : float
            无风险利率

        Returns:
        --------
        pd.DataFrame : 多股票对比明细表
        """
        # 创建输出文件夹
        os.makedirs(output_folder, exist_ok=True)

        # 获取文件列表
        if isinstance(file_list, str) and os.path.isdir(file_list):
            file_list = list(Path(file_list).glob("*.csv"))

        results = []

        for filepath in file_list:
            try:
                print(f"\n{'=' * 50}")
                print(f"正在分析: {filepath}")
                print('=' * 50)

                # 创建策略实例
                strategy = DoubleMovingAverageStrategy(
                    short_ma=short_ma,
                    long_ma=long_ma,
                    initial_capital=initial_capital,
                    commission_rate=commission_rate,
                    min_commission=min_commission,
                    stamp_tax_rate=stamp_tax_rate,
                    stop_loss=stop_loss,
                    risk_free_rate=risk_free_rate
                )

                # 运行分析
                strategy.load_data(str(filepath))
                strategy.preprocess_data()
                strategy.run_backtest()
                strategy.calculate_metrics()

                # 收集结果
                results.append({
                    '股票代码': strategy.stock_code,
                    '股票名称': strategy.stock_name,
                    '策略总收益率': strategy.metrics.get('总收益率', 0),
                    '策略年化收益率': strategy.metrics.get('年化收益率', 0),
                    '策略最大回撤率': strategy.metrics.get('最大回撤率', 0),
                    '策略夏普比率': strategy.metrics.get('夏普比率', 0),
                    '策略胜率': strategy.metrics.get('胜率', 0),
                    '策略盈亏比': strategy.metrics.get('盈亏比', 0),
                    '策略交易次数': strategy.metrics.get('交易次数', 0),
                    '买入持有总收益率': strategy.buyhold_metrics.get('总收益率', 0),
                    '买入持有最大回撤率': strategy.buyhold_metrics.get('最大回撤率', 0),
                    '买入持有夏普比率': strategy.buyhold_metrics.get('夏普比率', 0),
                    '超额总收益率': strategy.metrics.get('总收益率', 0) - strategy.buyhold_metrics.get('总收益率', 0),
                    '回撤改善': strategy.buyhold_metrics.get('最大回撤率', 0) - strategy.metrics.get('最大回撤率', 0)
                })

            except Exception as e:
                print(f"分析 {filepath} 时出错: {e}")
                continue

        # 创建结果DataFrame
        results_df = pd.DataFrame(results)

        if results_df.empty:
            print("没有有效的分析结果")
            return results_df

        # 保存明细表
        excel_path = os.path.join(output_folder, '双均线_多股票对比明细表.xlsx')
        results_df.to_excel(excel_path, index=False)
        print(f"对比明细表已保存至: {excel_path}")

        # 绘制对比图
        plot_path = os.path.join(output_folder, '双均线_多股票对比图.png')
        DoubleMovingAverageStrategy._plot_stock_comparison(results_df, plot_path)

        # 打印总体效果
        DoubleMovingAverageStrategy._print_stock_comparison_summary(results_df)

        return results_df

    @staticmethod
    def _plot_stock_comparison(results_df: pd.DataFrame, output_path: str) -> plt.Figure:
        """绘制多股票对比图"""
        if results_df.empty:
            print("没有可绘制的数据")
            return None

        # 处理无效值
        plot_df = results_df.copy()
        for col in plot_df.columns:
            if col not in ['股票代码', '股票名称']:
                plot_df[col] = plot_df[col].apply(lambda x: x if pd.notna(x) and not np.isinf(x) else 0)

        fig, axes = plt.subplots(2, 2, figsize=(16, 12))

        # 图1: 最大回撤分布对比直方图
        ax1 = axes[0, 0]

        # 过滤异常值
        mask = (plot_df['策略最大回撤率'] <= 1) & (plot_df['买入持有最大回撤率'] <= 1)
        plot_df_1 = plot_df[mask]

        bins = 30
        ax1.hist(plot_df_1['策略最大回撤率'], bins=bins, alpha=0.5,
                 label='策略回撤', color='red', edgecolor='black', density=True)
        ax1.hist(plot_df_1['买入持有最大回撤率'], bins=bins, alpha=0.5,
                 label='买入持有回撤', color='blue', edgecolor='black', density=True)

        # 添加均值线
        ax1.axvline(x=plot_df_1['策略最大回撤率'].mean(), color='red', linestyle='--',
                    linewidth=2, label=f'策略平均: {plot_df_1["策略最大回撤率"].mean():.2%}')
        ax1.axvline(x=plot_df_1['买入持有最大回撤率'].mean(), color='blue', linestyle='--',
                    linewidth=2, label=f'持有平均: {plot_df_1["买入持有最大回撤率"].mean():.2%}')

        ax1.set_xlabel('最大回撤')
        ax1.set_ylabel('股票个数')
        ax1.set_title('最大回撤分布对比直方图')
        ax1.legend()
        ax1.grid(True, alpha=0.3)

        # 图2: 本策略胜率 vs 盈亏比四象限散点图
        ax2 = axes[0, 1]

        # 处理盈亏比无限大的情况
        plot_df['盈亏比_clean'] = plot_df['策略盈亏比'].replace([np.inf, -np.inf], 0)

        scatter = ax2.scatter(plot_df['策略胜率'], plot_df['盈亏比_clean'],
                              c=plot_df['策略夏普比率'], cmap='RdYlGn', s=50, alpha=0.6)

        # 绘制象限分割线
        ax2.axhline(y=1, color='gray', linestyle='--', alpha=0.5, label='盈亏比=1')
        ax2.axvline(x=0.5, color='gray', linestyle='--', alpha=0.5, label='胜率=50%')

        ax2.set_xlabel('胜率')
        ax2.set_ylabel('盈亏比')
        ax2.set_title('胜率 vs 盈亏比四象限散点图')
        ax2.set_xlim(0, 1)
        ax2.grid(True, alpha=0.3)
        plt.colorbar(scatter, ax=ax2, label='夏普比率')
        ax2.legend()

        # 图3: 本策略收益率 vs 买入持有策略收益率散点图
        ax3 = axes[1, 0]

        # 设置坐标轴范围
        x_min, x_max = -1, 10
        y_min, y_max = -1, 5

        # 过滤异常值
        mask = (plot_df['买入持有总收益率'] >= x_min) & (plot_df['买入持有总收益率'] <= x_max) & \
               (plot_df['策略总收益率'] >= y_min) & (plot_df['策略总收益率'] <= y_max)
        plot_df_3 = plot_df[mask].copy()

        if not plot_df_3.empty:
            # 绘制45度线
            ax3.plot([x_min, x_max], [x_min, x_max], 'k--', alpha=0.5, label='45度线')

            scatter3 = ax3.scatter(plot_df_3['买入持有总收益率'], plot_df_3['策略总收益率'],
                                   c=plot_df_3['超额总收益率'], cmap='RdYlGn', s=50, alpha=0.6)
            plt.colorbar(scatter3, ax=ax3, label='超额收益')

        ax3.set_xlabel('买入持有收益率')
        ax3.set_ylabel('策略收益率')
        ax3.set_title('策略收益率 vs 买入持有收益率散点图')
        ax3.set_xlim(x_min, x_max)
        ax3.set_ylim(y_min, y_max)
        ax3.grid(True, alpha=0.3)
        ax3.legend()

        # 图4: 本策略超额收益 vs 回撤改善四象限散点图
        ax4 = axes[1, 1]

        # 设置纵轴范围
        y_max_4 = 1.5

        # 过滤异常值
        mask = plot_df['回撤改善'] <= y_max_4
        plot_df_4 = plot_df[mask].copy()

        if not plot_df_4.empty:
            scatter4 = ax4.scatter(plot_df_4['超额总收益率'], plot_df_4['回撤改善'],
                                   c=plot_df_4['策略夏普比率'], cmap='RdYlGn', s=50, alpha=0.6)
            plt.colorbar(scatter4, ax=ax4, label='策略夏普比率')

        # 绘制象限分割线
        ax4.axhline(y=0, color='gray', linestyle='--', alpha=0.5)
        ax4.axvline(x=0, color='gray', linestyle='--', alpha=0.5)

        ax4.set_xlabel('超额收益')
        ax4.set_ylabel('回撤改善')
        ax4.set_title('超额收益 vs 回撤改善四象限散点图')
        ax4.grid(True, alpha=0.3)

        plt.suptitle('多股票策略对比分析', fontsize=16, y=1.02)
        plt.tight_layout()

        # 保存图表
        plt.savefig(output_path, dpi=300, bbox_inches='tight')
        print(f"对比图已保存至: {output_path}")
        plt.show()

        return fig

    @staticmethod
    def _print_stock_comparison_summary(results_df: pd.DataFrame) -> None:
        """打印多股票对比总体效果"""
        if results_df.empty:
            print("没有数据可显示")
            return

        # 过滤有效数据
        valid_df = results_df.copy()

        print("\n" + "=" * 60)
        print("策略总体效果")
        print("=" * 60)
        print(f"分析股票总数: {len(valid_df)}")

        # 跑赢基准比例
        beat_count = (valid_df['策略总收益率'] > valid_df['买入持有总收益率']).sum()
        beat_ratio = beat_count / len(valid_df)
        print(f"跑赢基准比例: {beat_ratio:.1%} ({beat_count}/{len(valid_df)})")

        # 回撤改善比例
        improve_count = (valid_df['策略最大回撤率'] < valid_df['买入持有最大回撤率']).sum()
        improve_ratio = improve_count / len(valid_df)
        print(f"回撤改善比例: {improve_ratio:.1%} ({improve_count}/{len(valid_df)})")

        # 平均回撤
        print(f"策略平均回撤: {valid_df['策略最大回撤率'].mean():.2%} vs "
              f"买入持有平均回撤: {valid_df['买入持有最大回撤率'].mean():.2%}")

        # 平均夏普
        print(f"策略平均夏普: {valid_df['策略夏普比率'].mean():.2f} vs "
              f"买入持有平均夏普: {valid_df['买入持有夏普比率'].mean():.2f}")

        # 其他平均指标
        print(f"\n平均总收益率: {valid_df['策略总收益率'].mean():.2%}")
        print(f"平均交易次数: {valid_df['策略交易次数'].mean():.1f}")
        print(f"平均胜率: {valid_df['策略胜率'].mean():.2%}")
        print(f"平均盈亏比: {valid_df['策略盈亏比'].mean():.2f}")
        print("=" * 60)

    @staticmethod
    def compare_strategies(filepath: str,
                           ma_pairs: List[Tuple[int, int]],
                           output_folder: str = './strategy_comparison',
                           initial_capital: float = 100000,
                           commission_rate: float = 0.0001,
                           min_commission: float = 5,
                           stamp_tax_rate: float = 0.001,
                           stop_loss: float = None,
                           risk_free_rate: float = 0.02) -> pd.DataFrame:
        """
        比较同一只股票的不同参数组合

        Parameters:
        -----------
        filepath : str
            数据文件路径
        ma_pairs : List[Tuple[int, int]]
            均线参数对列表，如 [(5,20), (10,60), (20,120)]
        output_folder : str
            输出文件夹路径
        initial_capital : float
            初始资金
        commission_rate : float
            手续费率
        min_commission : float
            最低手续费
        stamp_tax_rate : float
            印花税率
        stop_loss : float, optional
            止损线
        risk_free_rate : float
            无风险利率

        Returns:
        --------
        pd.DataFrame : 参数对比详细数据
        """
        # 创建输出文件夹
        os.makedirs(output_folder, exist_ok=True)

        # 获取股票代码
        temp_strategy = DoubleMovingAverageStrategy()
        temp_strategy.load_data(filepath)
        stock_code = temp_strategy.stock_code

        results = []
        all_navs = {}  # 存储所有策略的净值序列，用于绘图

        for short_ma, long_ma in ma_pairs:
            try:
                print(f"\n测试均线组合: MA{short_ma} vs MA{long_ma}")
                print("-" * 40)

                strategy = DoubleMovingAverageStrategy(
                    short_ma=short_ma,
                    long_ma=long_ma,
                    initial_capital=initial_capital,
                    commission_rate=commission_rate,
                    min_commission=min_commission,
                    stamp_tax_rate=stamp_tax_rate,
                    stop_loss=stop_loss,
                    risk_free_rate=risk_free_rate
                )

                strategy.run_complete_analysis(filepath, output_folder)

                # 存储净值序列
                strategy_nav = strategy.data['total_asset'] / initial_capital
                all_navs[f'MA{short_ma}-{long_ma}'] = {
                    'nav': strategy_nav,
                    'dates': strategy.data['日期']
                }

                # 收集结果
                results.append({
                    '参数组合': f'MA{short_ma}-{long_ma}',
                    '总收益率': strategy.metrics.get('总收益率', 0),
                    '年化收益率': strategy.metrics.get('年化收益率', 0),
                    '最大回撤率': strategy.metrics.get('最大回撤率', 0),
                    '夏普比率': strategy.metrics.get('夏普比率', 0),
                    '胜率': strategy.metrics.get('胜率', 0),
                    '盈亏比': strategy.metrics.get('盈亏比', 0),
                    '交易次数': strategy.metrics.get('交易次数', 0),
                    '超额收益': strategy.metrics.get('总收益率', 0) - strategy.buyhold_metrics.get('总收益率', 0),
                    '回撤改善': strategy.buyhold_metrics.get('最大回撤率', 0) - strategy.metrics.get('最大回撤率', 0)
                })

            except Exception as e:
                print(f"测试 MA{short_ma}-{long_ma} 失败: {e}")
                continue

        # 创建结果DataFrame
        results_df = pd.DataFrame(results)

        # 添加买入持有策略
        bh_strategy = BuyHoldStrategy(
            initial_capital=initial_capital,
            commission_rate=commission_rate,
            stamp_duty_rate=stamp_tax_rate,
            min_commission=min_commission
        )

        # 加载数据并计算买入持有
        data = pd.read_csv(filepath, encoding=temp_strategy._detect_encoding(filepath))
        if '日期' in data.columns:
            data['日期'] = pd.to_datetime(data['日期'])
        data = data.sort_values('日期').reset_index(drop=True)

        bh_net_values = bh_strategy.run_backtest(data)
        bh_metrics = bh_strategy.calculate_metrics(risk_free_rate=risk_free_rate)

        bh_row = {
            '参数组合': '买入持有',
            '总收益率': bh_metrics.get('总收益率', 0),
            '年化收益率': bh_metrics.get('年化收益率', 0),
            '最大回撤率': bh_metrics.get('最大回撤', 0),
            '夏普比率': bh_metrics.get('夏普比率', 0),
            '胜率': 0,
            '盈亏比': 0,
            '交易次数': 1,
            '超额收益': 0,
            '回撤改善': 0
        }

        results_df = pd.concat([results_df, pd.DataFrame([bh_row])], ignore_index=True)

        # 保存详细数据
        excel_path = os.path.join(output_folder, f'{stock_code}_双均线参数对比详细数据.xlsx')
        results_df.to_excel(excel_path, index=False)
        print(f"参数对比详细数据已保存至: {excel_path}")

        # 绘制净值曲线
        plot_path = os.path.join(output_folder, f'{stock_code}_双均线参数对比净值曲线.png')
        DoubleMovingAverageStrategy._plot_strategies_comparison(
            all_navs, bh_net_values / initial_capital, data['日期'], plot_path
        )

        return results_df

    @staticmethod
    def _plot_strategies_comparison(all_navs: Dict, bh_nav: pd.Series,
                                    dates: pd.Series, output_path: str) -> plt.Figure:
        """绘制多策略净值对比图"""
        fig, ax = plt.subplots(figsize=(14, 8))

        # 绘制买入持有净值
        ax.semilogy(dates, bh_nav, label='买入持有', color='black',
                    linewidth=2, linestyle='--', alpha=0.7)

        # 绘制各策略净值
        colors = plt.cm.tab10(np.linspace(0, 1, len(all_navs)))

        for (label, data), color in zip(all_navs.items(), colors):
            ax.semilogy(data['dates'], data['nav'], label=label,
                        color=color, linewidth=1.5, alpha=0.8)

        ax.set_xlabel('日期')
        ax.set_ylabel('净值（对数坐标）')
        ax.set_title('不同参数双均线策略净值对比')
        ax.legend(loc='upper left', fontsize=9, ncol=2)
        ax.grid(True, alpha=0.3)

        plt.tight_layout()
        plt.savefig(output_path, dpi=300, bbox_inches='tight')
        print(f"净值曲线已保存至: {output_path}")
        plt.show()

        return fig

class RSIStrategy:
    """RSI策略主类"""

    def __init__(self,
                 initial_capital: float = 1000000,
                 commission_rate: float = 0.0001,
                 stamp_tax_rate: float = 0.001,
                 min_commission: float = 5,
                 stop_loss: float = None,
                 risk_free_rate: float = 0.03,
                 rsi_period: int = 14,
                 oversold_threshold: int = 30,
                 overbought_threshold: int = 70):
        """
        初始化RSI策略

        Args:
            initial_capital: 初始资金
            commission_rate: 佣金费率
            stamp_tax_rate: 印花税费率
            min_commission: 最低佣金
            stop_loss: 止损线（百分比，如0.1表示10%止损）
            risk_free_rate: 无风险利率
            rsi_period: RSI指标使用K线数量
            oversold_threshold: 超卖区对应RSI值
            overbought_threshold: 超买区对应RSI值
        """
        self.initial_capital = initial_capital
        self.commission_rate = commission_rate
        self.stamp_tax_rate = stamp_tax_rate
        self.min_commission = min_commission
        self.stop_loss = stop_loss
        self.risk_free_rate = risk_free_rate
        self.rsi_period = rsi_period
        self.oversold_threshold = oversold_threshold
        self.overbought_threshold = overbought_threshold

        # 数据相关
        self.data = None
        self.stock_code = None
        self.stock_name = None

        # 结果相关
        self.metrics = {}
        self.trade_records = []

    def _detect_encoding(self, file_path: str) -> str:
        """检测文件编码"""
        with open(file_path, 'rb') as f:
            result = chardet.detect(f.read(10000))
        return result['encoding'] or 'utf-8'

    def load_data(self, file_path: str) -> pd.DataFrame:
        """
        加载数据

        Args:
            file_path: 数据文件路径

        Returns:
            加载的数据框
        """
        # 检测文件编码
        encoding = self._detect_encoding(file_path)

        try:
            self.data = pd.read_csv(file_path, encoding=encoding)
        except UnicodeDecodeError:
            # 如果检测到的编码还是不对，尝试常见编码
            for enc in ['utf-8', 'gbk', 'gb2312', 'gb18030', 'latin1']:
                try:
                    self.data = pd.read_csv(file_path, encoding=enc)
                    break
                except UnicodeDecodeError:
                    continue

        # 从文件名提取股票代码
        file_stem = Path(file_path).stem
        if '_' in file_stem:
            parts = file_stem.split('_')
            if len(parts) >= 3:
                self.stock_code = parts[2]
            else:
                self.stock_code = file_stem
        else:
            self.stock_code = file_stem

        self.stock_name = self.stock_code

        # 打印数据列名，用于调试
        print(f"数据列名: {list(self.data.columns)}")

        # 确保日期格式正确 - 更灵活的列名处理
        date_col = None
        for col in ['日期', 'date', 'Date', 'trade_date']:
            if col in self.data.columns:
                date_col = col
                break

        if date_col:
            if date_col != '日期':
                self.data.rename(columns={date_col: '日期'}, inplace=True)
            self.data['日期'] = pd.to_datetime(self.data['日期'])
        else:
            raise ValueError("数据中找不到日期列，请确保列名为'日期'、'date'、'Date'或'trade_date'")

        # 确保收盘价列存在 - 更灵活的列名处理
        close_col = None
        for col in ['收盘', 'close', 'Close', '收盘价']:
            if col in self.data.columns:
                close_col = col
                break

        if close_col:
            if close_col != '收盘':
                self.data.rename(columns={close_col: '收盘'}, inplace=True)
        else:
            raise ValueError("数据中找不到收盘价列，请确保列名为'收盘'、'close'、'Close'或'收盘价'")

        # 确保数据按日期排序
        self.data = self.data.sort_values('日期').reset_index(drop=True)

        print(f"数据加载成功: {len(self.data)}行, 日期范围: {self.data['日期'].min()} 至 {self.data['日期'].max()}")

        return self.data

    def preprocess_data(self) -> pd.DataFrame:
        """
        数据处理：计算所需的策略指标和交易信号字段

        Returns:
            处理后的数据框
        """
        if self.data is None:
            raise ValueError("请先加载数据")

        # 创建停牌标识
        self.data['is_suspend'] = (self.data['异常情况'] == '停牌').astype(int)
        # 创建异常涨跌幅标识
        self.data['abnormal_buy_forbidden'] = (self.data['异常涨跌幅'] > 0).astype(int)
        self.data['abnormal_sell_forbidden'] = (self.data['异常涨跌幅'] < 0).astype(int)

        # 确保数据按日期排序
        self.data = self.data.sort_values('日期').reset_index(drop=True)

        # 计算RSI指标
        self.data['RSI'] = TechnicalIndicators.calculate_rsi(df=self.data, close_col='收盘', suspend_col="is_suspend", period= self.rsi_period).shift(1)
        # 生成交易信号
        self.data['交易信号'] = TechnicalIndicators.generate_signals_rsi(df=self.data, rsi_col='RSI', suspend_col="is_suspend", oversold_threshold= self.oversold_threshold, overbought_threshold= self.overbought_threshold)
        # 计算atr
        self.data['ATR'] = TechnicalIndicators.calculate_atr(self.data, period=14, high_col='最高', low_col='最低', close_col='收盘', suspend_col='is_suspend')

        return self.data

    def run_backtest(self,risk_manager=None) -> Dict:
        """
        运行回测

        Returns:
            绩效指标字典
        """
        if self.data is None or '交易信号' not in self.data.columns:
            raise ValueError("请先执行数据预处理")

        # 初始化数据列
        self.data['持仓状态'] = 0  # 持仓状态
        self.data['策略净值'] = 0.0  # 策略净值
        self.data['持仓市值'] = 0.0  # 持有市值
        self.data['可用资金'] = 0.0  # 可用现金
        self.data['总资产'] = 0.0  # 总资产
        self.data['回撤率'] = 0.0
        self.data['日收益率'] = 0.0  # 每日收益
        self.data['累计收益率'] = 0.0
        self.data['shares'] = 0 #持仓股数
        self.data['exit_reason'] = '' #卖出原因
        self.data['commission'] = 0.0  # 佣金
        self.data['stamp_tax'] = 0.0  # 印花税
        self.data['平均成本'] = 0.0 # 平均持仓成本（含买入手续费）

        # 风险管理相关字段
        if risk_manager is not None:
            self.data['add_count'] = 0 #第X次加仓
            self.data['avg_cost'] = 0.0  # 平均成本
            self.data['stop_loss_price'] = 0.0 #止损价
            self.data['stop_profit_price'] = 0.0 #止盈价
            self.data['last_buy_price'] = 0.0 #前次买入价
            self.data['initial_atr'] = 0.0 #本次交易初始atr

        # 交易记录相关
        self.trade_records = []
        current_trade = {}

        # 回测变量初始化
        pending = ""  # 延后一天交易
        holding = 0 # 持仓状态，0为未持仓，1为持仓
        for index, row in self.data.iterrows():
            signal = row['交易信号']
            prev_close = self.data.at[index-1, '收盘'] if index > 0 else None

            # 统一更新字段，如有交易后续覆盖
            self.data.at[index,'持仓状态'] = self.data.at[index-1, '持仓状态'] if index > 0 else 0  # 持仓状态
            self.data.at[index, 'shares'] = self.data.at[index - 1, 'shares'] if index > 0 else 0  # 持仓股数
            self.data.at[index,'持仓市值'] = self.data.at[index, 'shares'] * self.data.at[index, '收盘'] if index > 0 else 0.0  # 持有市值
            self.data.at[index,'可用资金'] = self.data.at[index-1, '可用资金'] if index > 0 else float(self.initial_capital)  # 可用现金
            self.data.at[index,'总资产'] = self.data.at[index,'持仓市值'] + self.data.at[index,'可用资金']  # 总资产
            self.data.at[index,'平均成本'] = self.data.at[index-1,'平均成本'] if index > 0 else 0.0

            # 记录当前风险状态（每行都记录）
            if risk_manager is not None:
                state = risk_manager.get_risk_state()
                self.data.at[index, 'add_count'] = state['add_count']
                self.data.at[index, 'avg_cost'] = state['avg_cost']
                self.data.at[index, 'stop_loss_price'] = state['stop_loss']
                self.data.at[index, 'stop_profit_price'] = state['stop_profit']
                self.data.at[index, 'last_buy_price'] = state['last_buy_price']
                self.data.at[index, 'initial_atr'] = state['initial_atr']

            # 处理止损止盈（无论有无信号）
            if risk_manager is not None and holding == 1:
                should_exit, exit_reason = risk_manager.check_stop(prev_close = prev_close,current_atr=row['ATR'])
                if should_exit:
                    if row['abnormal_sell_forbidden'] == 0:
                        # 执行卖出（止损止盈）
                        holding = 0
                        pending = ""
                        price = row['收盘']
                        shares = self.data.at[index, 'shares']
                        cash = self.data.at[index, '可用资金']

                        trade_record = self._do_sell(price,shares,cash,index,current_trade,exit_reason)
                        current_trade.update(trade_record)
                        self.trade_records.append(current_trade)
                        current_trade = {}

                        risk_manager.update_risk_state('close')


                        continue  # 当天不再处理其他交易
                    else:
                        pending = "卖出"

            # 有风险管理时的交易操作
            if risk_manager is not None:
                if signal == '':
                    if holding == 1:
                        # 延迟卖出
                        if pending == "卖出" and row['abnormal_sell_forbidden'] == 0:
                            # 执行卖出（延迟卖出）
                            holding = 0
                            pending = ""
                            price = row['收盘']
                            shares = self.data.at[index, 'shares']
                            cash = self.data.at[index, '可用资金']

                            trade_record = self._do_sell(price, shares, cash, index, current_trade)
                            current_trade.update(trade_record)
                            self.trade_records.append(current_trade)
                            current_trade = {}

                            risk_manager.update_risk_state('close')

                            continue
                        # 执行加仓
                        if risk_manager.is_can_add(prev_close):
                            price = row['收盘']
                            shares = self.data.at[index, 'shares']
                            cash = self.data.at[index, '可用资金']
                            avgcost_plus_shares = self.data.at[index, '平均成本'] * shares
                            # 执行加仓逻辑
                            add_info = risk_manager.calculate_add_position_size(price,shares,cash)
                            shares_to_add = add_info['shares']
                            if shares_to_add > 0:
                                trade_amount = shares_to_add * price
                                commission = max(trade_amount * self.commission_rate, self.min_commission)
                                total_cost = trade_amount + commission
                                if cash >= total_cost:
                                    holding = 1
                                    pending = ""
                                    shares += shares_to_add
                                    cash -= total_cost
                                    # 更新字段
                                    self.data.at[index, '持仓状态'] = 1
                                    self.data.at[index, 'shares'] = shares
                                    self.data.at[index, '持仓市值'] = shares * price
                                    self.data.at[index, '可用资金'] = cash
                                    self.data.at[index, '总资产'] = cash + shares * price
                                    self.data.at[index, 'commission'] = commission
                                    self.data.at[index, 'stamp_tax'] = 0
                                    self.data.at[index, '平均成本'] = (avgcost_plus_shares + total_cost)/shares

                                    risk_manager.update_risk_state(
                                        'add',
                                        old_shares=shares - shares_to_add,
                                        new_shares=shares_to_add,
                                        price=price,
                                        new_stop_loss=add_info.get('new_stop_loss')
                                    )

                                    updated_state = risk_manager.get_risk_state()
                                    self.data.at[index, 'add_count'] = updated_state['add_count']
                                    self.data.at[index, 'avg_cost'] = updated_state['avg_cost']
                                    self.data.at[index, 'stop_loss_price'] = updated_state['stop_loss']
                                    self.data.at[index, 'stop_profit_price'] = updated_state['stop_profit']
                                    self.data.at[index, 'last_buy_price'] = updated_state['last_buy_price']
                                    prev_trade = current_trade
                                    # 更新交易记录
                                    current_trade.update({
                                        f'第{self.data.at[index, 'add_count']}次加仓日期': self.data.at[index, '日期'],
                                        f'第{self.data.at[index, 'add_count']}次加仓股数': shares_to_add,
                                        f'第{self.data.at[index, 'add_count']}次加仓价格': price,
                                        f'第{self.data.at[index, 'add_count']}次加仓止损价': add_info.get('new_stop_loss'),
                                        '买入总金额': price * shares_to_add + prev_trade.get('买入总金额'),
                                        '买入总佣金': commission + prev_trade.get('买入总佣金'),
                                        '买入总股数': shares_to_add + prev_trade.get('买入总股数'),
                                        '买入总成本': total_cost + prev_trade.get('买入总成本'),
                                        '买入成本价': (total_cost + prev_trade.get('买入总成本')) / (shares_to_add + prev_trade.get('买入总股数')),
                                    })

                            continue
                    # 延迟建仓
                    elif pending == "买入" and row['abnormal_buy_forbidden'] == 0:
                        # 执行买入（延迟建仓）
                        price = row['收盘']
                        cash = self.data.at[index, '可用资金']
                        atr = self.data.at[index, 'ATR']
                        holding = 1
                        pending = ""

                        current_trade = self._do_buy(price, cash, index, atr, risk_manager)

                        continue

                elif signal == '买入':
                    # 实现建仓逻辑
                    if row['abnormal_buy_forbidden'] == 0:
                        price = row['收盘']
                        cash = self.data.at[index, '可用资金']
                        atr = self.data.at[index, 'ATR']
                        holding = 1
                        pending = ""

                        current_trade = self._do_buy(price, cash, index, atr, risk_manager)

                        continue
                    else:
                        pending = "买入"

                elif signal == '卖出' and holding == 1:
                    # 实现卖出逻辑
                    if row['abnormal_sell_forbidden'] == 0:
                        holding = 0
                        pending = ""
                        price = row['收盘']
                        shares = self.data.at[index, 'shares']
                        cash = self.data.at[index, '可用资金']

                        trade_record = self._do_sell(price, shares, cash, index, current_trade)
                        current_trade.update(trade_record)
                        self.trade_records.append(current_trade)
                        current_trade = {}

                        risk_manager.update_risk_state('close')

                        continue
                    else:
                        pending = "卖出"

            # 无风险管理时的交易操作
            else:
                if signal == '':
                    if holding == 1:
                        # 执行卖出（延迟卖出）
                        if pending == "卖出" and row['abnormal_sell_forbidden'] == 0:
                            # 执行卖出逻辑
                            holding = 0
                            pending = ""
                            price = row['收盘']
                            shares = self.data.at[index, 'shares']
                            cash = self.data.at[index, '可用资金']

                            trade_record = self._do_sell(price, shares, cash, index, current_trade)
                            current_trade.update(trade_record)
                            self.trade_records.append(current_trade)
                            current_trade = {}

                            continue

                    # 延迟建仓
                    elif pending == "买入" and row['abnormal_buy_forbidden'] == 0:
                        # 执行建仓逻辑
                        pending = ""
                        holding = 1
                        cash = self.data.at[index, '可用资金']
                        price = row['收盘']

                        current_trade = self._do_buy(price, cash, index)

                        continue

                # 建仓
                elif signal == '买入':
                    if row['abnormal_buy_forbidden'] == 0:
                        # 实现建仓逻辑
                        pending = ""
                        holding = 1
                        cash = self.data.at[index, '可用资金']
                        price = row['收盘']

                        current_trade = self._do_buy(price, cash, index)

                        continue
                    else:
                        pending = "买入"

                # 清仓卖出
                elif signal == '卖出' and holding == 1:
                    if row['abnormal_sell_forbidden'] == 0:
                        # 实现卖出逻辑
                        holding = 0
                        pending = ""
                        price = row['收盘']
                        shares = self.data.at[index, 'shares']
                        cash = self.data.at[index, '可用资金']

                        trade_record = self._do_sell(price, shares, cash, index, current_trade)
                        current_trade.update(trade_record)
                        self.trade_records.append(current_trade)
                        current_trade = {}

                        continue
                    else:
                        pending = "卖出"

        # 指标统一计算
        # 回撤率
        self.data['cum_max'] = self.data['总资产'].cummax()
        mask = self.data['cum_max'] > 0
        self.data.loc[mask, '回撤率'] = 1 - self.data.loc[mask, '总资产'] / self.data.loc[mask, 'cum_max']
        # 策略净值
        self.data['策略净值'] = self.data['总资产'] / float(self.initial_capital)
        # 策略日收益率
        self.data['日收益率'] = self.data['总资产'].pct_change()
        # 策略累计收益率
        self.data['累计收益率'] = self.data['策略净值'] - 1

        return self.calculate_metrics()

    def _do_buy(self,price,cash,index,atr=None,risk_manager=None):
        if risk_manager is not None:
            # 计算建仓股数
            buy_info = risk_manager.calculate_position_size(price, atr, cash)
            shares_to_buy = buy_info.get("shares")
            if shares_to_buy > 0:
                trade_amount = shares_to_buy * price
                commission = max(trade_amount * self.commission_rate, self.min_commission)
                total_cost = trade_amount + commission
                if cash >= total_cost:
                    shares = shares_to_buy
                    cash -= total_cost
                    # 更新字段
                    self.data.at[index, '持仓状态'] = 1
                    self.data.at[index, 'shares'] = shares
                    self.data.at[index, '持仓市值'] = shares * price
                    self.data.at[index, '可用资金'] = cash
                    self.data.at[index, '总资产'] = cash + shares * price
                    self.data.at[index, 'commission'] = commission
                    self.data.at[index, 'stamp_tax'] = 0
                    self.data.at[index, '平均成本'] = total_cost / shares

                    risk_manager.update_risk_state(
                        'open',
                        price=price,
                        stop_loss=buy_info['stop_loss'],
                        stop_profit=buy_info['stop_profit'],
                        atr_value=atr,
                        add_shares_list=buy_info['add_shares_list']
                    )

                    updated_state = risk_manager.get_risk_state()
                    self.data.at[index, 'add_count'] = updated_state['add_count']
                    self.data.at[index, 'avg_cost'] = updated_state['avg_cost']
                    self.data.at[index, 'stop_loss_price'] = updated_state['stop_loss']
                    self.data.at[index, 'stop_profit_price'] = updated_state['stop_profit']
                    self.data.at[index, 'last_buy_price'] = updated_state['last_buy_price']

                return {
                    '建仓日期': self.data.at[index, '日期'],
                    '建仓价': price,
                    '建仓股数': shares_to_buy,
                    '建仓止损价': buy_info.get('stop_loss'),
                    '买入总金额': price * shares_to_buy,
                    '买入总股数': shares_to_buy,
                    '买入总佣金': commission,
                    '买入总成本': total_cost,
                    '买入成本价': total_cost / shares_to_buy,
                }
        else:
            max_shares = int(cash / price / 100) * 100
            while max_shares > 0:
                trade_amount = max_shares * price
                commission = max(trade_amount * self.commission_rate, self.min_commission)
                total_cost = trade_amount + commission

                if cash >= total_cost:
                    shares = max_shares
                    cash -= total_cost
                    # 更新字段
                    self.data.at[index, '持仓状态'] = 1
                    self.data.at[index, 'shares'] = shares
                    self.data.at[index, '持仓市值'] = shares * price
                    self.data.at[index, '可用资金'] = cash
                    self.data.at[index, '总资产'] = cash + shares * price
                    self.data.at[index, 'commission'] = commission
                    self.data.at[index, 'stamp_tax'] = 0
                    self.data.at[index, '平均成本'] = total_cost / shares
                    break
                else:
                    max_shares -= 100

            return {
                '建仓日期': self.data.at[index, '日期'],
                '建仓价': price,
                '建仓股数': max_shares,
                '买入金额': price * max_shares,
                '买入佣金': self.data.at[index, 'commission'],
                '买入总成本': self.data.at[index, '平均成本'] * max_shares,
                '买入成本价': self.data.at[index, '平均成本'],
            }

    def _do_sell(self,price,shares,cash,index,current_trade,exit_reason=None):
        trade_amount = shares * price
        commission = max(trade_amount * self.commission_rate, self.min_commission)
        stamp_tax = trade_amount * self.stamp_tax_rate
        cash += trade_amount - commission - stamp_tax
        # 更新字段
        self.data.at[index, '持仓状态'] = 0
        self.data.at[index, 'shares'] = 0
        self.data.at[index, '持仓市值'] = 0.0
        self.data.at[index, '可用资金'] = cash
        self.data.at[index, '总资产'] = cash
        self.data.at[index, 'exit_reason'] = exit_reason if exit_reason is not None else '卖出信号'
        self.data.at[index, 'commission'] = commission
        self.data.at[index, 'stamp_tax'] = stamp_tax
        self.data.at[index, '平均成本'] = 0.0

        total_cost = current_trade.get('买入总成本') # 买入金额+手续费
        total_received = price * shares - commission - stamp_tax # 卖出金额-手续费-印花税
        pnl = total_received - total_cost
        return_pct = (pnl / total_cost) * 100 if total_cost > 0 else 0

        return {
            '卖出日期': self.data.at[index,'日期'],
            '卖出价': price,
            '卖出金额': price * shares,
            '卖出佣金': commission,
            '印花税': stamp_tax,
            '卖出净收入': total_received,
            '盈亏金额': pnl,
            '收益率': return_pct,
            '交易类型': exit_reason if exit_reason is not None else '卖出信号',
        }

    def calculate_metrics(self) -> Dict:
        """
        计算绩效指标

        Returns:
            绩效指标字典，包含本策略和买入持有策略的指标
        """
        if self.data is None:
            raise ValueError("请先运行回测")

        # 计算本策略指标
        strategy_metrics = self._calculate_strategy_metrics()

        # 计算买入持有指标
        hold_metrics = self._calculate_hold_metrics()

        # 合并指标
        self.metrics = {
            'strategy': strategy_metrics,
            'hold': hold_metrics
        }

        return self.metrics

    def _calculate_strategy_metrics(self) -> Dict:
        """计算本策略绩效指标"""
        total_assets = self.data['总资产'].values
        initial_capital = self.initial_capital

        # 总收益率
        total_return = (total_assets[-1] / initial_capital) - 1

        # 年化收益率
        trading_days = len(self.data[self.data.get('is_suspend')==0])
        years = trading_days / 245
        annual_return = (total_assets[-1] / initial_capital) ** (1 / years) - 1

        # 最大回撤
        max_drawdown = self.data['回撤率'].max()

        # 夏普比率
        daily_returns = self.data['日收益率'].dropna()
        if len(daily_returns) > 1 and daily_returns.std() != 0:
            excess_returns = daily_returns - self.risk_free_rate / 245
            sharpe_ratio = np.sqrt(245) * excess_returns.mean() / daily_returns.std()
        else:
            sharpe_ratio = 0

        # 交易次数
        total_trades = len(self.trade_records)

        # 胜率和盈亏比
        win_rate = 0
        profit_loss_ratio = 0

        if total_trades > 0:
            trades_df = pd.DataFrame(self.trade_records)
            win_trades = trades_df[trades_df['收益率'] > 0]
            loss_trades = trades_df[trades_df['收益率'] < 0]

            win_rate = (len(win_trades) / total_trades) if total_trades > 0 else 0

            # 计算平均盈利
            avg_win = win_trades['收益率'].mean() if len(win_trades) > 0 else 0

            if len(loss_trades) > 0:
                # 有亏损交易
                avg_loss = abs(loss_trades['收益率'].mean())
                profit_loss_ratio = avg_win / avg_loss if avg_loss > 0 else 0
            else:
                # 没有亏损交易，盈亏比设为一个大数或None
                profit_loss_ratio = float('inf') if avg_win > 0 else 0

        return {
            '总收益率': total_return * 100,
            '年化收益率': annual_return * 100,
            '最大回撤': max_drawdown * 100,
            '夏普比率': sharpe_ratio,
            '交易次数': total_trades,
            '胜率': win_rate * 100,
            '盈亏比': profit_loss_ratio
        }

    def _calculate_hold_metrics(self) -> Dict:
        """计算买入持有策略指标"""
        # 创建买入持有策略实例
        hold_strategy = BuyHoldStrategy(
            initial_capital=self.initial_capital,
            commission_rate=self.commission_rate,
            stamp_duty_rate=self.stamp_tax_rate,
            min_commission=self.min_commission
        )

        # 运行回测
        hold_strategy.run_backtest(self.data[['日期', '收盘']].copy())

        # 计算指标
        metrics = hold_strategy.calculate_metrics(risk_free_rate=self.risk_free_rate)

        return {
            '总收益率': metrics['总收益率']*100,
            '年化收益率': metrics['年化收益率']*100,
            '最大回撤': metrics['最大回撤']*100,
            '夏普比率': metrics['夏普比率'],
            '交易次数': 1,
            '胜率': 100 if metrics['总收益率']>0 else 0,
            '盈亏比': ''
        }

    def plot_result(self, save_path: str = None) -> plt.Figure:
        """
        绘制图表

        Args:
            save_path: 图表存放路径

        Returns:
            matplotlib图像对象
        """
        if self.data is None:
            raise ValueError("请先运行回测")

        # 创建图表
        fig, axes = plt.subplots(3, 2, figsize=(18, 16))
        fig.suptitle(f'{self.stock_name} - RSI策略回测结果', fontsize=16, fontweight='bold')

        # 获取信号点
        buy_signals = self.data[self.data['交易信号'] == '买入']
        sell_signals = self.data[self.data['exit_reason'] == '卖出信号']
        stop_signals = self.data[self.data['exit_reason'].isin(['止损', '止盈'])]

        # 1. 价格和买卖点
        ax1 = axes[0, 0]
        ax1.plot(self.data['日期'], self.data['收盘'], label='收盘价', color='blue', linewidth=1)

        ax1.scatter(buy_signals['日期'], buy_signals['收盘'], color='red', marker='^', s=80, label='买入', zorder=5)
        ax1.scatter(sell_signals['日期'], sell_signals['收盘'], color='green', marker='v', s=80, label='卖出', zorder=5)
        ax1.scatter(stop_signals['日期'], stop_signals['收盘'], color='orange', marker='v', s=80, label='止损/止盈',
                    zorder=5)

        ax1.set_title('价格与交易信号')
        ax1.set_ylabel('价格')
        ax1.legend()
        ax1.grid(True, alpha=0.3)

        # 2. RSI指标（加上买卖点标志）
        ax2 = axes[0, 1]
        ax2.plot(self.data['日期'], self.data['RSI'], label=f'RSI({self.rsi_period})', color='purple', linewidth=1.5)
        ax2.axhline(y=self.overbought_threshold, color='red', linestyle='--', alpha=0.5, label='超买线')
        ax2.axhline(y=self.oversold_threshold, color='green', linestyle='--', alpha=0.5, label='超卖线')
        ax2.axhline(y=50, color='gray', linestyle='--', alpha=0.3)

        ax2.fill_between(self.data['日期'], self.overbought_threshold, 100, alpha=0.1, color='red')
        ax2.fill_between(self.data['日期'], 0, self.oversold_threshold, alpha=0.1, color='green')

        # 在RSI图上添加买卖点标志
        ax2.scatter(buy_signals['日期'], buy_signals['RSI'], color='red', marker='^', s=80, label='买入', zorder=5)
        ax2.scatter(sell_signals['日期'], sell_signals['RSI'], color='green', marker='v', s=80, label='卖出', zorder=5)
        ax2.scatter(stop_signals['日期'], stop_signals['RSI'], color='orange', marker='v', s=80, label='止损/止盈', zorder=5)

        ax2.set_title('RSI指标')
        ax2.set_ylabel('RSI')
        ax2.set_ylim(0, 100)
        ax2.legend()
        ax2.grid(True, alpha=0.3)

        # 3. 持仓状态
        ax3 = axes[1, 0]
        ax3.fill_between(self.data['日期'], 0, self.data['持仓状态'] * self.data['持仓状态'].max() / 10,
                         color='blue', alpha=0.5, label='持仓')
        ax3.set_title('持仓状态')
        ax3.set_ylabel('持仓')
        ax3.legend()
        ax3.grid(True, alpha=0.3)

        # 4. 策略净值 vs 买入持有净值
        ax4 = axes[1, 1]

        # 计算买入持有净值
        first_price = self.data.iloc[0]['收盘']
        hold_nav = self.data['收盘'] / first_price

        ax4.plot(self.data['日期'], self.data['策略净值'], label='策略净值', color='red', linewidth=2)
        ax4.plot(self.data['日期'], hold_nav, label='买入持有净值', color='blue', linewidth=2, alpha=0.7)
        ax4.axhline(y=1, color='gray', linestyle='--', alpha=0.5)

        ax4.set_title('策略净值对比')
        ax4.set_ylabel('净值')
        ax4.legend()
        ax4.grid(True, alpha=0.3)

        # 5. 回撤曲线
        ax5 = axes[2, 0]
        ax5.fill_between(self.data['日期'], 0, self.data['回撤率'], color='red', alpha=0.3)
        ax5.plot(self.data['日期'], self.data['回撤率'], color='red', linewidth=1)
        ax5.axhline(y=0, color='black', linestyle='-', linewidth=0.5)

        ax5.set_title('策略回撤')
        ax5.set_ylabel('回撤率 (%)')
        ax5.grid(True, alpha=0.3)

        # 6. 日收益率分布
        ax6 = axes[2, 1]
        daily_returns = self.data['日收益率'].dropna()

        ax6.hist(daily_returns, bins=50, edgecolor='black', alpha=0.7, color='steelblue')
        ax6.axvline(x=0, color='red', linestyle='--', linewidth=1)
        ax6.axvline(x=daily_returns.mean(), color='green', linestyle='--', linewidth=1,
                    label=f'均值: {daily_returns.mean():.3f}%')

        ax6.set_title('日收益率分布')
        ax6.set_xlabel('日收益率 (%)')
        ax6.set_ylabel('频次')
        ax6.legend()
        ax6.grid(True, alpha=0.3)

        # 设置日期格式
        for ax in axes.flat:
            if hasattr(ax, 'xaxis'):
                ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m'))
                ax.xaxis.set_major_locator(mdates.MonthLocator(interval=3))
                ax.tick_params(axis='x', rotation=45)

        plt.tight_layout()

        if save_path:
            plt.savefig(save_path, dpi=300, bbox_inches='tight')

        return fig

    def write_report(self, report_path: str, risk_manager = None):
        """
        编写excel回测报告

        Args:
            report_path: 报告存放路径
        """
        if self.data is None or not self.metrics:
            raise ValueError("请先运行回测并计算指标")

        with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
            # Sheet1: 参数说明
            if risk_manager is not None:
                params_data = {
                    '参数名称': [
                        '初始资金',
                        '佣金费率',
                        '印花税费率',
                        '最低佣金',
                        '止损线',
                        '无风险利率',
                        'RSI周期',
                        '超卖阈值',
                        '超买阈值',
                        '风险管理',
                        '风险预算比率',
                        '加仓比例',
                        '加仓间隔(ATR倍数)',
                        '止损/止盈(ATR倍数)'
                    ],
                    '参数值': [
                        f"{self.initial_capital:,.0f}",
                        f"{self.commission_rate * 100:.3f}%",
                        f"{self.stamp_tax_rate * 100:.2f}%",
                        f"{self.min_commission}",
                        f"{self.stop_loss * 100:.1f}%" if self.stop_loss else '无',
                        f"{self.risk_free_rate * 100:.1f}%",
                        f"{self.rsi_period}",
                        f"{self.oversold_threshold}",
                        f"{self.overbought_threshold}",
                        "已启动",
                        f"{risk_manager.risk_percent * 100:.1f}%",
                        f"{risk_manager.add_ratios}",
                        f"{risk_manager.add_atr_multiple}",
                        f"{risk_manager.stop_atr_multiple}"
                    ]
                }
            else:
                params_data = {
                    '参数名称': [
                        '初始资金',
                        '佣金费率',
                        '印花税费率',
                        '最低佣金',
                        '止损线',
                        '无风险利率',
                        'RSI周期',
                        '超卖阈值',
                        '超买阈值',
                        '风险管理'
                    ],
                    '参数值': [
                        f"{self.initial_capital:,.0f}",
                        f"{self.commission_rate * 100:.3f}%",
                        f"{self.stamp_tax_rate * 100:.2f}%",
                        f"{self.min_commission}",
                        f"{self.stop_loss * 100:.1f}%" if self.stop_loss else '无',
                        f"{self.risk_free_rate * 100:.1f}%",
                        f"{self.rsi_period}",
                        f"{self.oversold_threshold}",
                        f"{self.overbought_threshold}",
                        "未启动"
                    ]
                }
            pd.DataFrame(params_data).to_excel(writer, sheet_name='参数说明', index=False)

            # Sheet2: 绩效指标
            metrics_data = {
                '指标名称': ['总收益率', '年化收益率', '最大回撤', '夏普比率', '交易次数', '胜率', '盈亏比'],
                '本策略指标值': [
                    f"{self.metrics['strategy']['总收益率']:.2f}%",
                    f"{self.metrics['strategy']['年化收益率']:.2f}%",
                    f"{abs(self.metrics['strategy']['最大回撤']):.2f}%",  # 取绝对值
                    f"{self.metrics['strategy']['夏普比率']:.2f}",
                    f"{self.metrics['strategy']['交易次数']}",
                    f"{self.metrics['strategy']['胜率']:.2f}%",
                    f"{self.metrics['strategy']['盈亏比']:.2f}"
                ],
                '买入持有指标值': [
                    f"{self.metrics['hold']['总收益率']:.2f}%",
                    f"{self.metrics['hold']['年化收益率']:.2f}%",
                    f"{abs(self.metrics['hold']['最大回撤']):.2f}%",  # 取绝对值
                    f"{self.metrics['hold']['夏普比率']:.2f}",
                    f"{self.metrics['hold']['交易次数']}",
                    f"{self.metrics['hold']['胜率']:.2f}%",
                    f""
                ]
            }
            pd.DataFrame(metrics_data).to_excel(writer, sheet_name='绩效指标', index=False)

            # Sheet3: 日度数据
            daily_columns = ['日期', '开盘', '收盘', '最高', '最低', '成交量', '成交额', '异常情况', 'RSI', '交易信号',
                             'ATR', '策略净值', '持仓市值', '可用资金', '总资产', '回撤率', '日收益率', '累计收益率',
                             '持仓股数', '止盈/止损', '止损线', '止盈线', '已加仓次数']
            column_mapping = {
                'shares': '持仓股数',
                'exit_reason': '止盈/止损',
                'stop_loss_price': '止损线',
                'stop_profit_price': '止盈线',
                'add_count': '已加仓次数'
            }

            # 复制数据并重命名
            daily_data = self.data.copy()
            daily_data.rename(columns=column_mapping, inplace=True)
            daily_data = daily_data[daily_columns].copy()

            daily_data['日期'] = daily_data['日期'].dt.strftime('%Y-%m-%d')

            # 格式化百分比列
            daily_data['回撤率'] = daily_data['回撤率'].abs().apply(lambda x: f"{x*100:.2f}%")
            daily_data['日收益率'] = daily_data['日收益率'].apply(lambda x: f"{x*100:.2f}%" if pd.notna(x) else "")
            daily_data['累计收益率'] = daily_data['累计收益率'].apply(lambda x: f"{x*100:.2f}%" if pd.notna(x) else "")

            # 格式化其他数值列
            daily_data['收盘'] = daily_data['收盘'].apply(lambda x: f"{x:.2f}")
            daily_data['开盘'] = daily_data['开盘'].apply(lambda x: f"{x:.2f}")
            daily_data['最高'] = daily_data['最高'].apply(lambda x: f"{x:.2f}")
            daily_data['最低'] = daily_data['最低'].apply(lambda x: f"{x:.2f}")
            daily_data['止损线'] = daily_data['止损线'].apply(lambda x: f"{x:.2f}")
            daily_data['止盈线'] = daily_data['止盈线'].apply(lambda x: f"{x:.2f}")
            daily_data['成交量'] = daily_data['成交量'].astype(int)
            daily_data['持仓股数'] = daily_data['持仓股数'].astype(int)
            daily_data['已加仓次数'] = daily_data['已加仓次数'].astype(int)
            daily_data['成交额'] = daily_data['成交额'].apply(lambda x: f"{x:.2f}")
            daily_data['RSI'] = daily_data['RSI'].apply(lambda x: f"{x:.2f}" if pd.notna(x) else "")
            daily_data['策略净值'] = daily_data['策略净值'].apply(lambda x: f"{x:.4f}")
            daily_data['持仓市值'] = daily_data['持仓市值'].apply(lambda x: f"{x:.2f}")
            daily_data['可用资金'] = daily_data['可用资金'].apply(lambda x: f"{x:.2f}")
            daily_data['总资产'] = daily_data['总资产'].apply(lambda x: f"{x:.2f}")

            daily_data.to_excel(writer, sheet_name='日度数据', index=False)

            # Sheet4: 交易记录
            if self.trade_records:
                pd.DataFrame(self.trade_records).to_excel(writer, sheet_name='交易记录', index=False)
            else:
                pd.DataFrame().to_excel(writer, sheet_name='交易记录', index=False)

    def run_complete_analysis(self, file_path: str, output_dir: str, risk_manager = None) -> Dict:
        """
        运行完整分析流程

        Args:
            file_path: 数据文件路径
            output_dir: 存放输出的文件夹

        Returns:
            绩效指标字典
        """
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)

        print("=" * 80)
        print(f"开始RSI策略分析...".center(80))
        print("=" * 80)

        # 加载数据
        self.load_data(file_path)
        print(f"股票代码: {self.stock_code}")
        print(f"数据行数: {len(self.data)}")

        # 数据预处理
        self.preprocess_data()
        print("数据预处理完成")

        # 运行回测
        self.run_backtest(risk_manager)
        print("回测完成")

        # 计算指标
        self.calculate_metrics()
        print("绩效指标计算完成")

        # 打印结果
        self._print_results()

        # 绘制图表
        plot_path = output_path / f"{self.stock_code}_RSI策略图表.png"
        self.plot_result(str(plot_path))
        print(f"图表已保存: {plot_path}")

        # 编写报告
        report_path = output_path / f"{self.stock_code}_RSI策略报告.xlsx"
        self.write_report(str(report_path),risk_manager)
        print(f"报告已保存: {report_path}")

        print("=" * 80)
        print("分析完成！".center(80))
        print("=" * 80)

        return self.metrics

    def _print_results(self):
        """打印回测结果"""
        s = self.metrics['strategy']
        h = self.metrics['hold']

        print("\n" + "=" * 100)
        print(f"{self.stock_code} - RSI策略回测结果".center(100))
        print("=" * 100)

        print(f"\n{'指标':<20} {'本策略':>15} {'买入持有':>15} {'对比':>15}")
        print("-" * 70)
        print(
            f"{'总收益率(%)':<20} {s['总收益率']:>15.2f} {h['总收益率']:>15.2f} {(s['总收益率'] - h['总收益率']):>+15.2f}")
        print(
            f"{'年化收益率(%)':<20} {s['年化收益率']:>15.2f} {h['年化收益率']:>15.2f} {(s['年化收益率'] - h['年化收益率']):>+15.2f}")
        print(
            f"{'最大回撤(%)':<20} {abs(s['最大回撤']):>15.2f} {abs(h['最大回撤']):>15.2f} {(abs(s['最大回撤']) - abs(h['最大回撤'])):>+15.2f}")
        print(
            f"{'夏普比率':<20} {s['夏普比率']:>15.2f} {h['夏普比率']:>15.2f} {(s['夏普比率'] - h['夏普比率']):>+15.2f}")
        print(f"{'交易次数':<20} {s['交易次数']:>15} {h['交易次数']:>15} {'-':>15}")
        print(f"{'胜率(%)':<20} {s['胜率']:>15.2f} {h['胜率']:>15} {'-':>15}")
        print(f"{'盈亏比':<20} {s['盈亏比']:>15.2f} {h['盈亏比']:>15} {'-':>15}")
        print("=" * 100)

    @staticmethod
    def compare_stocks(input_source: Union[str, List[str]],
                       initial_capital: float = 1000000,
                       commission_rate: float = 0.0001,
                       stamp_tax_rate: float = 0.001,
                       min_commission: float = 5,
                       stop_loss: float = None,
                       risk_free_rate: float = 0.03,
                       rsi_period: int = 14,
                       oversold_threshold: int = 30,
                       overbought_threshold: int = 70,
                       output_dir: str = "./output",
                       risk_manager = None) -> pd.DataFrame:
        """
        多股票对比

        Args:
            input_source: 文件夹路径或数据文件路径列表
            initial_capital: 初始资金
            commission_rate: 佣金费率
            stamp_tax_rate: 印花税费率
            min_commission: 最低佣金
            stop_loss: 止损线
            risk_free_rate: 无风险利率
            rsi_period: RSI周期
            oversold_threshold: 超卖阈值
            overbought_threshold: 超买阈值
            output_dir: 输出文件夹

        Returns:
            对比结果数据框
        """
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)

        # 获取所有股票文件
        stock_files = []
        stock_codes = []

        if isinstance(input_source, str):
            folder_path = Path(input_source)
            if not folder_path.exists():
                raise ValueError(f"文件夹不存在: {folder_path}")

            stock_files = list(folder_path.glob("*.csv"))
            for f in stock_files:
                stem = f.stem
                if '_' in stem:
                    parts = stem.split('_')
                    if len(parts) >= 3:
                        stock_codes.append(parts[2])
                    else:
                        stock_codes.append(stem)
                else:
                    stock_codes.append(stem)
        else:
            stock_files = [Path(f) for f in input_source]
            for f in stock_files:
                if not f.exists():
                    raise ValueError(f"文件不存在: {f}")
                stem = f.stem
                if '_' in stem:
                    parts = stem.split('_')
                    if len(parts) >= 3:
                        stock_codes.append(parts[2])
                    else:
                        stock_codes.append(stem)
                else:
                    stock_codes.append(stem)

        print(f"找到 {len(stock_files)} 个股票文件")

        # 执行回测
        results = []

        for i, (file_path, code) in enumerate(zip(stock_files, stock_codes)):
            print(f"\n▶ 正在处理 [{i + 1}/{len(stock_files)}] {code}...")

            try:
                # 创建策略实例
                strategy = RSIStrategy(
                    initial_capital=initial_capital,
                    commission_rate=commission_rate,
                    stamp_tax_rate=stamp_tax_rate,
                    min_commission=min_commission,
                    stop_loss=stop_loss,
                    risk_free_rate=risk_free_rate,
                    rsi_period=rsi_period,
                    oversold_threshold=oversold_threshold,
                    overbought_threshold=overbought_threshold
                )

                # 加载数据
                strategy.load_data(str(file_path))
                strategy.stock_code = code

                # 预处理
                strategy.preprocess_data()

                # 回测
                strategy.run_backtest(risk_manager)

                # 计算指标
                strategy.calculate_metrics()

                # 收集结果
                s = strategy.metrics['strategy']
                h = strategy.metrics['hold']

                # 计算回撤改善：正数表示策略回撤小于买入持有（表现更好）
                # 策略最大回撤（绝对值）越小越好，所以用买入持有回撤减去策略回撤
                strategy_dd_abs = abs(s['最大回撤'])
                hold_dd_abs = abs(h['最大回撤'])
                dd_improve = hold_dd_abs - strategy_dd_abs  # 正数表示改善，负数表示恶化

                results.append({
                    '股票代码': code,
                    '策略总收益率': s['总收益率'],
                    '策略年化收益率': s['年化收益率'],
                    '策略最大回撤率': strategy_dd_abs,  # 取绝对值
                    '策略夏普比率': s['夏普比率'],
                    '策略胜率': s['胜率'],
                    '策略盈亏比': s['盈亏比'],
                    '策略交易次数': s['交易次数'],
                    '买入持有总收益率': h['总收益率'],
                    '买入持有最大回撤率': hold_dd_abs,  # 取绝对值
                    '买入持有夏普比率': h['夏普比率'],
                    '超额总收益率': s['总收益率'] - h['总收益率'],
                    '回撤改善': dd_improve  # 正数表示改善
                })

                print(f"  ✓ 完成 - 收益率: {s['总收益率']:.2f}%")

            except Exception as e:
                print(f"  ✗ 失败: {str(e)}")
                continue

        if not results:
            raise ValueError("没有成功处理任何股票")

        # 创建对比数据框
        comparison = pd.DataFrame(results)

        # 保存明细表 - 格式化数值带%
        comparison_formatted = comparison.copy()
        for col in comparison_formatted.columns:
            if col == '股票代码':
                continue
            elif col in ['策略夏普比率', '买入持有夏普比率', '策略盈亏比']:
                comparison_formatted[col] = comparison_formatted[col].apply(lambda x: f"{x:.2f}")
            elif col == '策略交易次数':
                comparison_formatted[col] = comparison_formatted[col].apply(lambda x: f"{int(x)}")
            elif col in ['策略总收益率', '策略年化收益率', '策略最大回撤率', '策略胜率',
                         '买入持有总收益率', '买入持有最大回撤率', '超额总收益率', '回撤改善']:
                comparison_formatted[col] = comparison_formatted[col].apply(lambda x: f"{x:.2f}%")

        excel_path = output_path / "RSI_多股票对比明细表.xlsx"
        comparison_formatted.to_excel(excel_path, index=False)
        print(f"\n明细表已保存: {excel_path}")

        # 绘制对比图
        fig = RSIStrategy._plot_comparison_charts(comparison)
        chart_path = output_path / "RSI_多股票对比图.png"
        fig.savefig(chart_path, dpi=300, bbox_inches='tight')
        plt.show()
        print(f"对比图已保存: {chart_path}")

        # 控制台打印总体效果
        RSIStrategy._print_comparison_summary(comparison)

        return comparison

    @staticmethod
    def _plot_comparison_charts(comparison: pd.DataFrame) -> plt.Figure:
        """绘制多股票对比图表"""
        fig, axes = plt.subplots(2, 2, figsize=(18, 14))
        fig.suptitle('多股票RSI策略对比分析', fontsize=16, fontweight='bold')

        # a. 最大回撤分布对比直方图
        ax1 = axes[0, 0]
        strategy_dd = comparison['策略最大回撤率'].dropna()
        hold_dd = comparison['买入持有最大回撤率'].dropna()

        bins = np.linspace(min(strategy_dd.min(), hold_dd.min()),
                           max(strategy_dd.max(), hold_dd.max()), 20)

        ax1.hist(strategy_dd, bins=bins, alpha=0.7, color='red', edgecolor='black', label='RSI策略')
        ax1.hist(hold_dd, bins=bins, alpha=0.5, color='blue', edgecolor='black', label='买入持有')

        ax1.axvline(strategy_dd.mean(), color='darkred', linestyle='--', linewidth=2,
                    label=f'策略均值: {strategy_dd.mean():.1f}%')
        ax1.axvline(hold_dd.mean(), color='darkblue', linestyle='--', linewidth=2,
                    label=f'买入持有均值: {hold_dd.mean():.1f}%')

        ax1.set_xlabel('最大回撤 (%)')
        ax1.set_ylabel('股票个数')
        ax1.set_title('a. 最大回撤分布对比')
        ax1.legend()
        ax1.grid(True, alpha=0.3)

        # b. 胜率 vs 盈亏比四象限散点图
        ax2 = axes[0, 1]
        win_rate = comparison['策略胜率'].dropna()
        pl_ratio = comparison['策略盈亏比'].dropna()
        sharpe = comparison['策略夏普比率'].dropna()

        if len(win_rate) > 0:
            win_median = win_rate.median()
            pl_median = pl_ratio.median()

            ax2.axhline(y=pl_median, color='gray', linestyle='--', alpha=0.5)
            ax2.axvline(x=win_median, color='gray', linestyle='--', alpha=0.5)

            scatter = ax2.scatter(win_rate, pl_ratio, c=sharpe, cmap='RdYlGn',
                                  s=80, alpha=0.7, edgecolors='black', linewidth=0.5)

            plt.colorbar(scatter, ax=ax2, label='夏普比率')

        ax2.set_xlabel('胜率 (%)')
        ax2.set_ylabel('盈亏比')
        ax2.set_title('b. 策略胜率 vs 盈亏比分布')
        ax2.grid(True, alpha=0.3)

        # c. 策略收益率 vs 买入持有收益率散点图
        ax3 = axes[1, 0]
        strategy_return = comparison['策略总收益率'].dropna()
        hold_return = comparison['买入持有总收益率'].dropna()
        excess_return = comparison['超额总收益率'].dropna()

        if len(strategy_return) > 0:
            max_val = max(strategy_return.max(), hold_return.max())
            min_val = min(strategy_return.min(), hold_return.min())

            ax3.plot([min_val, max_val], [min_val, max_val], '--', color='gray', alpha=0.5, linewidth=2,
                     label='收益率相等')

            scatter = ax3.scatter(hold_return, strategy_return, c=excess_return, cmap='RdYlGn',
                                  s=80, alpha=0.7, edgecolors='black', linewidth=0.5)

            plt.colorbar(scatter, ax=ax3, label='超额收益(%)')

        ax3.set_xlabel('买入持有收益率 (%)')
        ax3.set_ylabel('策略收益率 (%)')
        ax3.set_title('c. 策略 vs 买入持有收益率对比')
        ax3.grid(True, alpha=0.3)

        # d. 超额收益 vs 回撤改善四象限散点图
        ax4 = axes[1, 1]
        excess_return = comparison['超额总收益率'].dropna()
        dd_improve = comparison['回撤改善'].dropna()
        sharpe = comparison['策略夏普比率'].dropna()

        if len(excess_return) > 0:
            ax4.axhline(y=0, color='gray', linestyle='--', alpha=0.5)
            ax4.axvline(x=0, color='gray', linestyle='--', alpha=0.5)

            scatter = ax4.scatter(excess_return, dd_improve, c=sharpe, cmap='RdYlGn',
                                  s=80, alpha=0.7, edgecolors='black', linewidth=0.5)

            plt.colorbar(scatter, ax=ax4, label='夏普比率')

        ax4.set_xlabel('超额收益 (%)')
        ax4.set_ylabel('回撤改善 (%)')
        ax4.set_title('d. 超额收益 vs 回撤改善分布')
        ax4.grid(True, alpha=0.3)

        plt.tight_layout()
        return fig

    @staticmethod
    def _print_comparison_summary(comparison: pd.DataFrame):
        """打印多股票对比总结"""
        print("\n" + "=" * 100)
        print("策略总体效果".center(100))
        print("=" * 100)

        # 跑赢基准比例
        beat_count = (comparison['超额总收益率'] > 0).sum()
        beat_pct = beat_count / len(comparison) * 100
        print(f"📊 跑赢基准比例: {beat_pct:.1f}% ({beat_count}/{len(comparison)})")

        # 回撤改善比例（回撤改善 > 0 表示策略回撤更小）
        improve_count = (comparison['回撤改善'] > 0).sum()
        improve_pct = improve_count / len(comparison) * 100
        print(f"📉 回撤改善比例: {improve_pct:.1f}% ({improve_count}/{len(comparison)})")

        # 平均回撤对比 - 取正数
        avg_strategy_dd = comparison['策略最大回撤率'].mean()
        avg_hold_dd = comparison['买入持有最大回撤率'].mean()
        print(f"📈 策略平均回撤 vs 买入持有平均回撤：{avg_strategy_dd:.1f}% vs {avg_hold_dd:.1f}%")

        # 平均夏普对比
        avg_strategy_sharpe = comparison['策略夏普比率'].mean()
        avg_hold_sharpe = comparison['买入持有夏普比率'].mean()
        print(f"⚡ 策略平均夏普 vs 买入持有平均夏普：{avg_strategy_sharpe:.2f} vs {avg_hold_sharpe:.2f}")

        # 平均指标
        avg_return = comparison['策略总收益率'].mean()
        avg_trades = comparison['策略交易次数'].mean()
        avg_win_rate = comparison['策略胜率'].mean()
        avg_pl_ratio = comparison['策略盈亏比'].mean()

        print(
            f"🎯 平均总收益率: {avg_return:.1f}%, 平均交易次数: {avg_trades:.0f}, 平均胜率: {avg_win_rate:.1f}%, 平均盈亏比: {avg_pl_ratio:.2f}")
        print("=" * 100)

    @staticmethod
    def compare_strategies(file_path: str,
                           param_combinations: List[Tuple[int, int, int]],
                           initial_capital: float = 1000000,
                           commission_rate: float = 0.0001,
                           stamp_tax_rate: float = 0.001,
                           min_commission: float = 5,
                           stop_loss: float = None,
                           risk_free_rate: float = 0.03,
                           output_dir: str = "./output") -> pd.DataFrame:
        """
        比较同一只股票，使用不同参数组合的RSI策略效果

        Args:
            file_path: 数据文件路径
            param_combinations: 参数组合列表，每个元素为 (rsi_period, oversold_threshold, overbought_threshold)
            initial_capital: 初始资金
            commission_rate: 佣金费率
            stamp_tax_rate: 印花税费率
            min_commission: 最低佣金
            stop_loss: 止损线
            risk_free_rate: 无风险利率
            output_dir: 输出文件夹

        Returns:
            对比结果数据框
        """
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)

        # 获取股票代码
        file_stem = Path(file_path).stem
        if '_' in file_stem:
            parts = file_stem.split('_')
            if len(parts) >= 3:
                stock_code = parts[2]
            else:
                stock_code = file_stem
        else:
            stock_code = file_stem

        print("=" * 100)
        print(f"RSI参数对比分析 - 股票: {stock_code}".center(100))
        print("=" * 100)

        print(f"\n待测试的RSI参数组合 ({len(param_combinations)}组):")
        for i, params in enumerate(param_combinations):
            print(f"  组合{i + 1}: RSI周期={params[0]}, 超卖阈值={params[1]}, 超买阈值={params[2]}")

        # 执行回测
        results = []
        nav_data = {}

        for i, params in enumerate(param_combinations):
            rsi_period, oversold, overbought = params

            print(f"\n▶ 测试组合{i + 1}: RSI周期={rsi_period}, 超卖={oversold}, 超买={overbought}")

            try:
                # 创建策略实例
                strategy = RSIStrategy(
                    initial_capital=initial_capital,
                    commission_rate=commission_rate,
                    stamp_tax_rate=stamp_tax_rate,
                    min_commission=min_commission,
                    stop_loss=stop_loss,
                    risk_free_rate=risk_free_rate,
                    rsi_period=rsi_period,
                    oversold_threshold=oversold,
                    overbought_threshold=overbought
                )

                # 加载数据
                strategy.load_data(file_path)

                # 预处理
                strategy.preprocess_data()

                # 回测
                strategy.run_backtest()

                # 计算指标
                strategy.calculate_metrics()

                # 存储净值数据
                nav_key = f"策略{i + 1}({rsi_period},{oversold},{overbought})"
                nav_data[nav_key] = strategy.data[['日期', '策略净值']].copy()

                # 收集结果
                s = strategy.metrics['strategy']
                h = strategy.metrics['hold']

                # 计算回撤改善：正数表示策略回撤小于买入持有（表现更好）
                strategy_dd_abs = abs(s['最大回撤'])
                hold_dd_abs = abs(h['最大回撤'])
                dd_improve = hold_dd_abs - strategy_dd_abs  # 正数表示改善，负数表示恶化

                results.append({
                    '参数组合': nav_key,
                    'rsi_period': rsi_period,
                    'oversold': oversold,
                    'overbought': overbought,
                    '总收益率': s['总收益率'],
                    '年化收益率': s['年化收益率'],
                    '最大回撤率': strategy_dd_abs,  # 取绝对值
                    '夏普比率': s['夏普比率'],
                    '胜率': s['胜率'],
                    '盈亏比': s['盈亏比'],
                    '交易次数': s['交易次数'],
                    '超额收益': s['总收益率'] - h['总收益率'],
                    '回撤改善': dd_improve,  # 正数表示改善
                    'strategy': strategy
                })

                print(f"  ✓ 完成 - 收益率: {s['总收益率']:.2f}%, 夏普: {s['夏普比率']:.2f}")

            except Exception as e:
                print(f"  ✗ 失败: {str(e)}")
                continue

        if not results:
            raise ValueError("没有成功测试任何参数组合")

        # 获取买入持有数据
        base_strategy = RSIStrategy(initial_capital=initial_capital)
        base_strategy.load_data(file_path)
        base_strategy.preprocess_data()
        base_strategy.run_backtest()
        base_strategy.calculate_metrics()
        hold_return = base_strategy.metrics['hold']['总收益率']
        hold_dd = abs(base_strategy.metrics['hold']['最大回撤'])  # 取绝对值
        hold_sharpe = base_strategy.metrics['hold']['夏普比率']
        hold_nav = base_strategy.data[['日期', '收盘']].copy()
        hold_nav['策略净值'] = hold_nav['收盘'] / hold_nav.iloc[0]['收盘']

        # 创建对比数据框
        comparison = pd.DataFrame(results)

        # 添加买入持有作为最后一行
        hold_row = {
            '参数组合': '买入持有',
            '总收益率': hold_return,
            '年化收益率': base_strategy.metrics['hold']['年化收益率'],
            '最大回撤率': hold_dd,  # 已取绝对值
            '夏普比率': hold_sharpe,
            '胜率': 100 if hold_return > 0 else 0,
            '盈亏比': None,
            '交易次数': 1,
            '超额收益': 0,
            '回撤改善': 0
        }
        comparison = pd.concat([comparison, pd.DataFrame([hold_row])], ignore_index=True)

        # 保存详细数据 - 格式化数值带%
        detail_columns = ['参数组合', '总收益率', '年化收益率', '最大回撤率', '夏普比率',
                          '胜率', '盈亏比', '交易次数', '超额收益', '回撤改善']
        detail_df = comparison[detail_columns].copy()

        # 格式化数值
        for col in detail_df.columns:
            if col == '参数组合':
                continue
            elif col in ['夏普比率', '盈亏比']:
                detail_df[col] = detail_df[col].apply(lambda x: f"{x:.2f}" if pd.notna(x) else "-")
            elif col == '交易次数':
                detail_df[col] = detail_df[col].apply(lambda x: f"{int(x)}" if pd.notna(x) else "-")
            elif col in ['总收益率', '年化收益率', '最大回撤率', '胜率', '超额收益', '回撤改善']:
                detail_df[col] = detail_df[col].apply(lambda x: f"{x:.2f}%" if pd.notna(x) else "-")

        excel_path = output_path / f"{stock_code}_RSI参数对比详细数据.xlsx"
        detail_df.to_excel(excel_path, index=False)
        print(f"\n详细数据已保存: {excel_path}")

        # 绘制净值曲线
        fig, ax = plt.subplots(figsize=(14, 8))

        # 绘制买入持有曲线
        ax.semilogy(hold_nav['日期'], hold_nav['策略净值'],
                    color='black', linewidth=2, linestyle='--', label='买入持有', alpha=0.7)

        # 为不同参数组合绘制曲线
        colors = plt.cm.tab10(np.linspace(0, 1, len(results)))

        for i, (r, color) in enumerate(zip(results, colors)):
            nav_df = nav_data[r['参数组合']]
            ax.semilogy(nav_df['日期'], nav_df['策略净值'],
                        color=color, linewidth=1.5, label=r['参数组合'], alpha=0.8)

        ax.set_xlabel('日期')
        ax.set_ylabel('净值 (对数坐标)')
        ax.set_title(f'不同RSI参数策略净值对比 - {stock_code}')
        ax.legend(loc='upper left', fontsize=9)
        ax.grid(True, alpha=0.3)

        # 设置日期格式
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m'))
        ax.xaxis.set_major_locator(mdates.MonthLocator(interval=3))
        ax.tick_params(axis='x', rotation=45)

        ax.axhline(y=1, color='gray', linestyle='-', linewidth=0.5, alpha=0.5)

        plt.tight_layout()

        chart_path = output_path / f"{stock_code}_RSI参数对比净值曲线.png"
        plt.savefig(chart_path, dpi=300, bbox_inches='tight')
        plt.show()
        print(f"净值曲线已保存: {chart_path}")

        return comparison

class BollStrategy:
    def __init__(self,data_path,min_commission=5,commission=0.0001, stamp_tax_rate=0.001,
                 risk_percent=0.02,add_ratios = [0.4, 0.3, 0.3],add_atr_multiple = 1.0,
                 stop_atr_multiple = 2.0,initial_capital=1000000,risk_free_rate=0.03,
                 boll_period=20,atr_period=14):
        self.datapath = data_path
        self.data = None
        self.data_dict = None # pandasdata数据线字典，默认为OHLCV,再preprocess函数中配置了指标对应字段
        self.stamp_tax_rate = stamp_tax_rate
        self.min_commission = min_commission
        self.commission = commission
        self.stamp_tax_rate = stamp_tax_rate
        self.risk_percent = risk_percent # 风险预算比率
        self.add_ratios = add_ratios # 加仓比率分配
        self.add_atr_multiple = add_atr_multiple # 加仓间隔(ATR倍数)
        self.stop_atr_multiple = stop_atr_multiple # 止损/止盈(ATR倍数)
        self.initial_capital = initial_capital
        self.buyandhold_cerebro = bt.Cerebro()
        self.use_risk_manager = True
        self.risk_free_rate = risk_free_rate
        self.boll_period = boll_period
        self.atr_period = atr_period
        self.stock_code = None
        self.PandasData = []
        self.result = []
        self.buyandhold_result = []

    def preprocess_data(self):
        with open(self.datapath, 'rb') as f:
            result = chardet.detect(f.read(10000))
            print(result)
        df = pd.read_csv(self.datapath, encoding=result['encoding'])
        print(df.columns)
        # 提取股票代码
        code_match = re.search(r'df_pre_(\d+)', self.datapath)
        if code_match:
            self.stock_code = code_match.group(1)
        else:
            self.stock_code = "XXXX"

        df['日期'] = pd.to_datetime(df['日期'])
        df['is_suspend'] = df['异常情况'].apply(lambda x: 1 if x == '停牌' else 0)
        df[['BB_Upper_Series', 'BB_Middle_Series', 'BB_Lower_Series']] = TechnicalIndicators.calculate_boll(df,close_col='收盘',suspend_col="is_suspend", period=self.boll_period)
        df['ATR'] = TechnicalIndicators.calculate_atr(df, suspend_col="is_suspend", period=self.atr_period)
        self.data = df
        self.data_dict = {
            'bb_upper': 'BB_Upper_Series',
            'bb_middle': 'BB_Middle_Series',
            'bb_lower': 'BB_Lower_Series',
            'atr': 'ATR'
        }
        self.PandasData.append(BaseTool.LoadAsPandasData(self.data, self.data_dict))

    def set_position_sizer(self):
        # 保存类的引用，而不是实例
        self.SizerClass = MyRMSizer
        self.sizer_params = {
            'risk_percent': self.risk_percent,
            'add_ratios': self.add_ratios,
            'add_atr_multiple': self.add_atr_multiple,
            'stop_atr_multiple': self.stop_atr_multiple
        }

    def _set_commission(self,broker):
        comminfo = MyCommInfo(perc=self.commission,min_commission=self.min_commission)
        broker.addcommissioninfo(comminfo)

    def run_backtest(self,plot=True,use_risk_manager=True):
        self.use_risk_manager = use_risk_manager
        class MyStrategy(bt.Strategy):
            params = (('use_risk_manager', True),)
            def __init__(self):
                self.order = None
                self.trades = []

            def next(self):
                if self.order is None:
                    if self.p.use_risk_manager == True:
                        # ========== 有持仓的情况 ==========
                        if self.position:
                            # ========== 止损止盈检查 ==========
                            should_exit, reason = self.sizer.check_stop(self.data)
                            if should_exit:
                                self.order = self.sell()
                                self.order.addinfo(trade_signal=3)
                                self.order.addinfo(sell_reason=reason)
                                print("止盈止损")
                                return
                            # ========== 加仓逻辑 ==========
                            if self.sizer.is_can_add(self.data):
                                self.order = self.buy()  # Sizer会自动判断加仓数量
                                self.order.addinfo(trade_signal=2)
                                print("加仓")
                                return
                            # ========== 信号卖出逻辑 ==========
                            if self.data.close[-1] > self.data.bb_upper[-1] and self.data.close[0] < self.data.bb_upper[0]:
                                self.order = self.sell()
                                self.order.addinfo(trade_signal=4)
                                self.order.addinfo(sell_reason='信号卖出')
                                print("卖出信号")
                                return

                        # ========== 建仓逻辑 ==========
                        if not self.position:
                            if self.data.close[-1] < self.data.bb_lower[-1] and self.data.close[0] > self.data.bb_lower[0] and (self.data.bb_upper[0]-self.data.bb_lower[0])/self.data.bb_middle[0]>=0.02:

                                self.order = self.buy()  # Sizer会自动计算数量
                                self.order.addinfo(trade_signal=1)
                                print("建仓")

                    else:
                        if self.position:
                            if self.data.close[-1] > self.data.bb_upper[-1] and self.data.close[0] < self.data.bb_upper[0]:
                                self.order = self.sell()
                                self.order.addinfo(trade_signal=4)
                                print("卖出信号")
                            return
                        else:
                            if self.data.close[-1] < self.data.bb_lower[-1] and self.data.close[0] > self.data.bb_lower[
                                0] and (self.data.bb_upper[0] - self.data.bb_lower[0]) / self.data.bb_middle[0] >= 0.02:
                                self.order = self.buy()  # Sizer会自动计算数量
                                self.order.addinfo(trade_signal=1)
                                print("建仓")

            def log(self, txt, dt=None, doprint=True):
                if doprint:
                    dt = dt or self.datas[0].datetime.date(0)
                    print(f'{dt.isoformat()} {txt}')

            def notify_order(self, order):
                self.observers.getbyname('myobserver').notify_order(order)
                if order.status in [order.Submitted, order.Accepted]:
                    return

                if order.status in [order.Completed]:
                    if order.isbuy():
                        self.log(
                            f'发生一笔买入，价格：{order.executed.price:.6f}，股数：{order.executed.size:.0f}，花费：{order.executed.value:.2f}，手续费：{order.executed.comm:.2f}')
                        self.buyprice = order.executed.price
                        self.buycomm = order.executed.comm
                    else:
                        sell_income = abs(order.executed.size) * order.executed.price
                        self.log(
                            f'发生一笔卖出，价格：{order.executed.price:.6f}，股数：{order.executed.size:.0f}，收入：{sell_income:.2f}，手续费：{order.executed.comm:.2f}，印花税：{order.stamp_tax:.2f}')
                    self.bars_executed = len(self)
                    self.analyzers.lnE.notify_order(order)
                    if self.p.use_risk_manager == True:
                        self.sizer.notify_order(order)
                    # 获取风险状态并存放在order中
                    risk_status = self.sizer.get_risk_state()
                    self.order.addinfo(risk_status=risk_status)

                elif order.status in [order.Canceled, order.Margin, order.Rejected]:
                    self.log(f'订单失败:{order.getstatusname()}')
                    if self.p.use_risk_manager == True:
                        self.sizer.notify_order(order)

                self.order = None

            def notify_trade(self, trade):
                if not trade.isclosed:
                    return
                trade.stamp_tax = self.broker._current_stamp_tax
                self.broker._current_stamp_tax = 0
                self.log(f'交易利润：{trade.pnl:.2f}，手续费：{trade.commission:.2f}，印花税：{trade.stamp_tax:.2f}，净利润：{trade.pnl - trade.commission-trade.stamp_tax:.2f}')
                self.trades.append(trade)

        class BuyAndHoldStrategy(bt.Strategy):
            def __init__(self):
                self.bought = False
                self.order = None

            def next(self):
                if not self.bought and len(self.data) > 0:
                    self.order = self.buy()
                    self.order.addinfo(trade_signal=1)

            def notify_order(self, order):
                self.observers.getbyname('myobserver').notify_order(order)
                if order.status == order.Completed:
                    self.bought = True
                    self.analyzers.lnE.notify_order(order)

            def stop(self):
                final_value = self.broker.getvalue()
                print(
                    f"买入持有策略: 最终资产 {final_value:.2f}, 收益率: {(final_value - self.broker.startingcash) / self.broker.startingcash * 100:.2f}%")

        for i in range(len(self.PandasData)):
            cerebro = bt.Cerebro(tradehistory=True)
            cerebro.adddata(self.PandasData[i])
            cerebro.addstrategy(MyStrategy,use_risk_manager=self.use_risk_manager)
            cerebro.broker = MyBroker(stamp_tax_rate=self.stamp_tax_rate)
            self._set_commission(cerebro.broker)
            if use_risk_manager==True:
                cerebro.addsizer(self.SizerClass,**self.sizer_params)
            else:
                cerebro.addsizer(MyAllInSizer)
            cerebro.broker.setcash(self.initial_capital)
            self._add_analyzers(cerebro)
            self._add_observers(cerebro)
            result = cerebro.run()
            self.result.append(result[0])
            if plot:
                cerebro.plot()
        print("主策略运行结果:", self.result)

        # ========== 买入持有策略运行 ==========
        for i in range(len(self.PandasData)):
            buyandhold_cerebro = bt.Cerebro()
            buyandhold_cerebro.adddata(self.PandasData[i])
            buyandhold_cerebro.addstrategy(BuyAndHoldStrategy)
            buyandhold_cerebro.broker = MyBroker(stamp_tax_rate=self.stamp_tax_rate)
            self._set_commission(cerebro.broker)
            buyandhold_cerebro.addsizer(MyAllInSizer)
            buyandhold_cerebro.broker.setcash(self.initial_capital)
            self._add_analyzers(buyandhold_cerebro)
            self._add_observers(buyandhold_cerebro)
            buyandhold_result = buyandhold_cerebro.run()
            self.buyandhold_result.append(buyandhold_result[0])

    def _add_analyzers(self, cerebro):
        """统一添加分析器"""
        cerebro.addanalyzer(bt.analyzers.Returns, _name='returns')
        cerebro.addanalyzer(bt.analyzers.AnnualReturn, _name='annual_return')
        cerebro.addanalyzer(bt.analyzers.SharpeRatio_A, _name='sharpe',
                            timeframe=bt.TimeFrame.Days,
                            riskfreerate=self.risk_free_rate,
                            factor=252)
        cerebro.addanalyzer(bt.analyzers.TimeDrawDown, _name='drawdown',
                            timeframe=bt.TimeFrame.Days)
        cerebro.addanalyzer(bt.analyzers.TradeAnalyzer, _name='trades')
        cerebro.addanalyzer(SignalEffectiveness, _name='lnE')

    def _get_metrics(self, result, strategy_name=""):
        """从回测结果中提取指标"""
        metrics_list = []
        for strat in result:
            metrics = {}

            # 打印基本信息
            print(f"\n{'=' * 30} {strategy_name} {'=' * 30}")
            print(f"最终现金: {strat.broker.getcash():.2f}")
            print(f"最终资产: {strat.broker.getvalue():.2f}")
            print(f"初始资金: {strat.broker.startingcash:.2f}")

            # 手动计算总收益率
            total_return = (strat.broker.getvalue() - strat.broker.startingcash) / strat.broker.startingcash
            metrics['总收益率'] = total_return * 100
            metrics['年化收益率'] = 0  # 先设默认值

            # 收益率分析器
            returns = strat.analyzers.returns.get_analysis()
            if returns:
                metrics['总收益率'] = returns.get('rtot', 0) * 100
                metrics['年化收益率'] = returns.get('rnorm100', 0)

            # 夏普比率
            sharpe = strat.analyzers.sharpe.get_analysis()
            metrics['夏普比率'] = sharpe.get('sharperatio', 0) if sharpe.get('sharperatio') else 0

            # 最大回撤
            drawdown = strat.analyzers.drawdown.get_analysis()
            metrics['最大回撤(%)'] = drawdown.get('maxdrawdown', 0)
            metrics['回撤修复天数'] = drawdown.get('maxdrawdownperiod', 0)

            # 交易统计
            trades = strat.analyzers.trades.get_analysis()

            # 使用 get() 方法安全获取值
            total_dict = trades.get('total', {})
            metrics['交易次数'] = total_dict.get('total', 0) if isinstance(total_dict, dict) else 0

            # 获取盈亏统计
            won_dict = trades.get('won', {})
            lost_dict = trades.get('lost', {})

            won_total = won_dict.get('total', 0) if isinstance(won_dict, dict) else 0
            lost_total = lost_dict.get('total', 0) if isinstance(lost_dict, dict) else 0
            total_trades = metrics['交易次数']

            if total_trades > 0:
                metrics['胜率(%)'] = (won_total / total_trades) * 100
            else:
                metrics['胜率(%)'] = 0

            # 盈亏比
            won_pnl_dict = won_dict.get('pnl', {}) if isinstance(won_dict, dict) else {}
            lost_pnl_dict = lost_dict.get('pnl', {}) if isinstance(lost_dict, dict) else {}

            won_pnl = won_pnl_dict.get('total', 0) if isinstance(won_pnl_dict, dict) else 0
            lost_pnl = abs(lost_pnl_dict.get('total', 0)) if isinstance(lost_pnl_dict, dict) else 0

            metrics['盈亏比'] = (won_pnl / lost_pnl) if lost_pnl > 0 else 0

            # 信号有效性
            try:
                lnE = strat.analyzers.lnE.get_analysis()
                if lnE and isinstance(lnE, dict):
                    avg_lnE = lnE.get('avg_lnE', 0)
                    if avg_lnE is None or math.isinf(avg_lnE):
                        metrics['平均ln(E)'] = 0
                    else:
                        metrics['平均ln(E)'] = avg_lnE
                else:
                    metrics['平均ln(E)'] = 0
            except Exception as e:
                print(f"获取ln(E)失败: {e}")
                metrics['平均ln(E)'] = 0

            # 手动计算卡玛比率
            annual_return = metrics.get('年化收益率', 0) / 100
            max_drawdown = metrics.get('最大回撤(%)', 0) / 100
            if max_drawdown > 0:
                metrics['卡玛比率'] = annual_return / max_drawdown
            else:
                metrics['卡玛比率'] = 0

            metrics_list.append(metrics)

        return metrics_list

    def _print_comparison(self):
        """打印对比报告"""
        print("\n" + "=" * 80)
        print("回测结果对比报告")
        print("=" * 80)

        # 获取指标
        strategy_metrics = self._get_metrics(self.result,"主策略")[0]
        bh_metrics = self._get_metrics(self.buyandhold_result,"买入持有策略")[0]

        # 打印表格
        print(f"\n{'指标':<20} {'策略':>15} {'买入持有':>15} {'超额收益':>15}")
        print("-" * 65)

        comparisons = [
            ('总收益率(%)', '总收益率', '总收益率'),
            ('年化收益率(%)', '年化收益率', '年化收益率'),
            ('夏普比率', '夏普比率', '夏普比率'),
            ('卡玛比率', '卡玛比率', '卡玛比率'),
            ('最大回撤(%)', '最大回撤(%)', '最大回撤(%)'),
            ('回撤修复天数', '回撤修复天数', '回撤修复天数'),
            ('交易次数', '交易次数', '交易次数'),
            ('胜率(%)', '胜率(%)', '胜率(%)'),
            ('盈亏比', '盈亏比', '盈亏比'),
            ('平均ln(E)', '平均ln(E)', '平均ln(E)'),
        ]

        for name, skey, bhkey in comparisons:
            s_val = strategy_metrics.get(skey, 0)
            bh_val = bh_metrics.get(bhkey, 0)

            if '回撤' in name and s_val != 0:
                # 回撤越小越好，超额 = 基准 - 策略
                excess = bh_val - s_val
            elif name in ['交易次数', '回撤修复天数']:
                excess = s_val - bh_val
            else:
                # 其他指标越大越好
                excess = s_val - bh_val

            # 格式化输出
            if '比率' in name or '夏普' in name or '卡玛' in name or '盈亏比' in name:
                print(f"{name:<20} {s_val:>15.2f} {bh_val:>15.2f} {excess:>15.2f}")
            elif '天数' in name:
                print(f"{name:<20} {s_val:>15.0f} {bh_val:>15.0f} {excess:>15.0f}")
            elif '交易次数' in name:
                print(f"{name:<20} {s_val:>15.0f} {bh_val:>15.0f} {excess:>15.0f}")
            else:
                print(f"{name:<20} {s_val:>15.2f} {bh_val:>15.2f} {excess:>15.2f}")

        print("=" * 80)

    def _add_observers(self,cerebro):
        """统一添加观察器"""
        cerebro.addobserver(MyObserver, _name='myobserver')
        cerebro.addobserver(bt.observers.DrawDown)
        cerebro.addobserver(bt.observers.TimeReturn)

    def get_observer_data(self, observer_name):
        """获取观察器数据（使用 array 属性，效率更高）"""
        observer = self.result[0].observers.getbyname(observer_name)
        if observer is None:
            return None
        # 观察器的第一个line名称
        line_names = observer.lines.getlinealiases()

        dates = self.result[0].data.datetime.array[:].tolist()
        datetime_dates = [bt.num2date(d) for d in dates]
        date_only = [d.date() for d in datetime_dates]
        data_dict = {'date': date_only}

        num = 0
        observer_data = []
        for line_name in line_names:
            observer_data.append(observer.lines[num].array[:len(observer.lines[num])].tolist())
            # 使用 array 属性直接获取底层数组
            data_dict[line_name] = observer_data[num]
            num += 1

        return pd.DataFrame(data_dict)

    def get_trade_data(self):
        TD_strategies = []
        # 遍历所有策略
        for i in range(len(self.result)):
            trades = self.result[i].trades
            TD_trades = []
            # 遍历单个策略内部所有交易
            for j in range(len(trades)):
                trade_datas = {}
                hist_first = trades[j].history[0]
                hist_last = trades[j].history[-1]
                # trade中第一个订单对象
                order_first = hist_first.event.order
                # trade中最后一个订单对象
                order_last = hist_last.event.order

                # 计算建仓日期
                open_date = bt.num2date(trades[j].dtopen).date()
                trade_datas["建仓日期"] = open_date.strftime('%Y-%m-%d')

                # 计算建仓价
                open_price = round(order_first.executed.price, 4)
                trade_datas["建仓价格"] = open_price

                # 计算建仓股数
                open_size = abs(order_first.executed.size)
                trade_datas["建仓股数"] = open_size

                # 计算建仓止损价
                open_stoploss_price = round(order_first.info.get('risk_status').get("stop_loss"), 4)
                trade_datas["建仓止损价"] = open_stoploss_price

                # 计算卖出日期
                close_date = bt.num2date(trades[j].dtclose).date()
                trade_datas["卖出日期"] = close_date

                # 计算卖出价格
                close_price = round(order_last.executed.price, 4)
                trade_datas["卖出价格"] = close_price

                # 计算总股数
                total_size = abs(order_last.executed.size)
                trade_datas["总股数"] = total_size

                # 计算每一笔trade的平均买入成本、加仓相关数据
                hist_cost = 0
                hist_size = 0
                # 对于除了最后一笔（卖出）外的所有历史订单进行循环处理
                for k in range(len(trades[j].history) - 1):
                    order = trades[j].history[k].event.order
                    # 每笔买入订单的成本
                    cost = order.executed.price * order.executed.size + order.executed.comm
                    # 每笔买入订单的股数
                    size = order.executed.size
                    hist_cost += cost
                    hist_size += size
                    # 对于每一笔加仓订单循环处理（排除k=0的第一笔建仓订单）
                    if k > 0:
                        add_date = bt.num2date(order.executed.dt).date() # 加仓日期
                        add_price = round(order.executed.price, 4) # 加仓价格
                        add_size = order.executed.size # 加仓数量
                        add_stoploss_price = round(order.info.get('risk_status').get("stop_loss"), 4) # 加仓止损价
                        trade_datas[f"第{k}次加仓日期"] = add_date
                        trade_datas[f"第{k}次加仓价格"] = add_price
                        trade_datas[f"第{k}次加仓股数"] = add_size
                        trade_datas[f"第{k}次加仓止损价"] = add_stoploss_price

                # 计算每一笔trade的平均成本
                avg_cost = round(hist_cost / hist_size, 4)
                trade_datas["平均成本"] = avg_cost

                # 计算盈亏金额
                net_profit = round(trades[j].pnl - trades[j].commission - trades[j].stamp_tax, 2)
                trade_datas["盈亏金额"] = net_profit

                # 计算收益率
                total_return = round(net_profit / (avg_cost * order_last.executed.size), 4)
                trade_datas["收益率"] = total_return

                # 计算持仓天数
                hold_days = (bt.num2date(trades[j].dtclose).date() - bt.num2date(trades[j].dtopen).date()).days
                trade_datas["持仓天数"] = hold_days

                # 计算离场类型
                sell_reason = order_last.info.get('sell_reason')
                trade_datas["离场类型"] = sell_reason

                TD_trades.append(trade_datas)
            TD_strategies.append(TD_trades)
        return TD_strategies

    def write_report(self, output_folder):
        """编写单个股票回测报告"""
        if not self.result:
            raise ValueError("请先运行回测")
        # 创建输出文件夹
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        # 定义Excel文件路径
        excel_path = os.path.join(output_folder, f'{self.stock_code}_布林带_回测报告.xlsx')

        # 准备参数说明
        if self.use_risk_manager == True:
            params_data = {
                '参数名称': [
                    '初始资金', '风险管理器', '风险预算比例', '加仓比例分配', '加仓间隔(ATR倍数)',
                    '佣金费率', '印花税率', '无风险利率',
                    '止损止盈设置(ATR倍数)', '布林带周期', 'ATR周期'
                ],
                '参数值': [
                    f'{self.initial_capital:,.0f}元',
                    '已启用',
                    f'{self.risk_percent * 100:.1f}%',
                    f'{self.add_ratios}',
                    f'{self.add_atr_multiple}倍ATR',
                    f'万{self.commission * 10000:.1f} ({self.min_commission}元起)',
                    f'{self.stamp_tax_rate * 1000:.1f}‰',
                    f'{self.risk_free_rate * 100:.1f}%',
                    f'{self.stop_atr_multiple}倍ATR',
                    str(self.boll_period),
                    str(self.atr_period)
                ]
            }
        else:
            params_data = {
                '参数名称': [
                    '初始资金', '风险管理器', '佣金费率', '印花税率',
                    '无风险利率', '布林带周期', 'ATR周期'
                ],
                '参数值': [
                    f'{self.initial_capital:,.0f}元',
                    '未启用',
                    f'万{self.commission * 10000:.1f} ({self.min_commission}元起)',
                    f'{self.stamp_tax_rate * 1000:.1f}‰',
                    f'{self.risk_free_rate * 100:.1f}%',
                    str(self.boll_period),
                    str(self.atr_period)
                ]
            }

        df_params = pd.DataFrame(params_data)

        # 准备绩效指标
        strategy_metrics = self._get_metrics(self.result, "主策略")[0]
        bh_metrics = self._get_metrics(self.buyandhold_result, "买入持有策略")[0]

        performance_data = {
            '指标名称': [
                '总收益率', '年化收益率', '最大回撤', '回撤恢复天数',
                '夏普比率', '卡玛比率', '交易次数',
                '胜率', '盈亏比', '平均ln(E)'
            ],
            '本策略指标值': [
                f'{strategy_metrics.get('总收益率', 0):.2f}%',
                f'{strategy_metrics.get('年化收益率', 0):.2f}%',
                f'{strategy_metrics.get('最大回撤(%)', 0):.2f}%',
                f'{strategy_metrics.get('回撤修复天数', 0):.0f}天',
                f'{strategy_metrics.get('夏普比率', 0):.2f}%',
                f'{strategy_metrics.get('卡玛比率', 0):.2f}%',
                f'{strategy_metrics.get('交易次数', 0):.0f}次',
                f'{strategy_metrics.get('胜率(%)', 0):.2f}%',
                f'{strategy_metrics.get('盈亏比', 0):.2f}',
                f'{strategy_metrics.get('平均ln(E)', 0):.2f}',
            ],
            '买入持有指标值': [
                f'{bh_metrics.get('总收益率', 0):.2f}%',
                f'{bh_metrics.get('年化收益率', 0):.2f}%',
                f'{bh_metrics.get('最大回撤(%)', 0):.2f}%',
                f'{bh_metrics.get('回撤修复天数', 0):.0f}天',
                f'{bh_metrics.get('夏普比率', 0):.2f}%',
                f'{bh_metrics.get('卡玛比率', 0):.2f}%',
                f'{bh_metrics.get('交易次数', 0):.0f}次',
                '-',
                '-',
                f'{bh_metrics.get('平均ln(E)', 0):.2f}',
            ]
        }
        df_performance = pd.DataFrame(performance_data)

        # 准备日度数据
        df_broker = self.get_observer_data('broker')
        df_returns = self.get_observer_data('timereturn')
        df_drawdown = self.get_observer_data('drawdown')
        df_tradesignal = self.get_observer_data('myobserver')
        # 删除所有子 DataFrame 的 date 列
        for df in [df_broker, df_returns, df_drawdown, df_tradesignal]:
            if df is not None and 'date' in df.columns:
                df.drop('date', axis=1, inplace=True)
        df_daily = self.data.join(df_broker).join(df_returns).join(df_drawdown).join(df_tradesignal)
        # 定义映射字典，将交易信号改为文字
        signal_map = {
            0: '',
            1: '建仓信号',
            2: '加仓',
            3: '止盈止损',
            4: '卖出信号'
        }
        # 应用映射
        df_daily['tradesignal'] = df_daily['tradesignal'].map(signal_map)

        # 准备交易记录
        trade_data = self.get_trade_data()
        dfs_trades = [pd.DataFrame(trades) for trades in trade_data if trades]

        # 写入Excel
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df_params.to_excel(writer, sheet_name='参数说明', index=False)
            df_performance.to_excel(writer, sheet_name='绩效指标', index=False)
            df_daily.to_excel(writer, sheet_name='日度数据', index=False)
            if len(dfs_trades[0]) > 0:
                dfs_trades[0].to_excel(writer, sheet_name='交易记录', index=False)
            else:
                pd.DataFrame({'说明': ['暂无交易记录']}).to_excel(writer, sheet_name='交易记录', index=False)

        print(f"回测报告已生成: {excel_path}")

    def plot_result(self):
        pass

    def run_complete_analysis(self, output_folder, use_risk_manager=True, plot=False):

        self.preprocess_data()
        self.set_position_sizer()
        self.run_backtest(plot=plot, use_risk_manager=use_risk_manager)
        self._print_comparison()
        self.write_report(output_folder)

    @staticmethod
    def compare_stocks(input_source: Union[str, List[str]],
                       min_commission=5, commission=0.0001, stamp_tax_rate=0.001,
                       risk_percent=0.02, add_ratios=[0.4, 0.3, 0.3], add_atr_multiple=1.0,
                       stop_atr_multiple=2.0, initial_capital=1000000, risk_free_rate=0.03,
                       boll_period=20, atr_period=14, output_dir = "./output",
                       use_risk_manager=True) -> pd.DataFrame:
        """
        多股票对比

        """
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)

        # 获取所有股票文件
        stock_files = []
        stock_codes = []

        if isinstance(input_source, str):
            folder_path = Path(input_source)
            if not folder_path.exists():
                raise ValueError(f"文件夹不存在: {folder_path}")

            # 修改这里：转换为字符串
            stock_files = [str(f) for f in folder_path.glob("*.csv")]
            for f in stock_files:  # f 现在是字符串
                # 从路径字符串中提取文件名
                file_name = Path(f).stem  # 需要转回 Path 来获取 stem
                if '_' in file_name:
                    parts = file_name.split('_')
                    if len(parts) >= 3:
                        stock_codes.append(parts[2])
                    else:
                        stock_codes.append(file_name)
                else:
                    stock_codes.append(file_name)
        else:
            stock_files = [str(Path(f)) for f in input_source]  # 转换为字符串
            for f in stock_files:
                file_path = Path(f)
                if not file_path.exists():
                    raise ValueError(f"文件不存在: {f}")
                stem = file_path.stem
                if '_' in stem:
                    parts = stem.split('_')
                    if len(parts) >= 3:
                        stock_codes.append(parts[2])
                    else:
                        stock_codes.append(stem)
                else:
                    stock_codes.append(stem)

        print(f"找到 {len(stock_files)} 个股票文件")

        # =============== 执行回测 ===============
        # 创建策略实例
        strategy = BollStrategy(data_path=None, min_commission=min_commission,
                                commission=commission, stamp_tax_rate=stamp_tax_rate, risk_percent=risk_percent,
                                add_ratios=add_ratios, add_atr_multiple=add_atr_multiple,
                                stop_atr_multiple=stop_atr_multiple, initial_capital=initial_capital,
                                risk_free_rate=risk_free_rate, boll_period=boll_period, atr_period=atr_period)

        for i, (file_path, code) in enumerate(zip(stock_files, stock_codes)):
            print(f"\n▶ 正在加载 [{i + 1}/{len(stock_files)}] {code}数据...")

            strategy.datapath = file_path
            # 数据挂载到strategy的PandasData属性中，格式为列表
            strategy.preprocess_data()

        # 设置统一的风险管理器（是否启用在run_strategy中配置）
        strategy.set_position_sizer()
        strategy.run_backtest(plot=False, use_risk_manager=use_risk_manager)

        # =============== 多股票对比报告 ===============
        # 定义Excel文件路径
        excel_path = os.path.join(output_path, f'布林带_多股票对比明细.xlsx')
        # 准备参数说明
        if use_risk_manager == True:
            params_data = {
                '参数名称': [
                    '初始资金', '风险管理器', '风险预算比例', '加仓比例分配', '加仓间隔(ATR倍数)',
                    '佣金费率', '印花税率', '无风险利率',
                    '止损止盈设置(ATR倍数)', '布林带周期', 'ATR周期'
                ],
                '参数值': [
                    f'{strategy.initial_capital:,.0f}元',
                    '已启用',
                    f'{strategy.risk_percent * 100:.1f}%',
                    f'{strategy.add_ratios}',
                    f'{strategy.add_atr_multiple}倍ATR',
                    f'万{strategy.commission * 10000:.1f} ({strategy.min_commission}元起)',
                    f'{strategy.stamp_tax_rate * 1000:.1f}‰',
                    f'{strategy.risk_free_rate * 100:.1f}%',
                    f'{strategy.stop_atr_multiple}倍ATR',
                    str(strategy.boll_period),
                    str(strategy.atr_period)
                ]
            }
        else:
            params_data = {
                '参数名称': [
                    '初始资金', '风险管理器', '佣金费率', '印花税率',
                    '无风险利率', '布林带周期', 'ATR周期'
                ],
                '参数值': [
                    f'{strategy.initial_capital:,.0f}元',
                    '未启用',
                    f'万{strategy.commission * 10000:.1f} ({strategy.min_commission}元起)',
                    f'{strategy.stamp_tax_rate * 1000:.1f}‰',
                    f'{strategy.risk_free_rate * 100:.1f}%',
                    str(strategy.boll_period),
                    str(strategy.atr_period)
                ]
            }

        df_params = pd.DataFrame(params_data)

        # 准备绩效指标
        strategy_metrics_list = strategy._get_metrics(strategy.result, "主策略")
        bh_metrics_list = strategy._get_metrics(strategy.buyandhold_result, "买入持有策略")
        performance_list = []
        for s, bh, stock_code in zip(strategy_metrics_list, bh_metrics_list, stock_codes):
            performance = {
                '股票代码': stock_code,
                '策略总收益率': s.get('总收益率')/100,
                '策略年化收益率': s.get('年化收益率')/100,
                '策略最大回撤率': s.get('最大回撤(%)')/100,
                '策略夏普比率': s.get('夏普比率')/100,
                '策略胜率': s.get('胜率(%)')/100,
                '策略盈亏比': s.get('盈亏比'),
                '策略交易次数': s.get('交易次数'),
                '买入持有总收益率': bh.get('总收益率')/100,
                '买入持有最大回撤率': bh.get('最大回撤(%)')/100,
                '买入持有夏普比率': bh.get('夏普比率')/100,
                '超额总收益率': s.get('总收益率')/100 - bh.get('总收益率')/100,
                '回撤改善': bh.get('最大回撤(%)')/100 - s.get('最大回撤(%)')/100,
                '平均ln(E)': s.get('平均ln(E)'),
            }
            performance_list.append(performance)
        df_performance = pd.DataFrame(performance_list)

        # 写入Excel
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df_params.to_excel(writer, sheet_name='参数说明', index=False)
            df_performance.to_excel(writer, sheet_name='绩效指标', index=False)

        print(f"多股票对比报告已生成: {excel_path}")

        # 绘制多股票对比图
        print("\n正在生成对比图...")
        fig = plt.figure(figsize=(20, 16))

        # 图1：最大回撤分布对比直方图
        ax1 = plt.subplot(2, 3, 1)
        ax1.hist(df_performance['策略最大回撤率'] * 100, bins=15, alpha=0.5, label='本策略', color='blue')
        ax1.hist(df_performance['买入持有最大回撤率'] * 100, bins=15, alpha=0.5, label='买入持有', color='red')
        ax1.set_xlabel('最大回撤 (%)')
        ax1.set_ylabel('股票个数')
        ax1.set_title('最大回撤分布对比')
        ax1.legend()

        # 图2：本策略胜率 vs 盈亏比四象限散点图
        ax2 = plt.subplot(2, 3, 2)
        scatter = ax2.scatter(df_performance['策略胜率'] * 100, df_performance['策略盈亏比'],
                              c=df_performance['策略夏普比率'], cmap='viridis', s=100, alpha=0.7)
        ax2.axhline(y=1, color='gray', linestyle='--', alpha=0.5)
        ax2.axvline(x=50, color='gray', linestyle='--', alpha=0.5)
        ax2.set_xlabel('胜率 (%)')
        ax2.set_ylabel('盈亏比')
        ax2.set_title('胜率 vs 盈亏比 (颜色=夏普比率)')
        plt.colorbar(scatter, ax=ax2, label='夏普比率')

        # 图3：本策略收益率 vs 买入持有策略收益率散点图
        ax3 = plt.subplot(2, 3, 3)
        scatter = ax3.scatter(df_performance['买入持有总收益率'] * 100, df_performance['策略总收益率'] * 100,
                              c=df_performance['超额总收益率'] * 100, cmap='RdYlGn', s=100, alpha=0.7)
        max_val = max(df_performance['买入持有总收益率'].max(), df_performance['策略总收益率'].max()) * 100
        min_val = min(df_performance['买入持有总收益率'].min(), df_performance['策略总收益率'].min()) * 100
        ax3.plot([min_val, max_val], [min_val, max_val], 'k--', alpha=0.5, label='45°线')
        ax3.set_xlabel('买入持有收益率 (%)')
        ax3.set_ylabel('本策略收益率 (%)')
        ax3.set_title('策略收益率 vs 买入持有收益率')
        plt.colorbar(scatter, ax=ax3, label='超额收益 (%)')

        # 图4：本策略超额收益 vs 回撤改善四象限散点图
        ax4 = plt.subplot(2, 3, 4)
        scatter = ax4.scatter(df_performance['超额总收益率'] * 100, df_performance['回撤改善'] * 100,
                              c=df_performance['策略夏普比率'], cmap='viridis', s=100, alpha=0.7)
        ax4.axhline(y=0, color='gray', linestyle='--', alpha=0.5)
        ax4.axvline(x=0, color='gray', linestyle='--', alpha=0.5)
        ax4.set_xlabel('超额收益率 (%)')
        ax4.set_ylabel('回撤改善 (%)')
        ax4.set_title('超额收益 vs 回撤改善 (颜色=夏普比率)')
        plt.colorbar(scatter, ax=ax4, label='夏普比率')

        # 图5：平均ln(E)直方图
        ax5 = plt.subplot(2, 3, 5)
        ax5.hist(df_performance['平均ln(E)'], bins=15, alpha=0.7, color='green', edgecolor='black')
        ax5.axvline(x=df_performance['平均ln(E)'].mean(), color='red', linestyle='--',
                    label=f'均值: {df_performance["平均ln(E)"].mean():.2f}')
        ax5.axvline(x=df_performance['平均ln(E)'].median(), color='orange', linestyle='--',
                    label=f'中位数: {df_performance["平均ln(E)"].median():.2f}')
        ax5.set_xlabel('平均ln(E)')
        ax5.set_ylabel('股票个数')
        ax5.set_title('平均ln(E)分布')
        ax5.legend()

        plt.tight_layout()

        # 保存对比图
        img_path = os.path.join(output_path, 'BOLL_多股票对比图.png')
        plt.savefig(img_path, dpi=150, bbox_inches='tight')
        plt.show()
        print(f"对比图已保存: {img_path}")

class MACDStrategy:
    """MACD策略类，负责策略层面的工作"""

    def __init__(self, initial_capital=1000000, commission_rate=0.0001,
                 stamp_tax_rate=0.001, min_commission=5, risk_free_rate=0.02,
                 macd_params=[12, 26, 9], atr_period=14):
        """
        初始化策略参数
        """
        self.initial_capital = initial_capital
        self.commission_rate = commission_rate
        self.stamp_tax_rate = stamp_tax_rate
        self.min_commission = min_commission
        self.risk_free_rate = risk_free_rate
        self.trading_days_per_year = 252

        self.fast, self.slow, self.signal = macd_params
        self.atr_period = atr_period

        self.data = None
        self.metrics = {}
        self.stock_code = None

        # 交易记录
        self.trades = []
        self.buy_trades = []

    def load_data(self, file_path):
        """
        加载数据

        Parameters:
        -----------
        file_path : str
            数据文件路径
        """
        # 尝试多种编码格式
        encodings = ['utf-8', 'gbk', 'gb2312', 'gb18030', 'utf-16', 'latin-1', 'cp936']
        df = None

        for encoding in encodings:
            try:
                df = pd.read_csv(file_path, encoding=encoding, engine='python')
                print(f"成功使用编码: {encoding}")
                break
            except UnicodeDecodeError:
                continue
            except Exception as e:
                print(f"编码 {encoding} 失败: {e}")
                continue

        if df is None:
            import chardet
            with open(file_path, 'rb') as f:
                raw_data = f.read()
                result = chardet.detect(raw_data)
                detected_encoding = result['encoding']
                print(f"检测到编码: {detected_encoding}")
                df = pd.read_csv(file_path, encoding=detected_encoding, engine='python')

        # 确保日期格式正确
        df['日期'] = pd.to_datetime(df['日期'])
        df = df.sort_values('日期').reset_index(drop=True)

        # 提取股票代码
        code_match = re.search(r'df_pre_(\d+)', file_path)
        if code_match:
            self.stock_code = code_match.group(1)
        else:
            self.stock_code = "XXXX"

        self.data = df
        print(f"数据加载完成，共 {len(self.data)} 行，时间范围: {self.data['日期'].min()} 至 {self.data['日期'].max()}")

    def preprocess_data(self):
        """
        数据处理：计算策略指标和ATR，生成交易信号
        """
        if self.data is None:
            raise ValueError("请先调用load_data加载数据")

        # 创建停牌标识
        if '异常情况' in self.data.columns:
            self.data['is_suspend'] = (self.data['异常情况'] == '停牌').astype(int)
        else:
            self.data['is_suspend'] = 0

        # 创建异常涨跌幅标识
        if '异常涨跌幅' in self.data.columns:
            # 正异常涨跌幅：不允许买入
            self.data['abnormal_buy_forbidden'] = (self.data['异常涨跌幅'] > 0).astype(int)
            # 负异常涨跌幅：不允许卖出
            self.data['abnormal_sell_forbidden'] = (self.data['异常涨跌幅'] < 0).astype(int)
        else:
            self.data['abnormal_buy_forbidden'] = 0
            self.data['abnormal_sell_forbidden'] = 0

        # 计算MACD指标
        fast_ema, slow_ema, dif, dea = TechnicalIndicators.calculate_macd(
            self.data, fast=self.fast, slow=self.slow, signal=self.signal,
            close_col='收盘', suspend_col='is_suspend'
        )
        self.data['12EMA'] = fast_ema
        self.data['26EMA'] = slow_ema
        self.data['DIF'] = dif
        self.data['DEA'] = dea

        # 计算ATR
        self.data['ATR'] = TechnicalIndicators.calculate_atr(
            self.data, period=self.atr_period,
            high_col='最高', low_col='最低', close_col='收盘',
            suspend_col='is_suspend'
        )

        # 生成原始交易信号
        self.data['原始信号'] = TechnicalIndicators.generate_signals(
            self.data, dif_col='DIF', dea_col='DEA', min_interval=5
        )

        # 处理异常情况导致的交易延迟
        self.data['交易信号'] = '无'
        self._process_abnormal_signals()

        # 计算ln(E)
        self.data['ln(E)'] = TechnicalIndicators.calculate_ln_e(
            self.data, signal_col='交易信号', close_col='收盘',
            high_col='最高', low_col='最低', window=10
        )

        # 添加shift后的收盘价
        self.data['收盘_shift1'] = self.data['收盘'].shift(1)

        print("数据处理完成，指标计算完毕")

    def _process_abnormal_signals(self):
        """处理异常情况导致的交易延迟"""
        df = self.data

        # 记录未处理的买入和卖出信号
        pending_buy = None  # (日期, 价格, 信号触发时的ATR)
        pending_sell = None  # (日期, 价格)

        for idx in range(len(df)):
            current_date = df.loc[idx, '日期']
            original_signal = df.loc[idx, '原始信号']

            # 检查是否有待处理的买入信号
            if pending_buy is not None:
                # 检查当前交易日是否允许买入（非停牌且非异常涨跌幅正数）
                if df.loc[idx, 'is_suspend'] == 0 and df.loc[idx, 'abnormal_buy_forbidden'] == 0:
                    # 可以交易，执行买入
                    df.loc[idx, '交易信号'] = '买入'
                    df.loc[idx, '原始信号'] = '买入'
                    pending_buy = None
                # 否则继续等待下一个交易日

            # 检查是否有待处理的卖出信号
            if pending_sell is not None:
                # 检查当前交易日是否允许卖出（非停牌且非异常涨跌幅负数）
                if df.loc[idx, 'is_suspend'] == 0 and df.loc[idx, 'abnormal_sell_forbidden'] == 0:
                    # 可以交易，执行卖出
                    df.loc[idx, '交易信号'] = '卖出'
                    df.loc[idx, '原始信号'] = '卖出'
                    pending_sell = None
                # 否则继续等待下一个交易日

            # 处理新产生的信号
            if original_signal == '买入' and pending_buy is None and pending_sell is None:
                # 检查当前交易日是否允许买入
                if df.loc[idx, 'is_suspend'] == 0 and df.loc[idx, 'abnormal_buy_forbidden'] == 0:
                    # 可以立即交易
                    df.loc[idx, '交易信号'] = '买入'
                else:
                    # 不能交易，加入待处理队列
                    pending_buy = (current_date, df.loc[idx, '收盘'], df.loc[idx, 'ATR'])

            elif original_signal == '卖出' and pending_sell is None and pending_buy is None:
                # 检查当前交易日是否允许卖出
                if df.loc[idx, 'is_suspend'] == 0 and df.loc[idx, 'abnormal_sell_forbidden'] == 0:
                    # 可以立即交易
                    df.loc[idx, '交易信号'] = '卖出'
                else:
                    # 不能交易，加入待处理队列
                    pending_sell = (current_date, df.loc[idx, '收盘'])

    def run_backtest(self, risk_manager=None):
        """
        运行回测

        Parameters:
        -----------
        risk_manager : RiskManage, optional
            风险管理器实例，如果不传则全仓买卖，不设止盈止损，不加仓
        """
        if self.data is None:
            raise ValueError("请先调用load_data和preprocess_data")

        df = self.data.copy()

        # 初始化回测数据列
        df['position'] = 0.0
        df['shares'] = 0.0
        df['cash'] = 0.0
        df['portfolio_value'] = 0.0
        df['returns'] = 0.0
        df['buy_cost'] = 0.0
        df['avg_cost'] = 0.0
        df['stop_loss'] = 0.0
        df['stop_profit'] = 0.0
        df['highest_close'] = 0.0
        df['add_position_flag'] = 0
        df['exit_type'] = ''
        df['last_buy_price'] = 0.0
        df['initial_atr'] = 0.0

        # 确保列类型正确
        float_cols = ['position', 'shares', 'cash', 'portfolio_value', 'returns',
                      'buy_cost', 'avg_cost', 'stop_loss', 'stop_profit', 'highest_close']
        for col in float_cols:
            df[col] = df[col].astype(float)

        # 初始化状态变量
        cash = float(self.initial_capital)
        shares = 0.0
        avg_cost = 0.0
        position_status = 0
        buy_date = None
        last_buy_price = 0.0
        buy_idx = -1
        use_risk_manager = risk_manager is not None

        # 交易记录
        self.trades = []
        self.buy_trades = []

        # 净值序列
        nav_sequence = []
        dates_sequence = []

        # 主回测循环
        for idx in range(len(df)):
            current_price = float(df.loc[idx, '收盘'])
            current_atr = float(df.loc[idx, 'ATR']) if not pd.isna(df.loc[idx, 'ATR']) else 0.0
            current_signal = df.loc[idx, '交易信号']

            # 获取前一个非停牌日的收盘价
            prev_close = None
            j = idx - 1
            while j >= 0 and prev_close is None:
                if df.loc[j, 'is_suspend'] == 0:
                    prev_close = df.loc[j, '收盘']
                j -= 1

            # 检查止盈止损（仅当启用风险管理器时）
            if use_risk_manager and position_status == 1 and shares > 0:
                is_stop, stop_type = risk_manager.check_stop(prev_close,
                                                             current_atr)
                if is_stop:
                    cash, shares, _ = self._execute_sell(idx, current_price, cash, shares, avg_cost, df)
                    if shares == 0:
                        position_status = 0
                        avg_cost = 0.0
                        buy_date = None
                        buy_idx = -1
                        last_buy_price = 0.0
                        df.loc[idx, 'exit_type'] = stop_type

            # 处理建仓信号
            if position_status == 0 and current_signal == '买入':
                if use_risk_manager and current_atr > 0:
                    # 使用风险管理器计算建仓股数
                    pos_info = risk_manager.calculate_position_size(current_price, current_atr, cash)
                    if pos_info['shares'] > 0:
                        cost = pos_info['shares'] * current_price
                        commission = max(cost * self.commission_rate, self.min_commission)
                        total_cost = cost + commission

                        if total_cost <= cash:
                            cash -= total_cost
                            shares = float(pos_info['shares'])
                            avg_cost = float(current_price)
                            position_status = 1
                            buy_date = df.loc[idx, '日期']
                            buy_idx = idx
                            last_buy_price = current_price

                            risk_manager.update_risk_state(
                                'open',
                                price=current_price,
                                stop_loss=pos_info['stop_loss'],
                                stop_profit=pos_info['stop_profit'],
                                atr_value=pos_info['atr_value'],
                                add_shares_list=pos_info['add_shares_list']
                            )

                            self.buy_trades.append({
                                'date': df.loc[idx, '日期'],
                                'price': current_price,
                                'cost': pos_info['shares'] * current_price
                            })
                else:
                    # 未启用风险管理器：全仓买入
                    # 计算可买股数（整百股），并调整直到资金足够
                    max_shares = int(cash / (current_price * 100)) * 100

                    # 循环减少股数直到资金足够
                    while max_shares > 0:
                        cost = max_shares * current_price
                        commission = max(cost * self.commission_rate, self.min_commission)
                        total_cost = cost + commission

                        if total_cost <= cash:
                            break
                        max_shares -= 100

                    if max_shares > 0:
                        cost = max_shares * current_price
                        commission = max(cost * self.commission_rate, self.min_commission)
                        total_cost = cost + commission

                        cash -= total_cost
                        shares = float(max_shares)
                        avg_cost = float(current_price)
                        position_status = 1
                        buy_date = df.loc[idx, '日期']
                        buy_idx = idx
                        last_buy_price = current_price

                        self.buy_trades.append({
                            'date': df.loc[idx, '日期'],
                            'price': current_price,
                            'cost': cost
                        })

            # 处理加仓（仅当启用风险管理器时）
            if use_risk_manager and position_status == 1 and shares > 0 and idx > buy_idx:
                if risk_manager.is_can_add(prev_close):
                    add_info = risk_manager.calculate_add_position_size(current_price, shares, cash)
                    if add_info['shares'] > 0:
                        cost = add_info['shares'] * current_price
                        commission = max(cost * self.commission_rate, self.min_commission)
                        total_cost = cost + commission

                        if total_cost <= cash:
                            cash -= total_cost
                            old_shares = shares
                            shares += add_info['shares']
                            avg_cost = (avg_cost * old_shares + current_price * add_info['shares']) / shares

                            risk_manager.update_risk_state(
                                'add',
                                price=current_price,
                                old_shares=old_shares,
                                new_shares=add_info['shares'],
                                new_stop_loss=add_info['new_stop_loss']
                            )

                            last_buy_price = current_price

            # 处理卖出信号
            if position_status == 1 and current_signal == '卖出':
                cash, shares, _ = self._execute_sell(idx, current_price, cash, shares, avg_cost, df)
                if shares == 0:
                    position_status = 0
                    avg_cost = 0.0
                    buy_date = None
                    buy_idx = -1
                    last_buy_price = 0.0
                    if use_risk_manager:
                        risk_manager.update_risk_state('close')
                    df.loc[idx, 'exit_type'] = '死叉'

            # 更新持仓数据
            df.loc[idx, 'position'] = float(position_status)
            df.loc[idx, 'shares'] = float(shares)
            df.loc[idx, 'cash'] = float(cash)
            df.loc[idx, 'portfolio_value'] = float(cash + shares * current_price)
            df.loc[idx, 'avg_cost'] = float(avg_cost)
            df.loc[idx, 'last_buy_price'] = float(last_buy_price)

            if use_risk_manager:
                risk_state = risk_manager.get_risk_state()
                df.loc[idx, 'stop_loss'] = float(risk_state['stop_loss'])
                df.loc[idx, 'stop_profit'] = float(risk_state['stop_profit'])
                df.loc[idx, 'highest_close'] = float(risk_state['highest_close'])
                df.loc[idx, 'add_position_flag'] = risk_state['add_count']
                df.loc[idx, 'initial_atr'] = float(risk_state['initial_atr'])

            # 记录净值
            nav_sequence.append(cash + shares * current_price)
            dates_sequence.append(df.loc[idx, '日期'])

        # 计算收益率
        df['returns'] = df['portfolio_value'].pct_change()
        df['cumulative_returns'] = df['portfolio_value'] / self.initial_capital - 1

        # 计算买入持有策略
        buy_hold = BuyHoldStrategy(
            self.initial_capital, self.commission_rate,
            self.stamp_tax_rate, self.min_commission
        )
        buy_hold.run_backtest(df)
        df['buy_hold_value'] = buy_hold.net_values.values
        df['buy_hold_returns'] = df['buy_hold_value'].pct_change()
        df['buy_hold_cumulative'] = df['buy_hold_value'] / self.initial_capital - 1

        self.data = df
        self.use_risk_manager = use_risk_manager

        print("回测完成")

    def _execute_sell(self, idx, price, cash, shares, avg_cost, df):
        """执行卖出操作"""
        if shares > 0:
            sale_amount = shares * price
            commission = max(sale_amount * self.commission_rate, self.min_commission)
            stamp_tax = sale_amount * self.stamp_tax_rate
            total_cost = commission + stamp_tax
            cash += (sale_amount - total_cost)

            # 记录完整交易
            if len(self.buy_trades) > 0:
                self.trades.append({
                    'buy_date': self.buy_trades[-1]['date'],
                    'sell_date': df.loc[idx, '日期'],
                    'buy_price': self.buy_trades[-1]['price'],
                    'sell_price': price,
                    'profit': sale_amount - total_cost - self.buy_trades[-1]['cost']
                })

            shares = 0.0
            return cash, shares, sale_amount - total_cost

        return cash, shares, 0

    def calculate_metrics(self):
        """计算绩效指标"""
        if self.data is None:
            raise ValueError("请先运行回测")

        # 确保买入持有策略的数据已正确计算
        # 如果buy_hold_value列不存在，重新计算买入持有策略
        if 'buy_hold_value' not in self.data.columns:
            buy_hold = BuyHoldStrategy(
                self.initial_capital, self.commission_rate,
                self.stamp_tax_rate, self.min_commission
            )
            buy_hold.run_backtest(self.data)
            self.data['buy_hold_value'] = buy_hold.net_values
            self.data['buy_hold_returns'] = self.data['buy_hold_value'].pct_change()
            self.data['buy_hold_cumulative'] = self.data['buy_hold_value'] / self.initial_capital - 1

        # 计算策略指标
        strategy_metrics = self._calculate_strategy_metrics()

        # 计算买入持有指标
        buy_hold_metrics = self._calculate_buy_hold_metrics()

        # 计算交易相关指标
        complete_trades = [t for t in self.trades if t.get('buy_date') is not None and t.get('profit') is not None]

        if len(complete_trades) > 0:
            win_trades = [t for t in complete_trades if t.get('profit', 0) > 0]
            win_rate = len(win_trades) / len(complete_trades)
            avg_profit = np.mean([t['profit'] for t in win_trades]) if win_trades else 0
            loss_trades = [t for t in complete_trades if t.get('profit', 0) <= 0]
            avg_loss = abs(np.mean([t['profit'] for t in loss_trades])) if loss_trades else 0
            profit_loss_ratio = avg_profit / avg_loss if avg_loss > 0 else 0
        else:
            win_rate = 0
            profit_loss_ratio = 0

        # 平均ln(E)
        df_lnE = self.data[self.data['交易信号'] == '买入']['ln(E)'].dropna()
        avg_lnE = df_lnE.mean() if len(df_lnE) > 0 else 0

        # 加仓贡献率
        add_contribution_rate = self._calculate_add_contribution_rate()

        self.metrics = {
            'strategy': strategy_metrics,
            'buy_hold': buy_hold_metrics,
            'win_rate': win_rate,
            'profit_loss_ratio': profit_loss_ratio,
            'avg_lnE': avg_lnE,
            'add_contribution_rate': add_contribution_rate,
            'trade_count': len(complete_trades)
        }

        # 控制台打印对比表格
        self._print_metrics_table()

        return self.metrics

    def _calculate_strategy_metrics(self):
        """计算策略绩效指标"""
        # 确保有累计收益率
        if 'cumulative_returns' not in self.data.columns:
            self.data['cumulative_returns'] = self.data['portfolio_value'] / self.initial_capital - 1

        returns_clean = self.data['returns'].dropna()
        if len(returns_clean) == 0:
            return {
                'total_return': 0, 'annual_return': 0, 'max_drawdown': 0,
                'sharpe_ratio': 0, 'calmar_ratio': 0,
                'recovery_time': None, 'max_recovery_time': 0
            }

        # 总收益率
        cumulative = self.data['cumulative_returns']
        if len(cumulative) > 0 and not pd.isna(cumulative.iloc[-1]):
            total_return = cumulative.iloc[-1]
        else:
            total_return = 0

        # 年化收益率
        days = len(self.data)
        if days > 0 and total_return > -1:
            annual_return = (1 + total_return) ** (self.trading_days_per_year / days) - 1
        else:
            annual_return = 0

        # 最大回撤
        rolling_max = self.data['portfolio_value'].expanding().max()
        drawdown = (self.data['portfolio_value'] - rolling_max) / rolling_max
        max_drawdown = drawdown.min() if len(drawdown) > 0 else 0

        # 夏普比率
        if len(returns_clean) > 0 and returns_clean.std() > 0:
            daily_rf = (1 + self.risk_free_rate) ** (1 / self.trading_days_per_year) - 1
            excess_returns = returns_clean - daily_rf
            sharpe_ratio = excess_returns.mean() / excess_returns.std() * np.sqrt(self.trading_days_per_year)
        else:
            sharpe_ratio = 0

        calmar_ratio = annual_return / abs(max_drawdown) if max_drawdown != 0 else 0

        # 计算最大回撤恢复时间
        max_recovery_time = 0
        in_drawdown = False
        drawdown_start_pos = None

        for i in range(len(drawdown)):
            if i == 0:
                continue
            if drawdown.iloc[i] < 0 and not in_drawdown:
                in_drawdown = True
                drawdown_start_pos = i
            elif in_drawdown and drawdown.iloc[i] == 0:
                in_drawdown = False
                if drawdown_start_pos is not None:
                    min_pos = drawdown_start_pos
                    for j in range(drawdown_start_pos, i + 1):
                        if drawdown.iloc[j] < drawdown.iloc[min_pos]:
                            min_pos = j
                    recovery_days = i - min_pos
                    max_recovery_time = max(max_recovery_time, recovery_days)

        return {
            'total_return': total_return,
            'annual_return': annual_return,
            'max_drawdown': max_drawdown,
            'sharpe_ratio': sharpe_ratio,
            'calmar_ratio': calmar_ratio,
            'recovery_time': None,
            'max_recovery_time': max_recovery_time
        }

    def _calculate_buy_hold_metrics(self):
        """计算买入持有策略绩效指标"""
        # 确保数据存在
        if 'buy_hold_cumulative' not in self.data.columns:
            # 重新计算买入持有策略
            buy_hold = BuyHoldStrategy(
                self.initial_capital, self.commission_rate,
                self.stamp_tax_rate, self.min_commission
            )
            buy_hold.run_backtest(self.data)
            self.data['buy_hold_value'] = buy_hold.net_values
            self.data['buy_hold_returns'] = self.data['buy_hold_value'].pct_change()
            self.data['buy_hold_cumulative'] = self.data['buy_hold_value'] / self.initial_capital - 1

        # 获取累计收益率序列
        cumulative = self.data['buy_hold_cumulative']

        # 总收益率
        if len(cumulative) > 0 and not pd.isna(cumulative.iloc[-1]):
            total_return = cumulative.iloc[-1]
        else:
            total_return = 0

        # 年化收益率
        days = len(self.data)
        if days > 0 and total_return > -1:
            annual_return = (1 + total_return) ** (self.trading_days_per_year / days) - 1
        else:
            annual_return = 0

        # 最大回撤
        if 'buy_hold_value' in self.data.columns:
            rolling_max = self.data['buy_hold_value'].expanding().max()
            drawdown = (self.data['buy_hold_value'] - rolling_max) / rolling_max
            max_drawdown = drawdown.min() if len(drawdown) > 0 else 0
        else:
            max_drawdown = 0

        # 夏普比率
        returns_clean = self.data['buy_hold_returns'].dropna()
        if len(returns_clean) > 0 and returns_clean.std() > 0:
            daily_rf = (1 + self.risk_free_rate) ** (1 / self.trading_days_per_year) - 1
            excess_returns = returns_clean - daily_rf
            sharpe_ratio = excess_returns.mean() / returns_clean.std() * np.sqrt(self.trading_days_per_year)
        else:
            sharpe_ratio = 0

        return {
            'total_return': total_return,
            'annual_return': annual_return,
            'max_drawdown': max_drawdown,
            'sharpe_ratio': sharpe_ratio
        }

    def _calculate_add_contribution_rate(self):
        """计算加仓贡献率"""
        total_profit_from_add = 0
        total_profit = 0

        # 从交易记录构建加仓信息
        trades_df = self._build_trade_records()

        if len(trades_df) > 0:
            for _, trade in trades_df.iterrows():
                try:
                    profit = float(trade['盈亏金额']) if trade['盈亏金额'] != '' else 0
                    total_profit += profit

                    base_shares = trade['建仓股数'] if isinstance(trade['建仓股数'], (int, float)) else 0
                    add1_shares = trade['第一次加仓股数'] if isinstance(trade['第一次加仓股数'], (int, float)) else 0
                    add2_shares = trade['第二次加仓股数'] if isinstance(trade['第二次加仓股数'], (int, float)) else 0
                    total_shares = base_shares + add1_shares + add2_shares

                    if total_shares > 0 and (add1_shares > 0 or add2_shares > 0):
                        add_ratio = (add1_shares + add2_shares) / total_shares
                        add_profit = profit * add_ratio
                        total_profit_from_add += add_profit
                except:
                    pass

        return total_profit_from_add / total_profit if total_profit != 0 else 0

    def _build_trade_records(self):
        """从回测数据构建交易记录"""
        trade_records = []
        df = self.data

        # 如果不启用风险管理器，使用简化版交易记录
        if hasattr(self, 'use_risk_manager') and not self.use_risk_manager:
            i = 0
            while i < len(df):
                if i > 0 and df.loc[i, 'position'] == 1 and df.loc[i - 1, 'position'] == 0:
                    buy_idx = i
                    buy_date = df.loc[buy_idx, '日期']
                    buy_price = df.loc[buy_idx, '收盘']
                    buy_shares = df.loc[buy_idx, 'shares']

                    if buy_shares < 100:
                        i += 1
                        continue

                    # 找到卖出点
                    j = buy_idx + 1
                    while j < len(df) and df.loc[j, 'position'] == 1:
                        j += 1

                    if j < len(df) and df.loc[j, 'position'] == 0:
                        sell_idx = j
                        sell_date = df.loc[sell_idx, '日期']
                        sell_price = df.loc[sell_idx, '收盘']
                        exit_type = df.loc[sell_idx, 'exit_type']
                        if exit_type == '' or pd.isna(exit_type):
                            exit_type = '死叉'

                        # 计算盈亏
                        buy_cost = buy_shares * buy_price
                        buy_commission = max(buy_cost * self.commission_rate, self.min_commission)
                        total_buy_cost = buy_cost + buy_commission

                        sale_amount = buy_shares * sell_price
                        sell_commission = max(sale_amount * self.commission_rate, self.min_commission)
                        stamp_tax = sale_amount * self.stamp_tax_rate
                        net_sale = sale_amount - sell_commission - stamp_tax

                        profit = net_sale - total_buy_cost
                        return_rate = profit / total_buy_cost * 100 if total_buy_cost > 0 else 0
                        hold_days = (sell_date - buy_date).days

                        trade = {
                            '买入日期': buy_date,
                            '卖出日期': sell_date,
                            '买入价': buy_price,
                            '卖出价': sell_price,
                            '股数': buy_shares,
                            '收益率': f'{return_rate:.2f}%',
                            '盈亏金额': f'{profit:.2f}'
                        }
                        trade_records.append(trade)
                        i = sell_idx + 1
                        continue
                    else:
                        i = j
                        continue
                else:
                    i += 1
            return pd.DataFrame(trade_records)

        # ========== 启用风险管理器时，使用原有的动态加仓版本 ==========
        i = 0
        while i < len(df):
            if i > 0 and df.loc[i, 'position'] == 1 and df.loc[i - 1, 'position'] == 0:
                buy_idx = i
                buy_date = df.loc[buy_idx, '日期']
                buy_price = df.loc[buy_idx, '收盘']
                buy_shares = df.loc[buy_idx, 'shares']
                buy_stop_loss = df.loc[buy_idx, 'stop_loss']

                if buy_shares < 500:
                    i += 1
                    continue

                # 获取加仓次数
                add_count_total = int(df.loc[buy_idx, 'add_position_flag']) if 'add_position_flag' in df.columns else 0

                # 动态构建交易记录字典
                trade = {
                    '建仓日期': buy_date,
                    '建仓价': buy_price,
                    '建仓股数': buy_shares,
                    '建仓止损价': buy_stop_loss,
                    '卖出日期': '',
                    '卖出价': '',
                    '平均成本': buy_price,
                    '总股数': buy_shares,
                    '收益率': '',
                    '盈亏金额': '',
                    '持仓天数': '',
                    '离场类型': ''
                }

                # 动态添加加仓字段
                for add_num in range(1, add_count_total + 1):
                    trade[f'第{add_num}次加仓日期'] = ''
                    trade[f'第{add_num}次加仓价'] = ''
                    trade[f'第{add_num}次加仓股数'] = ''
                    trade[f'第{add_num}次加仓止损价'] = ''

                total_shares = buy_shares
                total_cost = buy_shares * buy_price
                add_count = 1
                last_shares = total_shares

                j = buy_idx + 1
                while j < len(df) and df.loc[j, 'position'] == 1:
                    current_shares = df.loc[j, 'shares']
                    if current_shares > last_shares:
                        added_shares = current_shares - last_shares
                        add_price = df.loc[j, '收盘']
                        add_stop_loss = df.loc[j, 'stop_loss']

                        trade[f'第{add_count}次加仓日期'] = df.loc[j, '日期']
                        trade[f'第{add_count}次加仓价'] = add_price
                        trade[f'第{add_count}次加仓股数'] = added_shares
                        trade[f'第{add_count}次加仓止损价'] = add_stop_loss

                        total_cost += added_shares * add_price
                        total_shares = current_shares
                        trade['平均成本'] = total_cost / total_shares
                        trade['总股数'] = total_shares
                        add_count += 1
                        last_shares = current_shares

                    j += 1

                sell_idx = j
                if sell_idx < len(df) and df.loc[sell_idx, 'position'] == 0:
                    sell_date = df.loc[sell_idx, '日期']
                    sell_price = df.loc[sell_idx, '收盘']
                    exit_type = df.loc[sell_idx, 'exit_type']
                    if exit_type == '' or pd.isna(exit_type):
                        exit_type = '死叉'

                    # 计算总买入成本（包含手续费）
                    total_buy_cost_with_fee = 0
                    total_shares_count = 0

                    # 建仓
                    buy_cost = trade['建仓股数'] * trade['建仓价']
                    buy_commission = max(buy_cost * self.commission_rate, self.min_commission)
                    total_buy_cost_with_fee += buy_cost + buy_commission
                    total_shares_count += trade['建仓股数']

                    # 动态计算各次加仓
                    for add_num in range(1, add_count_total + 1):
                        add_shares = trade.get(f'第{add_num}次加仓股数', 0)
                        if add_shares != '' and add_shares != 0:
                            add_price = trade.get(f'第{add_num}次加仓价', 0)
                            add_cost = add_shares * add_price
                            add_commission = max(add_cost * self.commission_rate, self.min_commission)
                            total_buy_cost_with_fee += add_cost + add_commission
                            total_shares_count += add_shares

                    avg_cost_with_fee = total_buy_cost_with_fee / total_shares_count if total_shares_count > 0 else 0

                    sale_amount = total_shares_count * sell_price
                    sell_commission = max(sale_amount * self.commission_rate, self.min_commission)
                    stamp_tax = sale_amount * self.stamp_tax_rate
                    net_sale = sale_amount - sell_commission - stamp_tax

                    profit = net_sale - total_buy_cost_with_fee
                    return_rate = profit / total_buy_cost_with_fee * 100 if total_buy_cost_with_fee > 0 else 0
                    hold_days = (sell_date - buy_date).days

                    trade['卖出日期'] = sell_date
                    trade['卖出价'] = sell_price
                    trade['平均成本'] = avg_cost_with_fee
                    trade['收益率'] = f'{return_rate:.2f}%'
                    trade['盈亏金额'] = f'{profit:.2f}'
                    trade['持仓天数'] = hold_days
                    trade['离场类型'] = exit_type

                    trade_records.append(trade)
                    i = sell_idx + 1
                    continue
                else:
                    i = j
                    continue
            else:
                i += 1

        return pd.DataFrame(trade_records)

    def _print_metrics_table(self):
        """打印绩效指标对比表格"""
        s = self.metrics['strategy']
        bh = self.metrics['buy_hold']

        print("\n" + "=" * 80)
        print("策略绩效指标对比")
        print("=" * 80)
        print(f"{'指标':<25} {'本策略':>15} {'买入持有':>15}")
        print("-" * 55)
        print(f"{'总收益率':<25} {s['total_return'] * 100:>14.2f}% {bh['total_return'] * 100:>14.2f}%")
        print(f"{'年化收益率':<25} {s['annual_return'] * 100:>14.2f}% {bh['annual_return'] * 100:>14.2f}%")
        print(f"{'最大回撤率':<25} {abs(s['max_drawdown']) * 100:>14.2f}% {abs(bh['max_drawdown']) * 100:>14.2f}%")
        print(f"{'最大回撤恢复时间(交易日)':<25} {s['recovery_time'] if s['recovery_time'] is not None else 'N/A':>15} "
              f"{bh['recovery_time'] if bh.get('recovery_time') is not None else 'N/A':>15}")
        print(f"{'最大恢复时间(交易日)':<25} {s['max_recovery_time']:>15} {bh.get('max_recovery_time', 'N/A'):>15}")
        print(f"{'夏普比率':<25} {s['sharpe_ratio']:>15.2f} {bh['sharpe_ratio']:>15.2f}")
        print(f"{'卡玛比率':<25} {s['calmar_ratio']:>15.2f} {'-':>15}")
        print(f"{'交易次数':<25} {self.metrics['trade_count']:>15}")
        print(f"{'胜率':<25} {self.metrics['win_rate'] * 100:>14.2f}%")
        print(f"{'盈亏比':<25} {self.metrics['profit_loss_ratio']:>15.2f}")
        print(f"{'平均ln(E)':<25} {self.metrics['avg_lnE']:>15.2f}")
        print("=" * 80)

    def plot_result(self, save_path=None):
        """绘制图表"""
        if self.data is None:
            raise ValueError("请先运行回测")

        df = self.data

        fig = plt.figure(figsize=(16, 28))

        # 子图1：价格和12EMA、26EMA + 买卖点标记
        ax1 = plt.subplot(8, 1, 1)
        ax1.plot(df['日期'], df['收盘'], label='收盘价', color='black', linewidth=1)
        ax1.plot(df['日期'], df['12EMA'], label='12EMA', color='blue', linewidth=0.8)
        ax1.plot(df['日期'], df['26EMA'], label='26EMA', color='red', linewidth=0.8)

        buy_points = df[df['交易信号'] == '买入']
        sell_points = df[df['交易信号'] == '卖出']
        if len(buy_points) > 0:
            ax1.scatter(buy_points['日期'], buy_points['收盘'], color='red', marker='^', s=30, label='买入', alpha=0.7)
        if len(sell_points) > 0:
            ax1.scatter(sell_points['日期'], sell_points['收盘'], color='green', marker='v', s=30, label='卖出',
                        alpha=0.7)

        ax1.set_ylabel('价格')
        ax1.legend(loc='upper left', fontsize=8)
        ax1.grid(True, alpha=0.3)
        ax1.set_title('价格、EMA指标及交易信号', fontsize=12)

        # 子图2：DIF、DEA和0轴
        ax2 = plt.subplot(8, 1, 2)
        ax2.plot(df['日期'], df['DIF'], label='DIF', color='blue', linewidth=1)
        ax2.plot(df['日期'], df['DEA'], label='DEA', color='red', linewidth=1)
        ax2.axhline(y=0, color='gray', linestyle='--', linewidth=0.8, alpha=0.5, label='0轴')

        ax2.fill_between(df['日期'], df['DIF'], df['DEA'],
                         where=(df['DIF'] >= df['DEA']),
                         facecolor='red', alpha=0.3, label='DIF > DEA')
        ax2.fill_between(df['日期'], df['DIF'], df['DEA'],
                         where=(df['DIF'] < df['DEA']),
                         facecolor='green', alpha=0.3, label='DIF < DEA')

        ax2.set_ylabel('MACD值')
        ax2.legend(loc='upper left', fontsize=8)
        ax2.grid(True, alpha=0.3)
        ax2.set_title('MACD指标 (DIF, DEA, 0轴)', fontsize=12)

        # 子图3：持仓状态
        ax3 = plt.subplot(8, 1, 3)
        ax3.fill_between(df['日期'], df['position'], 0, alpha=0.5, color='blue')
        ax3.set_ylabel('持仓状态')
        ax3.set_ylim([-0.1, 1.1])
        ax3.set_title('持仓状态', fontsize=12)
        ax3.grid(True, alpha=0.3)

        # 子图4：策略净值 vs 买入持有净值（对数坐标）
        ax4 = plt.subplot(8, 1, 4)
        ax4.semilogy(df['日期'], df['portfolio_value'], label='策略净值', color='blue', linewidth=1.5)
        ax4.semilogy(df['日期'], df['buy_hold_value'], label='买入持有净值', color='red', linewidth=1.5)
        ax4.axhline(y=self.initial_capital, color='gray', linestyle='--', alpha=0.5)
        ax4.set_ylabel('净值（对数坐标）')
        ax4.legend(loc='upper left', fontsize=8)
        ax4.grid(True, alpha=0.3)
        ax4.set_title('策略净值 vs 买入持有净值', fontsize=12)

        # 子图5：回撤曲线
        ax5 = plt.subplot(8, 1, 5)
        rolling_max = df['portfolio_value'].expanding().max()
        drawdown = (df['portfolio_value'] - rolling_max) / rolling_max
        ax5.fill_between(df['日期'], drawdown * 100, 0, alpha=0.5, color='red', label='回撤')
        ax5.set_ylabel('回撤 (%)')
        ax5.set_title('回撤曲线', fontsize=12)
        ax5.legend(loc='lower left', fontsize=8)
        ax5.grid(True, alpha=0.3)

        # 子图6：每日收益率直方图
        ax6 = plt.subplot(8, 1, 6)
        returns = df['returns'].dropna() * 100
        if len(returns) > 0:
            ax6.hist(returns, bins=50, alpha=0.7, color='blue', edgecolor='black')
            ax6.axvline(x=returns.mean(), color='red', linestyle='--', linewidth=1.5,
                        label=f'均值: {returns.mean():.2f}%')
            ax6.axvline(x=returns.median(), color='green', linestyle='--', linewidth=1.5,
                        label=f'中位数: {returns.median():.2f}%')
            ax6.set_xlabel('收益率 (%)')
            ax6.set_ylabel('频数')
            ax6.set_title('每日收益率分布', fontsize=12)
            ax6.legend(fontsize=8)
            ax6.grid(True, alpha=0.3)

        # 子图7：滚动年化收益率折线图
        ax7 = plt.subplot(8, 1, 7)
        rolling_window = 252
        rolling_returns = []
        rolling_dates = []
        for i in range(rolling_window, len(df)):
            if df['portfolio_value'].iloc[i - rolling_window] > 0:
                period_returns = (df['portfolio_value'].iloc[i] / df['portfolio_value'].iloc[i - rolling_window] - 1)
                annualized_return = (1 + period_returns) ** (self.trading_days_per_year / rolling_window) - 1
                rolling_returns.append(annualized_return * 100)
                rolling_dates.append(df['日期'].iloc[i])

        if len(rolling_returns) > 0:
            ax7.plot(rolling_dates, rolling_returns, color='green', linewidth=1.5)
            ax7.axhline(y=0, color='gray', linestyle='--', alpha=0.5)
            mean_rolling = np.mean(rolling_returns)
            ax7.axhline(y=mean_rolling, color='orange', linestyle='--', alpha=0.7, label=f'均值: {mean_rolling:.2f}%')
            ax7.set_ylabel('年化收益率 (%)')
            ax7.set_title('滚动年化收益率（252天窗口）', fontsize=12)
            ax7.legend(fontsize=8)
            ax7.grid(True, alpha=0.3)

        # 子图8：滚动年化收益率直方图
        ax8 = plt.subplot(8, 1, 8)
        if len(rolling_returns) > 0:
            ax8.hist(rolling_returns, bins=30, alpha=0.7, color='orange', edgecolor='black')
            mean_return = np.mean(rolling_returns)
            std_return = np.std(rolling_returns)
            ax8.axvline(x=mean_return, color='red', linestyle='--', linewidth=1.5, label=f'均值: {mean_return:.2f}%')
            ax8.axvline(x=mean_return + std_return, color='blue', linestyle='--', linewidth=1,
                        label=f'+1σ: {mean_return + std_return:.2f}%')
            ax8.axvline(x=mean_return - std_return, color='blue', linestyle='--', linewidth=1,
                        label=f'-1σ: {mean_return - std_return:.2f}%')
            ax8.set_xlabel('年化收益率 (%)')
            ax8.set_ylabel('频数')
            ax8.set_title('滚动年化收益率分布', fontsize=12)
            ax8.legend(fontsize=8)
            ax8.grid(True, alpha=0.3)

        plt.tight_layout()

        if save_path:
            plt.savefig(save_path, dpi=150, bbox_inches='tight')
            print(f"图片已保存至: {save_path}")

        plt.show()

    def write_report(self, save_path):
        """编写回测报告"""
        if self.data is None:
            raise ValueError("请先运行回测")

        # 准备参数说明
        if hasattr(self, 'use_risk_manager') and self.use_risk_manager:
            params_data = {
                '参数名称': [
                    '初始资金', '风险管理器', '风险预算比例', '加仓比例分配', '加仓间隔(ATR倍数)',
                    '佣金费率', '最低佣金', '印花税率', '无风险利率',
                    '交易天数(年)', '止损设置(ATR倍数)', '止盈设置(ATR倍数)',
                    'EMA周期1', 'EMA周期2', 'DEA周期', 'ATR周期'
                ],
                '参数值': [
                    f'{self.initial_capital:,.0f}元',
                    '已启用',
                    '2.0%',
                    '40%, 30%, 30%',
                    '1倍ATR',
                    f'万{self.commission_rate * 10000:.1f} ({self.min_commission}元起)',
                    f'{self.min_commission}元',
                    f'{self.stamp_tax_rate * 1000:.1f}‰',
                    f'{self.risk_free_rate * 100:.1f}%',
                    str(self.trading_days_per_year),
                    '2倍ATR',
                    '2倍ATR',
                    str(self.fast),
                    str(self.slow),
                    str(self.signal),
                    str(self.atr_period)
                ]
            }
        else:
            params_data = {
                '参数名称': [
                    '初始资金', '风险管理器', '佣金费率', '最低佣金', '印花税率',
                    '无风险利率', '交易天数(年)', 'EMA周期1', 'EMA周期2', 'DEA周期', 'ATR周期'
                ],
                '参数值': [
                    f'{self.initial_capital:,.0f}元',
                    '未启用',
                    f'万{self.commission_rate * 10000:.1f} ({self.min_commission}元起)',
                    f'{self.min_commission}元',
                    f'{self.stamp_tax_rate * 1000:.1f}‰',
                    f'{self.risk_free_rate * 100:.1f}%',
                    str(self.trading_days_per_year),
                    str(self.fast),
                    str(self.slow),
                    str(self.signal),
                    str(self.atr_period)
                ]
            }

        df_params = pd.DataFrame(params_data)

        # 准备绩效指标
        s = self.metrics['strategy']
        bh = self.metrics['buy_hold']
        performance_data = {
            '指标名称': [
                '总收益率', '年化收益率', '最大回撤', '最大回撤恢复时间(交易日)',
                '最大恢复时间(交易日)', '夏普比率', '卡玛比率', '交易次数',
                '胜率', '盈亏比', '平均ln(E)'
            ],
            '本策略指标值': [
                f'{s["total_return"] * 100:.2f}%',
                f'{s["annual_return"] * 100:.2f}%',
                f'{abs(s["max_drawdown"]) * 100:.2f}%',
                str(s["recovery_time"] if s["recovery_time"] is not None else 'N/A'),
                str(s["max_recovery_time"]),
                f'{s["sharpe_ratio"]:.2f}',
                f'{s["calmar_ratio"]:.2f}',
                str(self.metrics['trade_count']),
                f'{self.metrics["win_rate"] * 100:.2f}%',
                f'{self.metrics["profit_loss_ratio"]:.2f}',
                f'{self.metrics["avg_lnE"]:.2f}'
            ],
            '买入持有指标值': [
                f'{bh["total_return"] * 100:.2f}%',
                f'{bh["annual_return"] * 100:.2f}%',
                f'{abs(bh["max_drawdown"]) * 100:.2f}%',
                'N/A',
                'N/A',
                f'{bh["sharpe_ratio"]:.2f}',
                '-',
                '-',
                '-',
                '-',
                '-'
            ]
        }
        df_performance = pd.DataFrame(performance_data)

        # 准备日度数据
        df = self.data
        rolling_window = 252
        rolling_annual_returns = []
        for i in range(len(df)):
            if i >= rolling_window:
                period_returns = df['portfolio_value'].iloc[i] / df['portfolio_value'].iloc[i - rolling_window] - 1
                annualized_return = (1 + period_returns) ** (self.trading_days_per_year / rolling_window) - 1
                rolling_annual_returns.append(annualized_return * 100)
            else:
                rolling_annual_returns.append(np.nan)

        drawdown_positive = []
        for i in range(len(df)):
            if i == 0:
                drawdown_positive.append(0)
            else:
                rolling_max_val = df['portfolio_value'].iloc[:i + 1].max()
                dd = (rolling_max_val - df['portfolio_value'].iloc[i]) / rolling_max_val * 100
                drawdown_positive.append(dd)

        daily_data = pd.DataFrame()
        daily_data['日期'] = df['日期']
        daily_data['股票代码'] = self.stock_code
        daily_data['开盘'] = df['开盘'] if '开盘' in df.columns else np.nan
        daily_data['最高'] = df['最高']
        daily_data['最低'] = df['最低']
        daily_data['收盘'] = df['收盘']
        daily_data['异常情况'] = df['异常情况'] if '异常情况' in df.columns else ''
        daily_data['异常涨跌幅'] = df['异常涨跌幅'] if '异常涨跌幅' in df.columns else 0
        daily_data['12EMA'] = df['12EMA'] if '12EMA' in df.columns else np.nan
        daily_data['26EMA'] = df['26EMA'] if '26EMA' in df.columns else np.nan
        daily_data['DIF'] = df['DIF']
        daily_data['DEA'] = df['DEA']
        daily_data['ATR'] = df['ATR']
        daily_data['交易信号'] = df['交易信号']
        daily_data['ln(E)'] = df['ln(E)']
        daily_data['策略净值'] = df['portfolio_value'].round(2)
        daily_data['持仓市值'] = (df['shares'] * df['收盘']).round(2)
        daily_data['可用资金'] = df['cash'].round(2)
        daily_data['总资产'] = df['portfolio_value'].round(2)
        daily_data['回撤率'] = [round(x / 100, 4) for x in drawdown_positive]
        daily_data['日收益率'] = df['returns'].round(4)
        daily_data['累计收益率'] = df['cumulative_returns'].round(4)
        daily_data['滚动年化收益率'] = [round(x / 100, 4) if not pd.isna(x) else np.nan for x in rolling_annual_returns]

        # 准备交易记录
        df_trades = self._build_trade_records()

        # 写入Excel
        with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
            df_params.to_excel(writer, sheet_name='参数说明', index=False)
            df_performance.to_excel(writer, sheet_name='绩效指标', index=False)
            daily_data.to_excel(writer, sheet_name='日度数据', index=False)
            if len(df_trades) > 0:
                df_trades.to_excel(writer, sheet_name='交易记录', index=False)
            else:
                pd.DataFrame({'说明': ['暂无交易记录']}).to_excel(writer, sheet_name='交易记录', index=False)

        # 格式化Excel
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, Alignment, PatternFill

            wb = load_workbook(save_path)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                header_alignment = Alignment(horizontal='center', vertical='center')

                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    ws.column_dimensions[column_letter].width = adjusted_width

                data_alignment = Alignment(horizontal='center', vertical='center')
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        cell.alignment = data_alignment

            wb.save(save_path)
        except Exception as e:
            print(f"Excel格式优化失败: {e}")

        print(f"回测报告已生成: {save_path}")

    def run_complete_analysis(self, file_path, output_folder, risk_manager=None):
        """
        完整分析流程
        """
        # 创建输出文件夹
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        # 加载数据
        self.load_data(file_path)

        # 预处理数据
        self.preprocess_data()

        # 运行回测（risk_manager为None时，不启用风险管理器）
        self.run_backtest(risk_manager)

        # 计算绩效指标
        self.calculate_metrics()

        # 绘制图表
        img_path = os.path.join(output_folder, f'{self.stock_code}_MACD_ATR_回测结果.png')
        self.plot_result(img_path)

        # 生成报告
        report_path = os.path.join(output_folder, f'{self.stock_code}_MACD_ATR_回测报告.xlsx')
        self.write_report(report_path)

        print(f"\n分析完成！结果保存在: {output_folder}")

    @staticmethod
    def compare_stocks(files_or_folder, init_params=None, risk_manager=None,
                       output_folder='compare_result', save_detail=True):
        """
        多股票对比

        Parameters:
        -----------
        files_or_folder : str or list
            文件夹路径或文件路径列表
        init_params : dict, optional
            策略初始化参数字典
        risk_manager : RiskManage or None, optional
            风险管理器实例，如果为None则不启用风险管理器（全仓买卖）
        output_folder : str
            输出文件夹路径
        save_detail : bool, default True
            是否保存每只股票的详细报告（图表和Excel）
            True: 每只股票生成独立的报告文件夹
            False: 只生成对比表格和对比图
        """
        if init_params is None:
            init_params = {}

        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        # 获取文件列表
        if isinstance(files_or_folder, str):
            if os.path.isdir(files_or_folder):
                files = [os.path.join(files_or_folder, f) for f in os.listdir(files_or_folder)
                         if f.endswith('.csv')]
            else:
                files = [files_or_folder]
        else:
            files = files_or_folder

        total_files = len(files)
        print(f"\n{'=' * 80}")
        print(f"多股票对比开始")
        print(f"共 {total_files} 只股票待处理")
        print(f"{'=' * 80}\n")

        results = []

        for idx, file_path in enumerate(files, 1):
            try:
                # 提取股票代码
                code_match = re.search(r'(\d+)', file_path)
                stock_code = code_match.group(1) if code_match else 'XXXX'

                print(f"\n[{idx}/{total_files}] 处理股票: {stock_code}")
                print(f"文件: {os.path.basename(file_path)}")

                # 显示风险管理器状态
                if risk_manager is None:
                    print(f"  风险管理器: 未启用 (全仓模式)")
                else:
                    risk_params = risk_manager.get_risk_metrics()
                    print(f"  风险管理器: 已启用")
                    print(f"    风险预算: {risk_params.get('risk_percent', 0.02) * 100:.1f}%")
                    print(f"    加仓比例: {risk_params.get('add_ratios', [0.4, 0.3, 0.3])}")
                    print(f"    加仓间隔: {risk_params.get('add_atr_multiple', 1)}倍ATR")
                    print(f"    止损倍数: {risk_params.get('stop_atr_multiple', 2)}倍ATR")

                # 运行策略
                strategy = MACDStrategy(**init_params)

                if save_detail:
                    print(f"  详细报告: 保存到 {os.path.join(output_folder, stock_code)}")
                    strategy.run_complete_analysis(file_path, os.path.join(output_folder, stock_code), risk_manager)
                else:
                    print(f"  详细报告: 不保存 (仅计算指标)")
                    strategy.load_data(file_path)
                    strategy.preprocess_data()
                    strategy.run_backtest(risk_manager)
                    strategy.calculate_metrics()

                # 收集结果
                s = strategy.metrics['strategy']
                bh = strategy.metrics['buy_hold']

                result = {
                    '股票代码': stock_code,
                    '策略总收益率': s['total_return'],
                    '策略年化收益率': s['annual_return'],
                    '策略最大回撤率': abs(s['max_drawdown']),
                    '策略夏普比率': s['sharpe_ratio'],
                    '策略胜率': strategy.metrics['win_rate'],
                    '策略盈亏比': strategy.metrics['profit_loss_ratio'],
                    '策略交易次数': strategy.metrics['trade_count'],
                    '买入持有总收益率': bh['total_return'],
                    '买入持有最大回撤率': abs(bh['max_drawdown']),
                    '买入持有夏普比率': bh['sharpe_ratio'],
                    '超额总收益率': s['total_return'] - bh['total_return'],
                    '回撤改善': abs(bh['max_drawdown']) - abs(s['max_drawdown']),
                    '平均ln(E)': strategy.metrics['avg_lnE'],
                }
                results.append(result)

                # 简单显示结果
                print(f"  完成: 策略收益率={result['策略总收益率'] * 100:.2f}%, "
                      f"超额收益={result['超额总收益率'] * 100:.2f}%")

            except Exception as e:
                print(f"  ✗ 处理失败: {e}")
                import traceback
                traceback.print_exc()

        if not results:
            print("\n没有成功处理任何股票")
            return

        # 显示处理汇总
        print(f"\n{'=' * 80}")
        print(f"处理完成汇总")
        print(f"{'=' * 80}")
        print(f"成功处理: {len(results)}/{total_files} 只股票")
        print(f"失败: {total_files - len(results)} 只股票")

        # 创建对比明细表
        df_results = pd.DataFrame(results)
        detail_path = os.path.join(output_folder, 'MACD_多股票对比明细表.xlsx')
        df_results.to_excel(detail_path, index=False)
        print(f"\n对比明细表已保存: {detail_path}")

        # 绘制多股票对比图
        print("\n正在生成对比图...")
        fig = plt.figure(figsize=(20, 16))

        # 图1：最大回撤分布对比直方图
        ax1 = plt.subplot(2, 3, 1)
        ax1.hist(df_results['策略最大回撤率'] * 100, bins=15, alpha=0.5, label='本策略', color='blue')
        ax1.hist(df_results['买入持有最大回撤率'] * 100, bins=15, alpha=0.5, label='买入持有', color='red')
        ax1.set_xlabel('最大回撤 (%)')
        ax1.set_ylabel('股票个数')
        ax1.set_title('最大回撤分布对比')
        ax1.legend()

        # 图2：本策略胜率 vs 盈亏比四象限散点图
        ax2 = plt.subplot(2, 3, 2)
        scatter = ax2.scatter(df_results['策略胜率'] * 100, df_results['策略盈亏比'],
                              c=df_results['策略夏普比率'], cmap='viridis', s=100, alpha=0.7)
        ax2.axhline(y=1, color='gray', linestyle='--', alpha=0.5)
        ax2.axvline(x=50, color='gray', linestyle='--', alpha=0.5)
        ax2.set_xlabel('胜率 (%)')
        ax2.set_ylabel('盈亏比')
        ax2.set_title('胜率 vs 盈亏比 (颜色=夏普比率)')
        plt.colorbar(scatter, ax=ax2, label='夏普比率')

        # 图3：本策略收益率 vs 买入持有策略收益率散点图
        ax3 = plt.subplot(2, 3, 3)
        scatter = ax3.scatter(df_results['买入持有总收益率'] * 100, df_results['策略总收益率'] * 100,
                              c=df_results['超额总收益率'] * 100, cmap='RdYlGn', s=100, alpha=0.7)
        max_val = max(df_results['买入持有总收益率'].max(), df_results['策略总收益率'].max()) * 100
        min_val = min(df_results['买入持有总收益率'].min(), df_results['策略总收益率'].min()) * 100
        ax3.plot([min_val, max_val], [min_val, max_val], 'k--', alpha=0.5, label='45°线')
        ax3.set_xlabel('买入持有收益率 (%)')
        ax3.set_ylabel('本策略收益率 (%)')
        ax3.set_title('策略收益率 vs 买入持有收益率')
        plt.colorbar(scatter, ax=ax3, label='超额收益 (%)')

        # 图4：本策略超额收益 vs 回撤改善四象限散点图
        ax4 = plt.subplot(2, 3, 4)
        scatter = ax4.scatter(df_results['超额总收益率'] * 100, df_results['回撤改善'] * 100,
                              c=df_results['策略夏普比率'], cmap='viridis', s=100, alpha=0.7)
        ax4.axhline(y=0, color='gray', linestyle='--', alpha=0.5)
        ax4.axvline(x=0, color='gray', linestyle='--', alpha=0.5)
        ax4.set_xlabel('超额收益率 (%)')
        ax4.set_ylabel('回撤改善 (%)')
        ax4.set_title('超额收益 vs 回撤改善 (颜色=夏普比率)')
        plt.colorbar(scatter, ax=ax4, label='夏普比率')

        # 图5：平均ln(E)直方图
        ax5 = plt.subplot(2, 3, 5)
        ax5.hist(df_results['平均ln(E)'], bins=15, alpha=0.7, color='green', edgecolor='black')
        ax5.axvline(x=df_results['平均ln(E)'].mean(), color='red', linestyle='--',
                    label=f'均值: {df_results["平均ln(E)"].mean():.2f}')
        ax5.axvline(x=df_results['平均ln(E)'].median(), color='orange', linestyle='--',
                    label=f'中位数: {df_results["平均ln(E)"].median():.2f}')
        ax5.set_xlabel('平均ln(E)')
        ax5.set_ylabel('股票个数')
        ax5.set_title('平均ln(E)分布')
        ax5.legend()

        plt.tight_layout()

        # 保存对比图
        img_path = os.path.join(output_folder, 'MACD_多股票对比图.png')
        plt.savefig(img_path, dpi=150, bbox_inches='tight')
        plt.show()
        print(f"对比图已保存: {img_path}")

        # 打印总体效果
        print("\n" + "=" * 80)
        print("策略总体效果")
        print("=" * 80)

        beat_benchmark = (df_results['超额总收益率'] > 0).sum()
        print(f"跑赢基准比例: {beat_benchmark}/{len(df_results)} = {beat_benchmark / len(df_results) * 100:.2f}%")

        drawdown_improve = (df_results['回撤改善'] > 0).sum()
        print(f"回撤改善比例: {drawdown_improve}/{len(df_results)} = {drawdown_improve / len(df_results) * 100:.2f}%")

        print(f"策略平均回撤: {df_results['策略最大回撤率'].mean() * 100:.2f}%")
        print(f"买入持有平均回撤: {df_results['买入持有最大回撤率'].mean() * 100:.2f}%")

        print(f"策略平均夏普: {df_results['策略夏普比率'].mean():.2f}")
        print(f"买入持有平均夏普: {df_results['买入持有夏普比率'].mean():.2f}")

        print(f"平均ln(E):")
        print(f"  中位数: {df_results['平均ln(E)'].median():.2f}")
        print(f"  平均数: {df_results['平均ln(E)'].mean():.2f}")
        print(f"  最大值: {df_results['平均ln(E)'].max():.2f}")
        print(f"  最小值: {df_results['平均ln(E)'].min():.2f}")

        print(f"平均总收益率: {df_results['策略总收益率'].mean() * 100:.2f}%")
        print(f"平均买入次数: {df_results['策略交易次数'].mean():.2f}")
        print(f"平均胜率: {df_results['策略胜率'].mean() * 100:.2f}%")
        print(f"平均盈亏比: {df_results['策略盈亏比'].mean():.2f}")

        return df_results

    @staticmethod
    def compare_strategies(file_path, param_combinations, output_folder):
        """
        比较同一只股票，使用不同参数组合的MACD策略效果
        """
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        # 提取股票代码
        code_match = re.search(r'(\d+)', file_path)
        stock_code = code_match.group(1) if code_match else 'XXXX'

        results = []
        nav_series = {}

        # 加载基础数据（用于获取日期）
        base_strategy = MACDStrategy()
        base_strategy.load_data(file_path)
        dates = base_strategy.data['日期']  # 获取日期序列

        # 计算买入持有净值（使用日期索引）
        buy_hold = BuyHoldStrategy(1000000, 0.0001, 0.001, 5)
        buy_hold.run_backtest(base_strategy.data)
        bh_nav = buy_hold.net_values

        # 确保买入持有净值使用日期索引
        if len(bh_nav) == len(dates):
            bh_nav.index = dates

        for i, params in enumerate(param_combinations):
            print(f"测试参数组合 {i + 1}: {params}")

            try:
                strategy = MACDStrategy(
                    initial_capital=1000000,
                    macd_params=[params.get('fast', 12), params.get('slow', 26), params.get('signal', 9)],
                    atr_period=params.get('atr_period', 14)
                )
                strategy.load_data(file_path)
                strategy.preprocess_data()

                risk_manager = RiskManage(
                    risk_percent=params.get('risk_percent', 0.02),
                    add_ratios=params.get('add_ratios', [0.4, 0.3, 0.3]),
                    add_atr_multiple=params.get('add_atr_multiple', 1),
                    stop_atr_multiple=params.get('stop_atr_multiple', 2)
                )
                strategy.run_backtest(risk_manager)
                strategy.calculate_metrics()

                s = strategy.metrics['strategy']
                bh = strategy.metrics['buy_hold']

                param_name = f"fast{params.get('fast', 12)}_slow{params.get('slow', 26)}_sig{params.get('signal', 9)}"
                if 'risk_percent' in params:
                    param_name += f"_risk{params.get('risk_percent')}"
                if 'add_atr_multiple' in params:
                    param_name += f"_add{params.get('add_atr_multiple')}"

                # 获取净值序列并设置日期索引
                nav = strategy.data['portfolio_value'].values
                nav_series[param_name] = pd.Series(nav, index=dates)

                result = {
                    '参数组合': param_name,
                    '总收益率': s['total_return'],
                    '年化收益率': s['annual_return'],
                    '最大回撤率': abs(s['max_drawdown']),
                    '夏普比率': s['sharpe_ratio'],
                    '胜率': strategy.metrics['win_rate'],
                    '盈亏比': strategy.metrics['profit_loss_ratio'],
                    '交易次数': strategy.metrics['trade_count'],
                    '超额收益': s['total_return'] - bh['total_return'],
                    '回撤改善': abs(s['max_drawdown']) - abs(bh['max_drawdown'])
                }
                results.append(result)
            except Exception as e:
                print(f"参数组合 {params} 测试失败: {e}")

        # 创建详细数据表
        df_results = pd.DataFrame(results)

        # 添加买入持有策略行
        bh_result = {
            '参数组合': '买入持有策略',
            '总收益率': bh['total_return'],
            '年化收益率': bh['annual_return'],
            '最大回撤率': bh['max_drawdown'],
            '夏普比率': bh['sharpe_ratio'],
            '胜率': '-',
            '盈亏比': '-',
            '交易次数': '-',
            '超额收益': '-',
            '回撤改善': '-'
        }
        df_results = pd.concat([df_results, pd.DataFrame([bh_result])], ignore_index=True)

        # 保存Excel
        excel_path = os.path.join(output_folder, f'{stock_code}_MACD参数对比详细数据.xlsx')
        df_results.to_excel(excel_path, index=False)
        print(f"参数对比数据已保存: {excel_path}")

        # 绘制净值曲线
        plt.figure(figsize=(14, 8))

        # 绘制各策略净值曲线（使用统一的日期索引）
        for name, nav in nav_series.items():
            plt.semilogy(nav.index, nav.values, label=name, linewidth=1.5, alpha=0.8)

        # 绘制买入持有策略净值曲线
        plt.semilogy(bh_nav.index, bh_nav.values, label='买入持有策略', color='black', linewidth=2, linestyle='--')

        plt.xlabel('日期')
        plt.ylabel('净值（对数坐标）')
        plt.title(f'股票 {stock_code} 不同MACD参数对比')
        plt.legend(loc='upper left', fontsize=10)
        plt.grid(True, alpha=0.3)
        plt.tight_layout()

        img_path = os.path.join(output_folder, f'{stock_code}_MACD参数对比净值曲线.png')
        plt.savefig(img_path, dpi=150, bbox_inches='tight')
        plt.show()

        print(f"净值曲线已保存: {img_path}")

        return df_results

class StrategyCompare:
    """
    策略对比分析类

    用于比较多个量化策略的回测结果，生成对比报告和可视化图表。
    """

    def __init__(self, input_path: Union[str, List[str]], output_dir: str):
        """
        初始化策略对比类

        Parameters:
        -----------
        input_path : str or List[str]
            策略分析结果文件路径列表，或者包含所有文件的文件夹路径
        output_dir : str
            输出结果保存文件夹路径
        """
        self.input_path = input_path
        self.output_dir = output_dir
        self.strategy_data = {}
        self.summary_stats = {}
        self.load_errors = []

        os.makedirs(output_dir, exist_ok=True)
        self._load_data()

    def _extract_strategy_name(self, filename: str) -> str:
        """从文件名中获取“_”之前的部分作为策略名称"""
        name_without_ext = Path(filename).stem
        if '_' in name_without_ext:
            return name_without_ext.split('_')[0]
        return name_without_ext

    def _safe_mean(self, series: pd.Series) -> float:
        """安全地计算平均值，处理inf和异常值"""
        if series.dtype not in ['float64', 'int64']:
            try:
                series = pd.to_numeric(series, errors='coerce')
            except:
                return np.nan

        clean_series = series.replace([np.inf, -np.inf], np.nan)
        valid_data = clean_series[np.isfinite(clean_series)]

        if len(valid_data) == 0:
            return np.nan
        return valid_data.mean()

    def _convert_to_numeric(self, df: pd.DataFrame) -> pd.DataFrame:
        """将DataFrame中的数值列从字符串转换为数值类型"""
        numeric_fields = [
            '策略总收益率', '策略年化收益率', '策略最大回撤率', '策略夏普比率',
            '策略胜率', '策略盈亏比', '策略交易次数', '买入持有总收益率',
            '买入持有最大回撤率', '买入持有夏普比率', '超额总收益率', '回撤改善'
        ]

        df_converted = df.copy()

        for field in numeric_fields:
            if field in df_converted.columns:
                if df_converted[field].dtype == 'object' or df_converted[field].dtype == 'string':
                    try:
                        if df_converted[field].astype(str).str.contains('%').any():
                            df_converted[field] = df_converted[field].astype(str).str.replace('%', '')
                            df_converted[field] = pd.to_numeric(df_converted[field], errors='coerce') / 100
                        else:
                            df_converted[field] = pd.to_numeric(df_converted[field], errors='coerce')
                    except:
                        df_converted[field] = np.nan
                else:
                    df_converted[field] = pd.to_numeric(df_converted[field], errors='coerce')

                if df_converted[field].dtype in ['float64', 'int64']:
                    df_converted[field] = df_converted[field].replace([np.inf, -np.inf], np.nan)

        return df_converted

    def _load_data(self):
        """加载策略数据"""
        file_paths = []

        if isinstance(self.input_path, str):
            if os.path.isdir(self.input_path):
                for file in os.listdir(self.input_path):
                    if file.endswith(('.xlsx', '.xls')):
                        file_paths.append(os.path.join(self.input_path, file))
            else:
                raise ValueError(f"输入的路径不是文件夹: {self.input_path}")
        elif isinstance(self.input_path, list):
            file_paths = [f for f in self.input_path if f.endswith(('.xlsx', '.xls'))]
        else:
            raise TypeError("input_path必须是文件夹路径或文件路径列表")

        if not file_paths:
            raise ValueError("没有找到Excel文件")

        print(f"找到 {len(file_paths)} 个策略文件")

        loaded_count = 0
        for file_path in file_paths:
            try:
                strategy_name = self._extract_strategy_name(os.path.basename(file_path))
                print(f"\n正在加载策略: {strategy_name}")

                df = pd.read_excel(file_path)

                required_fields = ['策略总收益率', '买入持有总收益率']
                missing_critical = [f for f in required_fields if f not in df.columns]
                if missing_critical:
                    self.load_errors.append(f"{strategy_name}: 缺少关键字段 {missing_critical}")
                    continue

                df = self._convert_to_numeric(df)

                if df['策略总收益率'].isnull().all():
                    self.load_errors.append(f"{strategy_name}: 策略总收益率全部为NaN")
                    continue

                self.strategy_data[strategy_name] = df
                loaded_count += 1
                print(f"  ✓ 成功加载策略: {strategy_name}")

            except Exception as e:
                self.load_errors.append(f"{os.path.basename(file_path)}: 加载失败 - {str(e)}")

        if loaded_count == 0:
            raise ValueError("没有成功加载任何策略数据")

        print(f"\n总共成功加载 {loaded_count} 个策略")
        if self.load_errors:
            print(f"\n加载过程中出现 {len(self.load_errors)} 个错误:")
            for error in self.load_errors:
                print(f"  - {error}")

    def _calculate_summary_stats(self):
        """计算每个策略的汇总统计指标"""
        buy_hold_returns = []

        for strategy_name, df in self.strategy_data.items():
            print(f"\n计算策略 {strategy_name} 的统计指标:")

            try:
                valid_mask = df['策略总收益率'].notna() & np.isfinite(df['策略总收益率'])
                valid_count = valid_mask.sum()
                print(f"  有效样本数: {valid_count}/{len(df)}")

                if valid_count == 0:
                    continue

                if '买入持有总收益率' in df.columns:
                    both_valid = valid_mask & df['买入持有总收益率'].notna() & np.isfinite(df['买入持有总收益率'])
                    win_benchmark = (df.loc[both_valid, '策略总收益率'] > df.loc[
                        both_valid, '买入持有总收益率']).sum() / both_valid.sum() if both_valid.sum() > 0 else np.nan
                else:
                    win_benchmark = np.nan

                if '策略最大回撤率' in df.columns and '买入持有最大回撤率' in df.columns:
                    drawdown_mask = (df['策略最大回撤率'].notna() & np.isfinite(df['策略最大回撤率']) &
                                     df['买入持有最大回撤率'].notna() & np.isfinite(df['买入持有最大回撤率']))
                    drawdown_improve = (df.loc[drawdown_mask, '策略最大回撤率'] < df.loc[
                        drawdown_mask, '买入持有最大回撤率']).sum() / drawdown_mask.sum() if drawdown_mask.sum() > 0 else np.nan
                else:
                    drawdown_improve = np.nan

                avg_return = self._safe_mean(df['策略总收益率'])
                avg_win_rate = self._safe_mean(df['策略胜率']) * 100 if '策略胜率' in df.columns else np.nan
                avg_win_loss = self._safe_mean(df['策略盈亏比']) if '策略盈亏比' in df.columns else np.nan
                avg_trades = self._safe_mean(df['策略交易次数']) if '策略交易次数' in df.columns else np.nan

                print(f"  跑赢基准比率: {win_benchmark:.4f}" if not np.isnan(win_benchmark) else "  跑赢基准比率: NaN")
                print(f"  平均收益率: {avg_return:.6f}" if not np.isnan(avg_return) else "  平均收益率: NaN")
                print(f"  平均胜率: {avg_win_rate:.4f}%" if not np.isnan(avg_win_rate) else "  平均胜率: NaN")
                print(f"  平均盈亏比: {avg_win_loss:.4f}" if not np.isnan(avg_win_loss) else "  平均盈亏比: NaN")

                self.summary_stats[strategy_name] = {
                    '策略名称': strategy_name,
                    '跑赢基准比率': win_benchmark,
                    '回撤改善比率': drawdown_improve,
                    '平均收益率': avg_return,
                    '平均胜率': avg_win_rate,
                    '平均盈亏比': avg_win_loss,
                    '平均交易次数': avg_trades
                }

                if '买入持有总收益率' in df.columns:
                    clean_bh = df['买入持有总收益率'].replace([np.inf, -np.inf], np.nan).dropna()
                    buy_hold_returns.extend(clean_bh.tolist())

            except Exception as e:
                print(f"  计算策略 {strategy_name} 时出错: {e}")
                continue

        avg_buy_hold_return = np.mean(buy_hold_returns) if buy_hold_returns else 0

        self.summary_stats['买入持有'] = {
            '策略名称': '买入持有',
            '跑赢基准比率': np.nan,
            '回撤改善比率': np.nan,
            '平均收益率': avg_buy_hold_return,
            '平均胜率': np.nan,
            '平均盈亏比': np.nan,
            '平均交易次数': 1.0
        }

        print(f"\n买入持有平均收益率: {avg_buy_hold_return:.6f}")

    def generate_excel_report(self) -> str:
        """生成多策略对比分析报告Excel文件"""
        print("\n生成Excel报告...")
        self._calculate_summary_stats()

        df_summary = pd.DataFrame(list(self.summary_stats.values()))
        df_to_save = df_summary.copy()

        df_display = df_summary.copy()
        df_display['跑赢基准比率'] = df_display['跑赢基准比率'].apply(
            lambda x: f"{x:.2%}" if pd.notna(x) and np.isfinite(x) else "-")
        df_display['回撤改善比率'] = df_display['回撤改善比率'].apply(
            lambda x: f"{x:.2%}" if pd.notna(x) and np.isfinite(x) else "-")
        df_display['平均收益率'] = df_display['平均收益率'].apply(
            lambda x: f"{x:.2%}" if pd.notna(x) and np.isfinite(x) else "-")
        df_display['平均胜率'] = df_display['平均胜率'].apply(
            lambda x: f"{x:.2f}%" if pd.notna(x) and np.isfinite(x) else "-")
        df_display['平均盈亏比'] = df_display['平均盈亏比'].apply(
            lambda x: f"{x:.2f}" if pd.notna(x) and np.isfinite(x) else "-")
        df_display['平均交易次数'] = df_display['平均交易次数'].apply(
            lambda x: f"{x:.0f}" if pd.notna(x) and np.isfinite(x) else "-")

        output_path = os.path.join(self.output_dir, '多策略对比分析报告.xlsx')

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df_display.to_excel(writer, sheet_name='策略对比报告', index=False)
            df_to_save.to_excel(writer, sheet_name='原始数据', index=False)

        print(f"✓ Excel报告已生成: {output_path}")
        return output_path

    def plot_scatter(self, strategy1: str, strategy2: str) -> str:
        """生成两个策略的收益率对比散点图"""
        if strategy1 not in self.strategy_data or strategy2 not in self.strategy_data:
            print(f"策略名称不存在，可用策略: {list(self.strategy_data.keys())}")
            return None

        df1 = self.strategy_data[strategy1]
        df2 = self.strategy_data[strategy2]

        if '股票代码' not in df1.columns or '股票代码' not in df2.columns:
            return None

        merged = pd.merge(
            df1[['股票代码', '策略总收益率']],
            df2[['股票代码', '策略总收益率']],
            on='股票代码',
            suffixes=(f'_{strategy1}', f'_{strategy2}')
        )

        if len(merged) == 0:
            return None

        plt.figure(figsize=(10, 8))

        x = merged[f'策略总收益率_{strategy1}']
        y = merged[f'策略总收益率_{strategy2}']

        valid_mask = ~(np.isnan(x) | np.isnan(y) | ~np.isfinite(x) | ~np.isfinite(y))
        x = x[valid_mask]
        y = y[valid_mask]

        if len(x) == 0:
            return None

        plt.scatter(x, y, alpha=0.6, edgecolors='white', linewidth=0.5)

        min_val = min(x.min(), y.min())
        max_val = max(x.max(), y.max())
        plt.plot([min_val, max_val], [min_val, max_val], 'r--', alpha=0.5, label='y=x')

        try:
            z = np.polyfit(x, y, 1)
            p = np.poly1d(z)
            plt.plot([min_val, max_val], [p(min_val), p(max_val)], 'b-', alpha=0.7,
                     label=f'趋势线 (斜率={z[0]:.3f})')
        except:
            pass

        plt.xlabel(f'{strategy1} 收益率')
        plt.ylabel(f'{strategy2} 收益率')
        plt.title(f'{strategy1} vs {strategy2} 收益率对比')
        plt.grid(True, alpha=0.3)
        plt.legend()

        corr = np.corrcoef(x, y)[0, 1]
        plt.text(0.05, 0.95, f'相关系数: {corr:.4f}\n样本数: {len(x)}',
                 transform=plt.gca().transAxes,
                 bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.5))

        plt.tight_layout()

        output_path = os.path.join(self.output_dir, f'scatter_{strategy1}_vs_{strategy2}.png')
        plt.savefig(output_path, dpi=300, bbox_inches='tight')
        plt.close()

        return output_path

    def plot_all_scatter(self):
        """为所有策略两两生成散点图"""
        strategy_names = list(self.strategy_data.keys())

        if len(strategy_names) < 2:
            return

        print("\n生成两两对比散点图...")
        for i in range(len(strategy_names)):
            for j in range(i + 1, len(strategy_names)):
                try:
                    self.plot_scatter(strategy_names[i], strategy_names[j])
                except Exception as e:
                    print(f"生成 {strategy_names[i]} vs {strategy_names[j]} 散点图失败: {e}")

    def plot_histograms(self) -> Dict[str, str]:
        """生成三个直方图：总收益率对比、最大回撤率对比、胜率对比"""
        print("\n生成分布直方图...")

        all_returns = []
        all_drawdowns = []
        all_win_rates = []
        strategy_labels = []

        colors = plt.cm.tab10(np.linspace(0, 1, len(self.strategy_data)))

        for i, (strategy_name, df) in enumerate(self.strategy_data.items()):
            returns = df['策略总收益率'].replace([np.inf, -np.inf],
                                                 np.nan).dropna() if '策略总收益率' in df.columns else pd.Series([])
            drawdowns = df['策略最大回撤率'].replace([np.inf, -np.inf],
                                                     np.nan).dropna() if '策略最大回撤率' in df.columns else pd.Series(
                [])
            win_rates = df['策略胜率'].replace([np.inf, -np.inf],
                                               np.nan).dropna() * 100 if '策略胜率' in df.columns else pd.Series([])

            if len(returns) > 0:
                all_returns.append(returns)
                all_drawdowns.append(drawdowns)
                all_win_rates.append(win_rates)
                strategy_labels.append(strategy_name)
                print(f"  {strategy_name}: 有效收益率样本数={len(returns)}")

        if not all_returns:
            return {}

        plt.figure(figsize=(14, 8))
        for i, (returns, label) in enumerate(zip(all_returns, strategy_labels)):
            plt.hist(returns, bins=30, alpha=0.5, label=label,
                     color=colors[i % len(colors)], edgecolor='black', linewidth=0.5)

        plt.xlabel('总收益率', fontsize=12)
        plt.ylabel('样本数', fontsize=12)
        plt.title('各策略总收益率分布对比', fontsize=14)
        plt.legend(loc='upper right')
        plt.grid(True, alpha=0.3)
        plt.tight_layout()
        returns_hist_path = os.path.join(self.output_dir, 'histogram_returns.png')
        plt.savefig(returns_hist_path, dpi=300, bbox_inches='tight')
        plt.close()

        plt.figure(figsize=(14, 8))
        for i, (drawdowns, label) in enumerate(zip(all_drawdowns, strategy_labels)):
            if len(drawdowns) > 0:
                plt.hist(drawdowns, bins=30, alpha=0.5, label=label,
                         color=colors[i % len(colors)], edgecolor='black', linewidth=0.5)

        plt.xlabel('最大回撤率', fontsize=12)
        plt.ylabel('样本数', fontsize=12)
        plt.title('各策略最大回撤率分布对比', fontsize=14)
        plt.legend(loc='upper right')
        plt.grid(True, alpha=0.3)
        plt.tight_layout()
        drawdowns_hist_path = os.path.join(self.output_dir, 'histogram_drawdowns.png')
        plt.savefig(drawdowns_hist_path, dpi=300, bbox_inches='tight')
        plt.close()

        plt.figure(figsize=(14, 8))
        for i, (win_rates, label) in enumerate(zip(all_win_rates, strategy_labels)):
            if len(win_rates) > 0:
                plt.hist(win_rates, bins=30, alpha=0.5, label=label,
                         color=colors[i % len(colors)], edgecolor='black', linewidth=0.5)

        plt.xlabel('胜率 (%)', fontsize=12)
        plt.ylabel('样本数', fontsize=12)
        plt.title('各策略胜率分布对比', fontsize=14)
        plt.legend(loc='upper right')
        plt.grid(True, alpha=0.3)
        plt.tight_layout()
        win_rates_hist_path = os.path.join(self.output_dir, 'histogram_win_rates.png')
        plt.savefig(win_rates_hist_path, dpi=300, bbox_inches='tight')
        plt.close()

        return {
            'returns_hist': returns_hist_path,
            'drawdowns_hist': drawdowns_hist_path,
            'win_rates_hist': win_rates_hist_path
        }

    def run_full_analysis(self) -> Dict:
        """运行完整的策略对比分析"""
        print("=" * 50)
        print("开始策略对比分析")
        print("=" * 50)

        print(f"\n当前成功加载的策略: {list(self.strategy_data.keys())}")

        results = {}
        results['excel_report'] = self.generate_excel_report()
        self.plot_all_scatter()
        results.update(self.plot_histograms())

        print("=" * 50)
        print("策略对比分析完成!")
        print(f"所有结果已保存至: {self.output_dir}")
        print("=" * 50)

        return results